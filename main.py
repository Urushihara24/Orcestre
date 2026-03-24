import asyncio
import csv
import json
import logging
import math
import os
import random
import re
import threading
import time
import types as py_types
from datetime import datetime, timedelta, date, timezone, time as dt_time
from enum import Enum
from functools import wraps
from typing import Any, List, Optional
from zoneinfo import ZoneInfo

import telebot
from apscheduler.schedulers.background import BackgroundScheduler
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import and_, case, func, or_
from telebot import types
from telebot.apihelper import ApiTelegramException

# Загружаем .env до импортов модулей, которые читают os.getenv на import-time.
load_dotenv()

from epic_api_client import (
    send_friend_request_with_device,
    check_friend_status_with_device,
    cancel_friend_request_with_device,
    remove_friend_with_device,
    verify_account_health_with_device,
    change_display_name_with_device,
)
from epic_device_auth import EpicDeviceAuthGenerator, append_device_auth_to_file
from campaign_settings import (
    campaign_ui_label,
    campaign_ui_num,
    get_campaign_send_mode,
    get_campaign_sender_pick_mode,
    set_campaign_send_mode,
    set_campaign_sender_pick_mode,
    target_required_senders as campaign_target_required_senders,
)
from db_models import (
    Account,
    Campaign,
    DB_URL,
    LogEvent,
    NicknameChangeTask,
    Proxy,
    SessionLocal,
    Setting,
    Target,
    Task,
)
from recheck_manager import RecheckPair, RecheckQueuePlanner

# ============================================================
# КОНФИГУРАЦИЯ
# ============================================================

BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "").strip()

# Admin configuration:
# - Backward compatible: ADMIN_TELEGRAM_ID (single id)
# - Preferred: ADMIN_TELEGRAM_IDS="id1,id2,..." (comma/space separated)
def _parse_admin_ids() -> set[int]:
    raw = os.getenv("ADMIN_TELEGRAM_IDS", "").strip()
    ids: set[int] = set()
    if raw:
        for part in raw.replace(";", ",").replace(" ", ",").split(","):
            part = part.strip()
            if not part:
                continue
            try:
                ids.add(int(part))
            except Exception:
                continue
    one = os.getenv("ADMIN_TELEGRAM_ID", "").strip()
    if one:
        try:
            ids.add(int(one))
        except Exception:
            one = ""
    return ids


ADMIN_IDS = _parse_admin_ids()
ADMIN_ID = min(ADMIN_IDS) if ADMIN_IDS else 0

MAX_TASKS_PER_TICK = int(os.getenv("MAX_TASKS_PER_TICK", "10"))
PROCESS_TICK_SECONDS = int(os.getenv("PROCESS_TICK_SECONDS", "20"))
DAILY_RESET_HOUR_UTC = int(os.getenv("DAILY_RESET_HOUR_UTC", "0"))
WORKER_INSTANCE = os.getenv("WORKER_INSTANCE", "worker-1").strip() or "worker-1"

DEFAULT_DAILY_LIMIT = int(os.getenv("DEFAULT_DAILY_LIMIT", "10"))
DEFAULT_SEND_JITTER_MIN_SEC = int(os.getenv("DEFAULT_SEND_JITTER_MIN_SEC", "60"))
DEFAULT_SEND_JITTER_MAX_SEC = int(os.getenv("DEFAULT_SEND_JITTER_MAX_SEC", "600"))
DEFAULT_TIMEZONE = "Europe/Moscow"
DEFAULT_TARGET_SENDERS_COUNT = int(os.getenv("DEFAULT_TARGET_SENDERS_COUNT", "1"))
DEFAULT_MIN_REQUEST_INTERVAL_SEC = int(os.getenv("DEFAULT_MIN_REQUEST_INTERVAL_SEC", "30"))
DEFAULT_MAX_REQUEST_INTERVAL_SEC = int(os.getenv("DEFAULT_MAX_REQUEST_INTERVAL_SEC", "40"))
DEFAULT_HOURLY_API_LIMIT = int(os.getenv("DEFAULT_HOURLY_API_LIMIT", "40"))
DEFAULT_DAILY_API_LIMIT = int(os.getenv("DEFAULT_DAILY_API_LIMIT", "500"))
DEFAULT_SEND_API_COST = int(os.getenv("DEFAULT_SEND_API_COST", "3"))
DEFAULT_CHECK_API_COST = int(os.getenv("DEFAULT_CHECK_API_COST", "2"))
DEFAULT_RECHECK_DAILY_LIMIT = int(os.getenv("DEFAULT_RECHECK_DAILY_LIMIT", "500"))
DEFAULT_DAILY_REPEAT_CAMPAIGN_ENABLED = os.getenv("DAILY_REPEAT_CAMPAIGN_ENABLED", "0").strip().lower() in {"1", "true", "yes"}

DRY_RUN = os.getenv("DRY_RUN", "0").strip().lower() in {"1", "true", "yes"}
DRY_RUN_ACCEPT_RATE = float(os.getenv("DRY_RUN_ACCEPT_RATE", "0.3"))
DRY_RUN_CHECK_DELAY_SEC = int(os.getenv("DRY_RUN_CHECK_DELAY_SEC", "60"))
# Safety gate for real Epic friend-request POSTs.
# Keep disabled by default; enable explicitly only when you're ready.
SEND_REQUESTS_ENABLED = os.getenv("SEND_REQUESTS_ENABLED", "0").strip().lower() in {"1", "true", "yes"}

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
logging.getLogger("apscheduler").setLevel(logging.WARNING)
LOG_RETENTION_DAYS = int(os.getenv("LOG_RETENTION_DAYS", "30"))
LIST_PAGE_SIZE = int(os.getenv("LIST_PAGE_SIZE", "25"))
TARGETS_PAGE_SIZE = int(os.getenv("TARGETS_PAGE_SIZE", "10"))
SENDERS_PAGE_SIZE = int(os.getenv("SENDERS_PAGE_SIZE", "15"))

if not BOT_TOKEN:
    raise SystemExit("❌ TELEGRAM_BOT_TOKEN не задан в .env")

if ADMIN_ID == 0:
    raise SystemExit("❌ ADMIN_TELEGRAM_ID/ADMIN_TELEGRAM_IDS не задан(ы) или пустые в .env")


def is_admin(user_id: int) -> bool:
    try:
        return int(user_id) in ADMIN_IDS
    except Exception:
        return False


# ============================================================
# ПЕРЕЧИСЛЕНИЯ СТАТУСОВ
# ============================================================

class AccountStatus(Enum):
    ACTIVE = "active"
    BANNED = "banned"
    LOCKED = "locked"
    PAUSED = "paused"
    MANUAL = "manual"


class TargetStatus(Enum):
    NEW = "new"
    PENDING = "pending"
    SENT = "sent"
    ACCEPTED = "accepted"
    REJECTED = "rejected"
    FAILED = "failed"


class TaskStatus(Enum):
    QUEUED = "queued"
    RUNNING = "running"
    DONE = "done"
    FAILED = "failed"
    POSTPONED = "postponed"
    CANCELLED = "cancelled"


class NicknameChangeStatus(Enum):
    QUEUED = "queued"
    RUNNING = "running"
    DONE = "done"
    FAILED = "failed"
    POSTPONED = "postponed"
    SKIPPED = "skipped"


ACTIVE_SEND_TASK_STATUSES = (
    TaskStatus.QUEUED.value,
    TaskStatus.POSTPONED.value,
    TaskStatus.RUNNING.value,
)
QUEUED_OR_POSTPONED_TASK_STATUSES = (
    TaskStatus.QUEUED.value,
    TaskStatus.POSTPONED.value,
)


TARGET_STATUS_RU = {
    TargetStatus.NEW.value: "новая",
    TargetStatus.PENDING.value: "в ожидании",
    TargetStatus.SENT.value: "отправлено",
    TargetStatus.ACCEPTED.value: "принято",
    TargetStatus.REJECTED.value: "отклонено",
    TargetStatus.FAILED.value: "ошибка",
}

TARGET_SEND_BASE_STATUSES = (
    TargetStatus.NEW.value,
    TargetStatus.PENDING.value,
    TargetStatus.SENT.value,
)
TARGET_SEND_REPEAT_EXTRA_STATUSES = (
    TargetStatus.REJECTED.value,
    TargetStatus.FAILED.value,
    TargetStatus.ACCEPTED.value,
)
TARGET_RECHECK_ELIGIBLE_STATUSES = (
    TargetStatus.PENDING.value,
    TargetStatus.SENT.value,
    TargetStatus.ACCEPTED.value,
)

NICKNAME_ALLOWED_RE = re.compile(r"^[A-Za-z0-9_]{3,16}$")
NICKNAME_FALLBACK_LIMIT = 12
NICKNAME_VISUAL_SUBS: dict[str, tuple[str, ...]] = {
    "a": ("4",),
    "b": ("8",),
    "e": ("3",),
    "g": ("9",),
    "i": ("1",),
    "l": ("1",),
    "o": ("0",),
    "q": ("9",),
    "s": ("5",),
    "t": ("7",),
    "z": ("2",),
    "0": ("o",),
    "1": ("i", "l"),
    "2": ("z",),
    "3": ("e",),
    "4": ("a",),
    "5": ("s",),
    "7": ("t",),
    "8": ("b",),
    "9": ("g", "q"),
}

PRECHECK_SKIP_REASONS = (
    "precheck_accepted_skip",
    "precheck_pending_skip",
    "precheck_accepted_skip_requeued",
    "precheck_pending_skip_requeued",
    "idempotent_request_skip",
    "idempotent_request_skip_requeued",
)


def target_status_ru(code: str) -> str:
    return TARGET_STATUS_RU.get(code, code or "неизвестно")


# ============================================================
# DATABASE
# ============================================================

DB_LOCK = threading.Lock()


# ============================================================
# HELPERS
# ============================================================

def utc_now() -> datetime:
    """UTC datetime без tzinfo (совместимо с текущими naive DateTime колонками)."""
    return datetime.now(timezone.utc).replace(tzinfo=None)


def utc_today() -> date:
    return utc_now().date()

def next_daily_reset_utc(now_utc: datetime) -> datetime:
    """
    Вернуть ближайший момент сброса дневных лимитов (UTC, naive).
    now_utc: naive datetime в UTC.
    """
    reset_today = datetime.combine(now_utc.date(), dt_time(hour=DAILY_RESET_HOUR_UTC, minute=0, second=0))
    if now_utc < reset_today:
        return reset_today
    return reset_today + timedelta(days=1)


def db_exec(fn):
    """Выполнить функцию в потокобезопасном контексте БД"""
    with DB_LOCK:
        db = SessionLocal()
        try:
            return fn(db)
        finally:
            db.close()


def log_event(level: str, message: str):
    """Логировать событие через отдельную сессию"""

    def _inner(db):
        db.add(LogEvent(level=level, message=message[:4000]))
        db.commit()

    db_exec(_inner)
    logger.info(f"[{level.upper()}] {message}")


def log_event_in_db(db, level: str, message: str):
    """Логировать событие, используя уже открытую сессию (без db_exec)."""
    db.add(LogEvent(level=level, message=message[:4000]))
    db.commit()
    logger.info(f"[{level.upper()}] {message}")

def log_event_in_tx(db, level: str, message: str):
    """Лог в рамках текущей транзакции (без commit)."""
    db.add(LogEvent(level=level, message=message[:4000]))
    logger.info(f"[{level.upper()}] {message}")

def safe_split_lines(text: str) -> List[str]:
    return [ln.strip() for ln in (text or "").splitlines() if ln.strip()]


def split_multi_values(text: str) -> List[str]:
    """
    Разобрать список значений из текста:
    - по строкам
    - по ',' или ';' внутри строки
    """
    values: List[str] = []
    for line in safe_split_lines(text):
        parts = re.split(r"[;,]", line)
        for p in parts:
            v = (p or "").strip()
            if v:
                values.append(v)
    return values

def md_inline_code(value: str) -> str:
    """
    Безопасно вставить пользовательское значение в Markdown как `inline code`.
    Telegram Markdown ломается на незакрытых/вложенных backticks, поэтому убираем их.
    """
    s = (value or "").replace("\n", " ").replace("\r", " ")
    s = s.replace("`", "'")
    return f"`{s}`"


def windows_human(windows_json: str) -> str:
    raw = (windows_json or "").strip() or "[]"
    try:
        data = json.loads(raw)
    except Exception:
        return raw
    if not data:
        return "24/7"
    parts = []
    for w in data[:3]:
        days = ",".join(str(x) for x in (w.get("days") or []))
        parts.append(f"days={days} {w.get('from', '00:00')}-{w.get('to', '23:59')}")
    if len(data) > 3:
        parts.append("...")
    return "; ".join(parts)

def prune_log_events_job():
    cutoff = utc_now() - timedelta(days=LOG_RETENTION_DAYS)

    def _inner(db):
        deleted = (
            db.query(LogEvent)
            .filter(LogEvent.created_at < cutoff)
            .delete(synchronize_session=False)
        )
        db.commit()
        return deleted

    deleted = db_exec(_inner)
    if deleted:
        logger.info(f"[INFO] 🧹 Pruned log_events: {deleted}")

def parse_windows_text(text: str) -> List[dict]:
    """Парсить окна активности"""
    import re
    windows = []
    for line in safe_split_lines(text):
        m_days = re.search(r"days=([0-9,]+)", line)
        m_from = re.search(r"from=([0-9]{2}:[0-9]{2})", line)
        m_to = re.search(r"to=([0-9]{2}:[0-9]{2})", line)
        if not (m_days and m_from and m_to):
            raise ValueError("Формат: days=1,2,3 from=09:00 to=18:00")
        days = [int(x) for x in m_days.group(1).split(",")]
        windows.append({"days": days, "from": m_from.group(1), "to": m_to.group(1)})
    return windows


def is_in_window_utc(windows: List[dict], dt_utc: datetime) -> bool:
    if not windows:
        return True
    wd = dt_utc.isoweekday()  # 1..7
    prev_wd = 7 if wd == 1 else (wd - 1)
    hhmm = dt_utc.strftime("%H:%M")
    for w in windows:
        days = w.get("days", []) or []
        t_from = w.get("from", "00:00")
        t_to = w.get("to", "23:59")

        # Normal window within a day.
        if t_from <= t_to:
            if wd in days and t_from <= hhmm < t_to:
                return True
            continue

        # Crossing midnight, e.g. 22:00 -> 02:00.
        # Interpret "days" as the day when the window starts.
        if wd in days and hhmm >= t_from:
            return True
        if prev_wd in days and hhmm < t_to:
            return True
    return False


def get_setting(db, key: str, default: str = "") -> str:
    obj = db.query(Setting).filter(Setting.key == key).first()
    return obj.value if obj else default


def set_setting(db, key: str, value: str):
    obj = db.query(Setting).filter(Setting.key == key).first()
    if obj:
        obj.value = value
        obj.updated_at = utc_now()
    else:
        db.add(Setting(key=key, value=value))
    db.commit()


def ensure_runtime_settings() -> None:
    def _inner(db):
        defaults = {
            "runtime_timezone": DEFAULT_TIMEZONE,
            "target_senders_count": str(DEFAULT_TARGET_SENDERS_COUNT),
            "daily_repeat_campaign_enabled": "1" if DEFAULT_DAILY_REPEAT_CAMPAIGN_ENABLED else "0",
            "processing_enabled": "1",
            "min_request_interval_sec": str(DEFAULT_MIN_REQUEST_INTERVAL_SEC),
            "max_request_interval_sec": str(max(DEFAULT_MIN_REQUEST_INTERVAL_SEC, DEFAULT_MAX_REQUEST_INTERVAL_SEC)),
            "hourly_api_limit": str(DEFAULT_HOURLY_API_LIMIT),
            "daily_api_limit": str(DEFAULT_DAILY_API_LIMIT),
            "send_api_cost": str(DEFAULT_SEND_API_COST),
            "check_api_cost": str(DEFAULT_CHECK_API_COST),
            "nickname_change_api_cost": "2",
            "recheck_daily_limit": str(DEFAULT_RECHECK_DAILY_LIMIT),
            "jitter_min_sec": str(DEFAULT_SEND_JITTER_MIN_SEC),
            "jitter_max_sec": str(DEFAULT_SEND_JITTER_MAX_SEC),
            "sender_switch_min_sec": "0",
            "sender_switch_max_sec": "0",
            "new_send_requests_enabled": "1",
            "recheck_only_mode_enabled": "0",
        }
        for key, val in defaults.items():
            if db.query(Setting).filter(Setting.key == key).first() is None:
                db.add(Setting(key=key, value=val))
        # Bot works in MSK only.
        tz_obj = db.query(Setting).filter(Setting.key == "runtime_timezone").first()
        if tz_obj:
            tz_obj.value = DEFAULT_TIMEZONE
            tz_obj.updated_at = utc_now()
        db.commit()

    db_exec(_inner)


def get_campaign_or_default(db, campaign_id: Optional[int] = None) -> Optional[Campaign]:
    camp = None
    if campaign_id:
        camp = db.query(Campaign).filter(Campaign.id == int(campaign_id)).first()
    if camp is None:
        camp = db.query(Campaign).order_by(Campaign.id.asc()).first()
    return camp


def target_campaign_filter(db, campaign_id: int):
    camp = get_campaign_or_default(db, campaign_id)
    if not camp:
        return Target.campaign_id.is_(None)
    if camp.name == "Основная":
        return or_(Target.campaign_id == int(camp.id), Target.campaign_id.is_(None))
    return Target.campaign_id == int(camp.id)


def task_campaign_filter(db, campaign_id: int):
    camp = get_campaign_or_default(db, campaign_id)
    if not camp:
        return Task.campaign_id.is_(None)
    if camp.name == "Основная":
        return or_(Task.campaign_id == int(camp.id), Task.campaign_id.is_(None))
    return Task.campaign_id == int(camp.id)


def get_setting_int(db, key: str, default: int) -> int:
    try:
        return int(get_setting(db, key, str(default)))
    except Exception:
        return int(default)


def get_setting_bool(db, key: str, default: bool = False) -> bool:
    val = get_setting(db, key, "1" if default else "0").strip().lower()
    return val in {"1", "true", "yes", "on"}


def is_processing_enabled() -> bool:
    return bool(db_exec(lambda db: get_setting_bool(db, "processing_enabled", True)))


def set_processing_enabled(enabled: bool):
    db_exec(lambda db: set_setting(db, "processing_enabled", "1" if enabled else "0"))


def show_processing_status(chat_id: int):
    enabled = is_processing_enabled()
    show_menu_status(chat_id, "manage", f"ℹ️ Статус обработки: {'включена' if enabled else 'выключена'}.")


def is_new_send_requests_enabled(db) -> bool:
    return get_setting_bool(db, "new_send_requests_enabled", True)


def set_new_send_requests_enabled(db, enabled: bool):
    set_setting(db, "new_send_requests_enabled", "1" if enabled else "0")


def is_recheck_only_mode_enabled(db) -> bool:
    return get_setting_bool(db, "recheck_only_mode_enabled", False)


def set_recheck_only_mode_enabled(db, enabled: bool):
    set_setting(db, "recheck_only_mode_enabled", "1" if enabled else "0")


def validate_requested_nickname(raw_value: str) -> tuple[bool, str]:
    nickname = str(raw_value or "").strip()
    if not nickname:
        return False, "пустой ник"
    if not NICKNAME_ALLOWED_RE.fullmatch(nickname):
        return False, "формат: только латиница/цифры/_ и длина 3..16"
    return True, ""


def _nickname_fallback_candidates(nickname: str, limit: int = NICKNAME_FALLBACK_LIMIT) -> list[str]:
    """
    Построить кандидатов "похожих" ников (ASCII-safe):
    - визуальные подмены символов (o->0, s->5, ...)
    - суффиксы, если ник всё ещё занят.
    """
    base = str(nickname or "").strip()
    ok, _ = validate_requested_nickname(base)
    if not ok:
        return []

    out: list[str] = []
    seen: set[str] = {base}

    def _push(val: str):
        if len(out) >= int(limit):
            return
        if not val or val in seen:
            return
        v_ok, _ = validate_requested_nickname(val)
        if not v_ok:
            return
        seen.add(val)
        out.append(val)

    # 1) Single-char visual substitutions.
    for i, ch in enumerate(base):
        repls = NICKNAME_VISUAL_SUBS.get(ch.lower(), ())
        for repl in repls:
            cand = f"{base[:i]}{repl}{base[i + 1:]}"
            _push(cand)
            if len(out) >= int(limit):
                return out

    # 2) Suffix variants (append or overwrite tail).
    suffixes = ("1", "2", "3", "5", "7", "8", "9", "01", "02", "03", "11", "77")
    for suf in suffixes:
        if len(base) + len(suf) <= 16:
            cand = f"{base}{suf}"
        else:
            keep = max(0, 16 - len(suf))
            cand = f"{base[:keep]}{suf}"
        _push(cand)
        if len(out) >= int(limit):
            return out

    return out


def get_runtime_timezone(db) -> ZoneInfo:
    tz_name = get_setting(db, "runtime_timezone", DEFAULT_TIMEZONE).strip() or DEFAULT_TIMEZONE
    try:
        return ZoneInfo(tz_name)
    except Exception:
        return ZoneInfo(DEFAULT_TIMEZONE)


def local_day_bounds_utc_naive(db, now_utc: datetime) -> tuple[datetime, datetime]:
    """
    Границы текущих локальных суток (по runtime TZ) в формате UTC naive.
    """
    tz = get_runtime_timezone(db)
    now_local = now_utc.replace(tzinfo=timezone.utc).astimezone(tz)
    start_local = now_local.replace(hour=0, minute=0, second=0, microsecond=0)
    end_local = start_local + timedelta(days=1)
    start_utc = start_local.astimezone(timezone.utc).replace(tzinfo=None)
    end_utc = end_local.astimezone(timezone.utc).replace(tzinfo=None)
    return start_utc, end_utc


def _campaign_daily_intervals_utc(db, windows_json: str, now_utc: datetime) -> List[tuple[datetime, datetime]]:
    """
    Вернуть интервалы активности текущих локальных суток в UTC naive.
    Если окна не заданы/некорректны — вернуть 24/7 (полные сутки).
    """
    day_start_utc, day_end_utc = local_day_bounds_utc_naive(db, now_utc)
    raw = (windows_json or "").strip() or "[]"
    try:
        windows = json.loads(raw)
    except Exception:
        windows = []
    if not windows:
        return [(day_start_utc, day_end_utc)]

    tz = get_runtime_timezone(db)
    now_local = now_utc.replace(tzinfo=timezone.utc).astimezone(tz)
    day_local_start = now_local.replace(hour=0, minute=0, second=0, microsecond=0)
    weekday = day_local_start.isoweekday()  # 1..7
    prev_weekday = 7 if weekday == 1 else weekday - 1
    intervals: List[tuple[datetime, datetime]] = []

    def _add_local_interval(start_local: datetime, end_local: datetime):
        if end_local <= start_local:
            return
        start_utc = start_local.astimezone(timezone.utc).replace(tzinfo=None)
        end_utc = end_local.astimezone(timezone.utc).replace(tzinfo=None)
        start_utc = max(start_utc, day_start_utc)
        end_utc = min(end_utc, day_end_utc)
        if end_utc > start_utc:
            intervals.append((start_utc, end_utc))

    for w in windows:
        days = [int(x) for x in (w.get("days") or []) if str(x).isdigit()]
        t_from = str(w.get("from", "00:00"))
        t_to = str(w.get("to", "23:59"))
        try:
            h1, m1 = [int(x) for x in t_from.split(":")]
            h2, m2 = [int(x) for x in t_to.split(":")]
        except Exception:
            continue

        start_today = day_local_start.replace(hour=h1, minute=m1)
        end_today = day_local_start.replace(hour=h2, minute=m2)

        if t_from <= t_to:
            if weekday in days:
                _add_local_interval(start_today, end_today)
        else:
            # Окно через полночь.
            # Хвост прошлого дня, который попадает в текущие сутки.
            if prev_weekday in days:
                _add_local_interval(day_local_start, end_today)
            # Начало сегодняшнего дня, переходящее на завтра.
            if weekday in days:
                _add_local_interval(start_today, day_local_start + timedelta(days=1))

    if not intervals:
        return [(day_start_utc, day_end_utc)]
    intervals.sort(key=lambda x: x[0])
    return intervals


def _map_offset_to_intervals(intervals: List[tuple[datetime, datetime]], offset_sec: int) -> datetime:
    remaining = max(0, int(offset_sec))
    for start_dt, end_dt in intervals:
        span = max(1, int((end_dt - start_dt).total_seconds()))
        if remaining < span:
            return start_dt + timedelta(seconds=remaining)
        remaining -= span
    return intervals[-1][1] - timedelta(seconds=1)


def _window_elapsed_seconds(intervals: List[tuple[datetime, datetime]], now_utc: datetime) -> int:
    elapsed = 0
    for start_dt, end_dt in intervals:
        if now_utc <= start_dt:
            continue
        if now_utc >= end_dt:
            elapsed += max(0, int((end_dt - start_dt).total_seconds()))
        else:
            elapsed += max(0, int((now_utc - start_dt).total_seconds()))
    return max(0, elapsed)


def _window_total_seconds(intervals: List[tuple[datetime, datetime]]) -> int:
    return max(1, sum(max(1, int((e - s).total_seconds())) for s, e in intervals))


def _campaign_local_day_str(db, now_utc: datetime) -> str:
    tz = get_runtime_timezone(db)
    now_local = now_utc.replace(tzinfo=timezone.utc).astimezone(tz)
    return now_local.strftime("%Y-%m-%d")


def _campaign_effective_daily_limit(db, camp: Campaign, now_utc: datetime) -> int:
    """
    Freeze per-campaign daily limit for local day.
    When user changes goal limit via UI, cache is explicitly reset and new value applies immediately.
    """
    current_limit = max(1, int(camp.daily_limit_per_account or DEFAULT_DAILY_LIMIT))
    day_key = _campaign_local_day_str(db, now_utc)
    k_day = f"camp_daily_limit_day_{int(camp.id)}"
    k_val = f"camp_daily_limit_val_{int(camp.id)}"
    stored_day = str(get_setting(db, k_day, "") or "")
    stored_val = int(get_setting_int(db, k_val, current_limit))
    if stored_day != day_key:
        set_setting(db, k_day, day_key)
        set_setting(db, k_val, str(current_limit))
        return current_limit
    return max(1, int(stored_val))


def _campaign_pacing_gate(db, camp: Campaign, now_utc: datetime) -> tuple[bool, Optional[datetime], str]:
    """
    Global campaign pacing: spread today's planned send tasks across campaign window.
    """
    day_start_utc, day_end_utc = local_day_bounds_utc_naive(db, now_utc)
    intervals = _campaign_daily_intervals_utc(db, camp.active_windows_json or "[]", now_utc)
    total_window_sec = _window_total_seconds(intervals)
    elapsed_sec = _window_elapsed_seconds(intervals, now_utc)

    planned_today = int(
        db.query(func.count(Task.id))
        .filter(
            Task.campaign_id == int(camp.id),
            Task.task_type == "send_request",
            Task.status != TaskStatus.CANCELLED.value,
            Task.scheduled_for >= day_start_utc,
            Task.scheduled_for < day_end_utc,
        )
        .scalar()
        or 0
    )
    if planned_today <= 0:
        return True, None, "no_plan"

    done_today = int(
        db.query(func.count(Task.id))
        .filter(
            Task.campaign_id == int(camp.id),
            Task.task_type == "send_request",
            Task.status == TaskStatus.DONE.value,
            Task.completed_at >= day_start_utc,
            Task.completed_at < day_end_utc,
        )
        .scalar()
        or 0
    )

    if elapsed_sec >= total_window_sec:
        return True, None, "window_done"

    allowed_done_by_now = int((planned_today * elapsed_sec) // total_window_sec)
    if done_today <= allowed_done_by_now:
        return True, None, "ok"

    slot_sec = max(1, total_window_sec // max(1, planned_today))
    target_idx = max(0, done_today)
    next_offset = min(target_idx * slot_sec, max(0, total_window_sec - 1))
    next_at = _map_offset_to_intervals(intervals, next_offset)
    next_at = next_at + timedelta(seconds=random.randint(0, max(1, slot_sec // 3)))
    if next_at <= now_utc:
        next_at = now_utc + timedelta(seconds=max(30, slot_sec // 2))
    return False, next_at, "paced_by_window"


def _scheduled_send_time_spread(
    db,
    now_utc: datetime,
    acc: Account,
    campaign_id_for_tasks: Optional[int],
    daily_limit: int,
    min_s: int,
    max_s: int,
    windows_json: str,
) -> datetime:
    """
    Распределить отправки аккаунта по суточному окну:
    позиция слота = max(уже запланировано, текущее положение в окне).
    """
    day_start_utc, day_end_utc = local_day_bounds_utc_naive(db, now_utc)
    q = db.query(func.count(Task.id)).filter(
        Task.task_type == "send_request",
        Task.status != TaskStatus.CANCELLED.value,
        Task.account_id == int(acc.id),
        Task.scheduled_for >= day_start_utc,
        Task.scheduled_for < day_end_utc,
    )
    if campaign_id_for_tasks is None:
        q = q.filter(Task.campaign_id.is_(None))
    else:
        q = q.filter(Task.campaign_id == int(campaign_id_for_tasks))
    planned_today = int(q.scalar() or 0)

    intervals = _campaign_daily_intervals_utc(db, windows_json, now_utc)
    total_window_sec = _window_total_seconds(intervals)
    slot_sec = max(1, total_window_sec // max(1, int(daily_limit)))

    elapsed_now = _window_elapsed_seconds(intervals, now_utc)
    # Берём ближайший не-прошедший слот (ceil), чтобы не сжимать всё "в ближайшие минуты".
    current_slot_idx = (elapsed_now + slot_sec - 1) // slot_sec
    slot_idx = max(planned_today, current_slot_idx)
    if slot_idx >= daily_limit:
        slot_idx = max(0, daily_limit - 1)

    # Разводим аккаунты по фазе внутри слота, чтобы не бить пачкой в одну секунду.
    phase_span = max(1, slot_sec // 2)
    phase_shift = int(acc.id % phase_span)
    base_offset = int(slot_idx * slot_sec + phase_shift)
    base_offset = min(base_offset, max(0, total_window_sec - 1))
    base_dt = _map_offset_to_intervals(intervals, base_offset)

    jitter_cap = max(0, min(int(max_s), max(1, slot_sec // 3)))
    jitter_min = max(0, min(int(min_s), jitter_cap))
    jitter = random.randint(jitter_min, jitter_cap) if jitter_cap > 0 else 0
    dt = base_dt + timedelta(seconds=int(jitter))
    if dt < now_utc:
        dt = now_utc + timedelta(seconds=max(1, int(slot_sec // 3)))
    return dt


def is_in_window_for_account(db, windows: List[dict], dt_utc: datetime) -> bool:
    """
    Проверка окна по локальному TZ, заданному в settings.runtime_timezone.
    Хранение окон остаётся прежним (days/from/to), но интерпретация идёт в выбранной TZ.
    """
    try:
        tz = get_runtime_timezone(db)
        local_dt = dt_utc.replace(tzinfo=timezone.utc).astimezone(tz).replace(tzinfo=None)
    except Exception:
        local_dt = dt_utc
    return is_in_window_utc(windows, local_dt)


def enforce_api_rate_limit(db, acc: Account, now: datetime, api_cost: int) -> tuple[bool, Optional[datetime], str]:
    """
    Жёсткий rate-limit на аккаунт:
    - минимальный интервал между API-запросами
    - максимум API-запросов в час
    - максимум API-запросов в сутки
    """
    min_interval = max(0, get_setting_int(db, "min_request_interval_sec", DEFAULT_MIN_REQUEST_INTERVAL_SEC))
    max_interval = max(min_interval, get_setting_int(db, "max_request_interval_sec", DEFAULT_MAX_REQUEST_INTERVAL_SEC))
    hourly_limit = max(1, get_setting_int(db, "hourly_api_limit", DEFAULT_HOURLY_API_LIMIT))
    daily_limit = max(1, get_setting_int(db, "daily_api_limit", DEFAULT_DAILY_API_LIMIT))

    if acc.api_next_allowed_at and now < acc.api_next_allowed_at:
        return False, acc.api_next_allowed_at, "min_interval"

    if acc.last_api_request_at:
        elapsed = (now - acc.last_api_request_at).total_seconds()
        if elapsed < min_interval:
            wait_sec = max(1, int(min_interval - elapsed))
            return False, now + timedelta(seconds=wait_sec), "min_interval"

    if acc.api_hour_window_start is None:
        # Legacy safety: some rows may have counters without window start.
        # Initialize fresh window to avoid endless hourly postpones.
        acc.api_hour_window_start = now
        acc.api_hour_count = 0
    hour_start = acc.api_hour_window_start or now
    if (now - hour_start) >= timedelta(hours=1):
        acc.api_hour_window_start = now
        acc.api_hour_count = 0

    if acc.api_day_window_start is None:
        # Legacy safety for daily window metadata.
        acc.api_day_window_start = now
        acc.api_day_count = 0
    day_start = acc.api_day_window_start or now
    if (now - day_start) >= timedelta(days=1):
        acc.api_day_window_start = now
        acc.api_day_count = 0

    if (int(acc.api_hour_count or 0) + api_cost) > hourly_limit:
        next_at = (acc.api_hour_window_start or now) + timedelta(hours=1, seconds=1)
        return False, next_at, "hourly_limit"

    if (int(acc.api_day_count or 0) + api_cost) > daily_limit:
        next_at = (acc.api_day_window_start or now) + timedelta(days=1, seconds=1)
        return False, next_at, "daily_api_limit"

    acc.api_hour_count = int(acc.api_hour_count or 0) + api_cost
    acc.api_day_count = int(acc.api_day_count or 0) + api_cost
    acc.last_api_request_at = now
    acc.api_next_allowed_at = now + timedelta(seconds=random.randint(min_interval, max_interval))
    return True, None, "ok"


def target_required_senders(db, tgt: Target) -> int:
    # Compatibility wrapper for existing call sites/tests.
    fallback = get_setting_int(db, "target_senders_count", DEFAULT_TARGET_SENDERS_COUNT)
    return campaign_target_required_senders(db, tgt, fallback)


def jitter_seconds_with_db(db) -> int:
    """Джиттер, который использует уже открытую сессию (без db_exec)."""
    min_s = int(get_setting(db, "jitter_min_sec", str(DEFAULT_SEND_JITTER_MIN_SEC)))
    max_s = int(get_setting(db, "jitter_max_sec", str(DEFAULT_SEND_JITTER_MAX_SEC)))
    if max_s < min_s:
        max_s = min_s
    return random.randint(min_s, max_s)


def get_proxy_for_account(db, account_id: int) -> Optional[str]:
    """Получить прокси для аккаунта"""
    acc = db.query(Account).filter(Account.id == account_id).first()
    if not acc or not acc.proxy_id:
        return None
    proxy = db.query(Proxy).filter(Proxy.id == acc.proxy_id, Proxy.enabled == True).first()
    return proxy.url if proxy else None


def import_accounts_from_excel(path: str) -> tuple:
    """Импортировать аккаунты из Excel"""
    wb = load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0, 0, 0

    start_idx = 1 if rows and any("login" in str(x).lower() for x in rows[0]) else 0
    added = skipped = errors = 0

    def _inner(db):
        nonlocal added, skipped, errors
        parsed_rows = []
        seen_logins = set()
        for row in rows[start_idx:]:
            if not row or not (row[0] and len(row) > 1 and row[1]):
                errors += 1
                continue
            login = str(row[0]).strip()
            password = str(row[1]).strip()
            if login in seen_logins:
                skipped += 1
                continue
            seen_logins.add(login)
            parsed_rows.append((login, password))

        if not parsed_rows:
            db.commit()
            return

        existing_logins = {
            x[0] for x in db.query(Account.login).filter(Account.login.in_([p[0] for p in parsed_rows])).all()
        }
        for login, password in parsed_rows:
            if login in existing_logins:
                skipped += 1
                continue
            db.add(
                Account(
                    login=login,
                    password=password,
                    daily_limit=DEFAULT_DAILY_LIMIT,
                    active_windows_json="[]",
                    warmup_until=utc_now() + timedelta(minutes=5),
                )
            )
            added += 1
        db.commit()

    db_exec(_inner)
    return added, skipped, errors


def _parse_account_line(line: str) -> Optional[tuple[str, str]]:
    for sep in (":", ";", ",", "\t"):
        if sep in line:
            left, right = line.split(sep, 1)
            login = left.strip()
            password = right.strip()
            if login and password:
                return login, password
    return None


def import_accounts_from_text(path: str) -> tuple:
    """Импортировать аккаунты из txt/csv: login:password (или ; , tab)."""
    with open(path, "r", encoding="utf-8-sig") as fp:
        raw_lines = [ln.strip() for ln in fp.readlines()]

    lines = [ln for ln in raw_lines if ln and not ln.startswith("#")]
    added = skipped = errors = 0

    def _inner(db):
        nonlocal added, skipped, errors
        parsed_rows = []
        seen_logins = set()
        for ln in lines:
            parsed = _parse_account_line(ln)
            if not parsed:
                errors += 1
                continue

            login, password = parsed
            if login in seen_logins:
                skipped += 1
                continue
            seen_logins.add(login)
            parsed_rows.append((login, password))

        if not parsed_rows:
            db.commit()
            return

        existing_logins = {
            x[0] for x in db.query(Account.login).filter(Account.login.in_([p[0] for p in parsed_rows])).all()
        }
        for login, password in parsed_rows:
            if login in existing_logins:
                skipped += 1
                continue

            db.add(
                Account(
                    login=login,
                    password=password,
                    daily_limit=DEFAULT_DAILY_LIMIT,
                    active_windows_json="[]",
                    warmup_until=utc_now() + timedelta(minutes=5),
                )
            )
            added += 1
        db.commit()

    db_exec(_inner)
    return added, skipped, errors


def import_targets_from_text(path: str, campaign_id: Optional[int] = None) -> tuple:
    """Импортировать цели из txt/csv: по одному нику на строку."""
    with open(path, "r", encoding="utf-8-sig") as fp:
        names = [ln.strip() for ln in fp.readlines() if ln.strip() and not ln.strip().startswith("#")]

    def _inner(db):
        added = 0
        duplicate_in_payload = 0
        seen_names = set()
        unique_names = []
        for name in names:
            if name in seen_names:
                duplicate_in_payload += 1
                continue
            seen_names.add(name)
            unique_names.append(name)

        if not unique_names:
            return 0, duplicate_in_payload, 0

        camp = get_campaign_or_default(db, campaign_id)
        if not camp:
            return 0, duplicate_in_payload, 0
        required = max(1, int(camp.target_senders_count or DEFAULT_TARGET_SENDERS_COUNT))
        effective_campaign_id = int(camp.id)
        existing_names = {
            x[0]
            for x in db.query(Target.username)
            .filter(
                Target.username.in_(unique_names),
                target_campaign_filter(db, effective_campaign_id),
            )
            .all()
        }
        already_in_db = 0
        for name in unique_names:
            if name in existing_names:
                already_in_db += 1
                continue
            db.add(
                Target(
                    username=name,
                    campaign_id=effective_campaign_id,
                    status=TargetStatus.NEW.value,
                    priority=random.randint(1, 100),
                    required_senders=required,
                )
            )
            added += 1
        db.commit()
        return added, duplicate_in_payload, already_in_db

    added, duplicate_in_payload, already_in_db = db_exec(_inner)
    skipped = duplicate_in_payload + already_in_db
    return added, skipped, 0


def _parse_nickname_change_line(line: str) -> Optional[tuple[str, str]]:
    raw = str(line or "").strip()
    if not raw or raw.startswith("#"):
        return None
    for sep in (";", ",", "\t", ":"):
        if sep in raw:
            left, right = raw.split(sep, 1)
            login = str(left or "").strip()
            nickname = str(right or "").strip()
            if login and nickname:
                return login, nickname
    return None


def _load_nickname_change_rows_from_xlsx(path: str) -> list[tuple[str, str]]:
    wb = load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    start_idx = 0
    head = rows[0] or ()
    head_l = [str(x or "").strip().lower() for x in head[:2]]
    if any("email" in x or "почт" in x or "login" in x for x in head_l):
        start_idx = 1
    out: list[tuple[str, str]] = []
    for row in rows[start_idx:]:
        if not row:
            continue
        login = str((row[0] if len(row) > 0 else "") or "").strip()
        nickname = str((row[1] if len(row) > 1 else "") or "").strip()
        if login and nickname:
            out.append((login, nickname))
    return out


def _load_nickname_change_rows_from_text(path: str) -> list[tuple[str, str]]:
    rows: list[tuple[str, str]] = []
    with open(path, "r", encoding="utf-8-sig") as fp:
        content = fp.read()
    # First pass: csv parser (works for comma/semicolon/tab files with quoted cells).
    parsed_any = False
    for delim in (";", ",", "\t"):
        reader = csv.reader(content.splitlines(), delimiter=delim)
        candidate: list[tuple[str, str]] = []
        for parts in reader:
            if not parts:
                continue
            if len(parts) < 2:
                continue
            login = str(parts[0] or "").strip()
            nickname = str(parts[1] or "").strip()
            if login and nickname and not login.startswith("#"):
                candidate.append((login, nickname))
        if candidate:
            parsed_any = True
            rows.extend(candidate)
            break
    if parsed_any:
        return rows
    # Fallback: free-form lines.
    for line in content.splitlines():
        parsed = _parse_nickname_change_line(line)
        if parsed:
            rows.append(parsed)
    return rows


def import_nickname_change_tasks(path: str, source_file: str = "") -> tuple[int, int, int]:
    """
    Импортировать задания на смену ников:
    формат строк: login/email + новый ник.
    """
    lower_name = str(path or "").lower()
    if lower_name.endswith(".xlsx"):
        rows = _load_nickname_change_rows_from_xlsx(path)
    else:
        rows = _load_nickname_change_rows_from_text(path)

    added = 0
    skipped = 0
    errors = 0

    def _inner(db):
        nonlocal added, skipped, errors
        if not rows:
            return

        # Keep the latest value for duplicate logins in one payload.
        latest_by_login: dict[str, str] = {}
        for login, nickname in rows:
            key = str(login or "").strip().lower()
            if not key:
                errors += 1
                continue
            latest_by_login[key] = str(nickname or "").strip()

        if not latest_by_login:
            return

        wanted_logins = list(latest_by_login.keys())
        acc_rows = (
            db.query(Account)
            .filter(func.lower(Account.login).in_(wanted_logins))
            .all()
        )
        acc_by_login = {str(a.login or "").strip().lower(): a for a in acc_rows}

        active_jobs = (
            db.query(NicknameChangeTask.account_id)
            .filter(NicknameChangeTask.status.in_([
                NicknameChangeStatus.QUEUED.value,
                NicknameChangeStatus.POSTPONED.value,
                NicknameChangeStatus.RUNNING.value,
            ]))
            .all()
        )
        active_account_ids = {int(x[0]) for x in active_jobs if x and x[0]}

        for login_l, nickname in latest_by_login.items():
            acc = acc_by_login.get(login_l)
            if not acc:
                errors += 1
                continue
            ok_nick, _ = validate_requested_nickname(nickname)
            if not ok_nick:
                errors += 1
                continue
            if not (acc.epic_account_id and acc.device_id and acc.device_secret):
                errors += 1
                continue
            if int(acc.id) in active_account_ids:
                skipped += 1
                continue
            db.add(
                NicknameChangeTask(
                    account_id=int(acc.id),
                    requested_nick=str(nickname),
                    status=NicknameChangeStatus.QUEUED.value,
                    scheduled_for=utc_now(),
                    attempt_number=0,
                    max_attempts=3,
                    source_file=(source_file or "")[:256],
                )
            )
            active_account_ids.add(int(acc.id))
            added += 1
        db.commit()

    db_exec(_inner)
    return int(added), int(skipped), int(errors)


def show_nickname_change_status(chat_id: int):
    def _inner(db):
        by_status = dict(
            db.query(NicknameChangeTask.status, func.count(NicknameChangeTask.id))
            .group_by(NicknameChangeTask.status)
            .all()
        )
        last_rows = (
            db.query(NicknameChangeTask, Account.login, Account.epic_display_name)
            .join(Account, Account.id == NicknameChangeTask.account_id)
            .order_by(NicknameChangeTask.id.desc())
            .limit(15)
            .all()
        )
        return by_status, last_rows

    by_status, rows = db_exec(_inner)
    status_order = [
        NicknameChangeStatus.QUEUED.value,
        NicknameChangeStatus.POSTPONED.value,
        NicknameChangeStatus.RUNNING.value,
        NicknameChangeStatus.DONE.value,
        NicknameChangeStatus.FAILED.value,
        NicknameChangeStatus.SKIPPED.value,
    ]
    lines = ["📝 Смена ников: очередь"]
    lines.append("Статусы:")
    for key in status_order:
        lines.append(f"• {key}: {int(by_status.get(key, 0) or 0)}")
    lines.append("")
    lines.append("Подсказка: при nickname_taken бот пробует похожие варианты автоматически.")
    if rows:
        lines.append("")
        lines.append("Последние задания:")
        for task, login, display_name in rows:
            current = str(display_name or "").strip() or "-"
            final = str(task.final_nick or "").strip() or "-"
            err = str(task.last_error or "").strip()
            if len(err) > 80:
                err = err[:77] + "..."
            lines.append(
                f"#{task.id} acc#{task.account_id} {login}\n"
                f"  {current} -> {task.requested_nick} (итог: {final}) [{task.status}]"
                + (f"\n  err: {err}" if err else "")
            )
    show_menu_status(chat_id, "accounts", "\n".join(lines))


def export_results_to_excel(filename: str = "/tmp/results.xlsx") -> str:
    """Экспортировать результаты в Excel"""

    def _inner(db):
        targets = db.query(Target).all()
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        headers = ["Username", "Status", "Attempts", "Sent", "Accepted", "First", "Last", "Error"]
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        for t in targets:
            ws.append([
                t.username,
                t.status,
                t.attempt_count,
                t.sent_count,
                t.accepted_count,
                t.first_attempt_at.strftime("%Y-%m-%d %H:%M") if t.first_attempt_at else "",
                t.last_attempt_at.strftime("%Y-%m-%d %H:%M") if t.last_attempt_at else "",
                t.last_error or "",
            ])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(
                max(len(str(cell.value)) for cell in col) + 2, 50
            )
        wb.save(filename)
        return filename

    return db_exec(_inner)


# ============================================================
# CORE LOGIC
# ============================================================

def reset_daily_counters_job():
    """Сбросить суточные лимиты"""
    today = utc_today()

    def _inner(db):
        accs = db.query(Account).all()
        updated = 0
        for a in accs:
            if not a.last_reset_date or a.last_reset_date.date() < today:
                a.today_sent = 0
                a.last_reset_date = utc_now()
                updated += 1
        db.commit()
        return updated

    updated = db_exec(_inner)
    if updated:
        log_event("info", f"🔄 Сброшены суточные лимиты: {updated} аккаунтов")


def pick_best_account(db) -> Optional[Account]:
    """Выбрать лучший аккаунт"""
    now = utc_now()
    accs = db.query(Account).filter(
        Account.enabled == True,
        Account.status == AccountStatus.ACTIVE.value,
        Account.epic_account_id.isnot(None),
        Account.device_id.isnot(None),
        Account.device_secret.isnot(None),
        Account.today_sent < Account.daily_limit,
    ).all()
    valid = []
    for a in accs:
        if a.today_sent >= a.daily_limit:
            continue
        if a.warmup_until and now < a.warmup_until:
            continue
        windows = json.loads(a.active_windows_json or "[]")
        if not is_in_window_for_account(db, windows, now):
            continue
        valid.append(a)
    if not valid:
        return None
    valid.sort(key=lambda a: (a.today_sent, -a.total_sent))
    return valid[0]


def _done_sender_ids_for_target(db, target_id: int, camp: Optional[Campaign], now: datetime) -> set[int]:
    q = db.query(Task.account_id).filter(
        Task.target_id == int(target_id),
        Task.task_type == "send_request",
        Task.status == TaskStatus.DONE.value,
    )
    if camp is not None and bool(camp.daily_repeat_enabled):
        day_start_utc, day_end_utc = local_day_bounds_utc_naive(db, now)
        q = q.filter(
            Task.completed_at >= day_start_utc,
            Task.completed_at < day_end_utc,
        )
    return {int(x[0]) for x in q.distinct().all()}


def _precheck_skipped_sender_ids_for_target(db, target_id: int, camp: Optional[Campaign], now: datetime) -> set[int]:
    q = db.query(Task.account_id).filter(
        Task.target_id == int(target_id),
        Task.task_type == "send_request",
        Task.last_error.in_(PRECHECK_SKIP_REASONS),
    )
    if camp is not None and bool(camp.daily_repeat_enabled):
        day_start_utc, day_end_utc = local_day_bounds_utc_naive(db, now)
        q = q.filter(
            Task.completed_at >= day_start_utc,
            Task.completed_at < day_end_utc,
        )
    return {int(x[0]) for x in q.distinct().all()}


def _pick_replacement_account_for_target(
    db,
    target_id: int,
    camp: Optional[Campaign],
    now: datetime,
    excluded_ids: set[int],
) -> Optional[Account]:
    candidate_accounts = (
        db.query(Account)
        .filter(
            Account.enabled == True,
            Account.status == AccountStatus.ACTIVE.value,
            Account.epic_account_id.isnot(None),
            Account.device_id.isnot(None),
            Account.device_secret.isnot(None),
        )
        .order_by(Account.today_sent.asc(), Account.id.asc())
        .all()
    )

    done_sender_ids = _done_sender_ids_for_target(db, int(target_id), camp, now)
    blocked_ids = set(int(x) for x in excluded_ids) | set(done_sender_ids)

    if camp is not None:
        campaign_daily_limit = _campaign_effective_daily_limit(db, camp, now)
        try:
            camp_windows = json.loads(camp.active_windows_json or "[]")
        except Exception:
            camp_windows = []
    else:
        campaign_daily_limit = 0
        camp_windows = []

    for acc in candidate_accounts:
        aid = int(acc.id)
        if aid in blocked_ids:
            continue
        if acc.warmup_until and now < acc.warmup_until:
            continue
        try:
            windows = json.loads(acc.active_windows_json or "[]")
        except Exception:
            windows = []
        if not is_in_window_for_account(db, windows, now):
            continue
        if camp_windows and not is_in_window_for_account(db, camp_windows, now):
            continue
        if int(acc.today_sent or 0) >= int(acc.daily_limit or 0):
            continue
        if camp is not None:
            sent_today = campaign_sent_today_for_account(db, int(acc.id), int(camp.id), now)
            if int(sent_today) >= int(campaign_daily_limit):
                continue
        has_active_pair = db.query(Task.id).filter(
            Task.task_type == "send_request",
            Task.target_id == int(target_id),
            Task.account_id == int(acc.id),
            Task.status.in_(ACTIVE_SEND_TASK_STATUSES),
        ).first()
        if has_active_pair:
            continue
        return acc
    return None


def _enqueue_replacement_send_task(
    db,
    target: Target,
    camp: Optional[Campaign],
    now: datetime,
    excluded_ids: set[int],
    source_reason: str,
) -> bool:
    pick_excluded = set(int(x) for x in excluded_ids)
    pick_excluded |= _precheck_skipped_sender_ids_for_target(db, int(target.id), camp, now)
    replacement = _pick_replacement_account_for_target(
        db,
        target_id=int(target.id),
        camp=camp,
        now=now,
        excluded_ids=pick_excluded,
    )
    if replacement is None:
        return False

    if camp is not None:
        jitter_min = max(0, int(camp.jitter_min_sec or DEFAULT_SEND_JITTER_MIN_SEC))
        jitter_max = max(jitter_min, int(camp.jitter_max_sec or DEFAULT_SEND_JITTER_MAX_SEC))
        campaign_id = int(camp.id)
    else:
        jitter_min = max(0, get_setting_int(db, "jitter_min_sec", DEFAULT_SEND_JITTER_MIN_SEC))
        jitter_max = max(jitter_min, get_setting_int(db, "jitter_max_sec", DEFAULT_SEND_JITTER_MAX_SEC))
        campaign_id = None

    scheduled_for = now + timedelta(seconds=random.randint(int(jitter_min), int(jitter_max)))
    db.add(
        Task(
            task_type="send_request",
            status=TaskStatus.QUEUED.value,
            campaign_id=campaign_id,
            account_id=int(replacement.id),
            target_id=int(target.id),
            scheduled_for=scheduled_for,
            max_attempts=int(target.max_attempts or 3),
            last_error=str(source_reason or "replacement_after_precheck_skip"),
        )
    )
    target.status = TargetStatus.PENDING.value
    target.first_attempt_at = target.first_attempt_at or now
    return True


def pick_best_account_with_reservations_excluded(db, reserved: dict, excluded_ids: set[int]) -> Optional[Account]:
    now = utc_now()
    accs = db.query(Account).filter(
        Account.enabled == True,
        Account.status == AccountStatus.ACTIVE.value,
        Account.epic_account_id.isnot(None),
        Account.device_id.isnot(None),
        Account.device_secret.isnot(None),
    ).all()
    valid = []
    for a in accs:
        if int(a.id) in excluded_ids:
            continue
        if a.warmup_until and now < a.warmup_until:
            continue
        windows = json.loads(a.active_windows_json or "[]")
        if not is_in_window_for_account(db, windows, now):
            continue
        reserved_count = int(reserved.get(a.id, 0))
        effective = int(a.today_sent or 0) + reserved_count
        if effective >= int(a.daily_limit or 0):
            continue
        valid.append((effective, int(a.today_sent or 0), -int(a.total_sent or 0), a))
    if not valid:
        return None
    valid.sort(key=lambda t: (t[0], t[1], t[2], t[3].id))
    return valid[0][3]


def create_tasks_for_new_targets(db, limit: int = 500, campaign_id: Optional[int] = None) -> int:
    """
    Создать send_request задачи с fanout:
    для каждой цели назначаем заявки с разных аккаунтов до required_senders.
    """
    log_event_in_db(db, "info", "⚙️ create_tasks_for_new_targets: старт")
    if not is_new_send_requests_enabled(db):
        return 0
    created = 0
    now = utc_now()
    # Резерв по аккаунтам: уже существующие активные send_request задачи.
    # Нужен, чтобы не назначать сотни целей на один аккаунт, пока today_sent не растёт.
    reserved_by_account = {}
    for acc_id, cnt in (
        db.query(Task.account_id, func.count(Task.id))
        .filter(
            Task.task_type == "send_request",
            Task.status.in_(ACTIVE_SEND_TASK_STATUSES),
        )
        .group_by(Task.account_id)
        .all()
    ):
        reserved_by_account[int(acc_id)] = int(cnt)

    if campaign_id is None:
        # Legacy compatibility mode for old rows without campaign_id.
        campaign_filter = Target.campaign_id.is_(None)
        campaign_id_for_tasks: Optional[int] = None
        repeat_daily = get_setting_bool(db, "daily_repeat_campaign_enabled", False)
    else:
        camp = db.query(Campaign).filter(Campaign.id == int(campaign_id)).first()
        if camp is None:
            return 0
        if not camp.enabled:
            return 0
        campaign_filter = target_campaign_filter(db, int(camp.id))
        campaign_id_for_tasks = int(camp.id)
        repeat_daily = bool(camp.daily_repeat_enabled)
        if camp.name == "Основная":
            repeat_daily = repeat_daily or get_setting_bool(db, "daily_repeat_campaign_enabled", False)
    if repeat_daily:
        target_statuses = list(TARGET_SEND_BASE_STATUSES + TARGET_SEND_REPEAT_EXTRA_STATUSES)
    else:
        target_statuses = list(TARGET_SEND_BASE_STATUSES)

    day_start_utc, day_end_utc = local_day_bounds_utc_naive(db, now)

    targets = db.query(Target).filter(
        campaign_filter,
        Target.status.in_(target_statuses),
    ).order_by(
        Target.id.asc()
    ).limit(limit).all()

    # User flow is simplified: sequential sender dispatch is always used.
    dispatch_mode = "sequential"

    camp_for_cfg = db.query(Campaign).filter(Campaign.id == int(campaign_id_for_tasks)).first() if campaign_id_for_tasks else None
    if campaign_id_for_tasks is not None:
        min_s = max(0, int((camp_for_cfg.jitter_min_sec if camp_for_cfg else DEFAULT_SEND_JITTER_MIN_SEC) or DEFAULT_SEND_JITTER_MIN_SEC))
        max_s = max(min_s, int((camp_for_cfg.jitter_max_sec if camp_for_cfg else DEFAULT_SEND_JITTER_MAX_SEC) or DEFAULT_SEND_JITTER_MAX_SEC))
        if camp_for_cfg:
            default_daily_limit_per_account = _campaign_effective_daily_limit(db, camp_for_cfg, now)
        else:
            default_daily_limit_per_account = max(1, int(DEFAULT_DAILY_LIMIT))
        default_windows_json = (camp_for_cfg.active_windows_json if camp_for_cfg else "[]") or "[]"
    else:
        min_s = max(0, get_setting_int(db, "jitter_min_sec", DEFAULT_SEND_JITTER_MIN_SEC))
        max_s = max(min_s, get_setting_int(db, "jitter_max_sec", DEFAULT_SEND_JITTER_MAX_SEC))
        default_daily_limit_per_account = max(1, get_setting_int(db, "daily_limit_default", DEFAULT_DAILY_LIMIT))
        default_windows_json = "[]"

    target_states = []
    for t in targets:
        if campaign_id_for_tasks is not None and not t.campaign_id:
            t.campaign_id = int(campaign_id_for_tasks)
        required = max(1, target_required_senders(db, t))
        active_assigned_accounts = {
            int(x[0]) for x in db.query(Task.account_id).filter(
                Task.target_id == t.id,
                Task.task_type == "send_request",
                Task.status.in_(ACTIVE_SEND_TASK_STATUSES),
            ).distinct().all()
        }
        if repeat_daily:
            # Daily repeat: cap is per local day.
            # Next day planner should be able to pick next senders again.
            done_today_assigned_accounts = {
                int(x[0]) for x in db.query(Task.account_id).filter(
                    Task.target_id == t.id,
                    Task.task_type == "send_request",
                    Task.status == TaskStatus.DONE.value,
                    Task.completed_at >= day_start_utc,
                    Task.completed_at < day_end_utc,
                ).distinct().all()
            }
            assigned_accounts = active_assigned_accounts | done_today_assigned_accounts
            skipped_accounts = _precheck_skipped_sender_ids_for_target(db, int(t.id), camp_for_cfg, now)
        else:
            # Non-repeat mode: lifetime cap for sender uniqueness per target.
            assigned_accounts = {
                int(x[0]) for x in db.query(Task.account_id).filter(
                    Task.target_id == t.id,
                    Task.task_type == "send_request",
                    Task.status != TaskStatus.CANCELLED.value,
                ).distinct().all()
            }
            skipped_accounts = _precheck_skipped_sender_ids_for_target(db, int(t.id), camp_for_cfg, now)
        missing = required - len(assigned_accounts)
        if missing <= 0:
            if t.status == TargetStatus.NEW.value:
                t.status = TargetStatus.PENDING.value
            continue
        target_states.append(
            {
                "target": t,
                "assigned": assigned_accounts,
                "skipped": skipped_accounts,
                "required": required,
                "missing": missing,
            }
        )

    switch_min_sec = max(0, get_setting_int(db, "sender_switch_min_sec", 0))
    switch_max_sec = max(switch_min_sec, get_setting_int(db, "sender_switch_max_sec", 0))
    switch_avg_sec = (switch_min_sec + switch_max_sec) // 2 if switch_max_sec > 0 else 0

    active_for_phase = db.query(Account.id).filter(
        Account.enabled == True,
        Account.status == AccountStatus.ACTIVE.value,
        Account.epic_account_id.isnot(None),
        Account.device_id.isnot(None),
        Account.device_secret.isnot(None),
    ).order_by(Account.id.asc()).all()
    account_phase = {}
    for idx, row in enumerate(active_for_phase):
        account_phase[int(row[0])] = int(idx * switch_avg_sec) if switch_avg_sec > 0 else 0

    def _assign_task(
        acc: Account,
        state: dict,
        seq_rank: int = 0,
        seq_block_sec: int = 0,
        seq_extra_sec: int = 0,
        scheduled_override: Optional[datetime] = None,
    ) -> bool:
        t = state["target"]
        assigned_accounts = state["assigned"]
        skipped_accounts = state["skipped"]
        if int(acc.id) in skipped_accounts:
            return False
        if int(acc.id) in assigned_accounts or int(state["missing"]) <= 0:
            return False
        existing_active_pair = db.query(Task.id).filter(
            Task.task_type == "send_request",
            Task.target_id == int(t.id),
            Task.account_id == int(acc.id),
            Task.status.in_(ACTIVE_SEND_TASK_STATUSES),
        ).first()
        if existing_active_pair:
            if int(acc.id) not in assigned_accounts:
                assigned_accounts.add(int(acc.id))
                state["missing"] = max(0, int(state["missing"]) - 1)
            return False
        daily_limit_per_account = default_daily_limit_per_account if campaign_id_for_tasks is not None else max(1, int(acc.daily_limit or DEFAULT_DAILY_LIMIT))
        windows_json = default_windows_json if campaign_id_for_tasks is not None else (acc.active_windows_json or "[]")
        if scheduled_override is None:
            scheduled_for = _scheduled_send_time_spread(
                db,
                now,
                acc,
                campaign_id_for_tasks,
                daily_limit_per_account,
                min_s,
                max_s,
                windows_json,
            )
        else:
            scheduled_for = scheduled_override
        if dispatch_mode == "sequential" and seq_block_sec > 0 and seq_rank > 0:
            scheduled_for = scheduled_for + timedelta(seconds=int(seq_rank * seq_block_sec))
        if dispatch_mode == "sequential" and seq_extra_sec > 0:
            scheduled_for = scheduled_for + timedelta(seconds=int(seq_extra_sec))
        if dispatch_mode != "sequential" and switch_avg_sec > 0:
            scheduled_for = scheduled_for + timedelta(seconds=int(account_phase.get(int(acc.id), 0)))
        db.add(
            Task(
                task_type="send_request",
                status=TaskStatus.QUEUED.value,
                campaign_id=campaign_id_for_tasks,
                account_id=acc.id,
                target_id=t.id,
                scheduled_for=scheduled_for,
                max_attempts=t.max_attempts,
            )
        )
        t.status = TargetStatus.PENDING.value
        t.first_attempt_at = t.first_attempt_at or now
        state["missing"] = int(state["missing"]) - 1
        assigned_accounts.add(int(acc.id))
        reserved_by_account[acc.id] = int(reserved_by_account.get(acc.id, 0)) + 1
        return True

    if dispatch_mode == "sequential" and target_states:
        sender_cursor_key = f"sender_seq_cursor_campaign_{int(campaign_id_for_tasks or 0)}"
        cursor = max(0, get_setting_int(db, sender_cursor_key, 0))
        accs = db.query(Account).filter(
            Account.enabled == True,
            Account.status == AccountStatus.ACTIVE.value,
            Account.epic_account_id.isnot(None),
            Account.device_id.isnot(None),
            Account.device_secret.isnot(None),
        ).order_by(Account.id.asc()).all()
        valid_accs = []
        for a in accs:
            if a.warmup_until and now < a.warmup_until:
                continue
            windows = json.loads(a.active_windows_json or "[]")
            if not is_in_window_for_account(db, windows, now):
                continue
            valid_accs.append(a)
        if valid_accs:
            start_idx = int(cursor % len(valid_accs))
            ordered_accs = valid_accs[start_idx:] + valid_accs[:start_idx]
            sender_pick_mode = (
                get_campaign_sender_pick_mode(db, int(campaign_id_for_tasks))
                if campaign_id_for_tasks
                else "ordered"
            )
            if sender_pick_mode == "random":
                random.shuffle(ordered_accs)
            last_used_shift = 0
            campaign_send_mode = get_campaign_send_mode(db, int(campaign_id_for_tasks or 0)) if campaign_id_for_tasks else "sender_first"
            # Strict layered pacing:
            # sender#1 -> all eligible targets, then sender#2 -> all eligible targets, etc.
            # Start gap between sender blocks is auto-derived from today's window and daily limit.
            intervals = _campaign_daily_intervals_utc(db, default_windows_json, now)
            total_window_sec = _window_total_seconds(intervals)
            elapsed_now_sec = _window_elapsed_seconds(intervals, now)
            per_sender_daily = max(1, int(default_daily_limit_per_account))
            senders_count = max(1, len(ordered_accs))
            cycle_capacity = max(1, senders_count * per_sender_daily)
            auto_switch_gap_sec = max(1, total_window_sec // cycle_capacity)
            if switch_max_sec > 0:
                auto_switch_gap_sec = max(auto_switch_gap_sec, switch_min_sec)
                auto_switch_gap_sec = min(auto_switch_gap_sec, switch_max_sec)

            if campaign_send_mode == "target_first":
                # Target-first mode: nick#1 <- all senders, then nick#2 <- all senders.
                target_count = max(1, len(target_states))
                auto_target_gap_sec = max(1, total_window_sec // target_count)
                if switch_max_sec > 0:
                    auto_target_gap_sec = max(auto_target_gap_sec, switch_min_sec)
                    auto_target_gap_sec = min(auto_target_gap_sec, switch_max_sec)
                used_target_blocks = 0
                for state in target_states:
                    if int(state["missing"]) <= 0:
                        continue
                    block_start_offset = max(0, elapsed_now_sec + int(used_target_blocks * auto_target_gap_sec))
                    block_cursor_sec = 0
                    assigned_this_target = 0
                    for shift, acc in enumerate(ordered_accs):
                        if int(state["missing"]) <= 0:
                            break
                        if campaign_id_for_tasks is not None:
                            account_daily_cap = per_sender_daily
                        else:
                            account_daily_cap = max(1, int(acc.daily_limit or DEFAULT_DAILY_LIMIT))
                        effective = int(acc.today_sent or 0) + int(reserved_by_account.get(acc.id, 0))
                        if effective >= account_daily_cap:
                            continue
                        if int(acc.id) in state["assigned"]:
                            continue
                        if int(acc.id) in state["skipped"]:
                            continue
                        offset = min(max(0, total_window_sec - 1), max(0, block_start_offset + block_cursor_sec))
                        scheduled_seq = _map_offset_to_intervals(intervals, int(offset))
                        if scheduled_seq < now:
                            scheduled_seq = now + timedelta(seconds=1)
                        if _assign_task(
                            acc,
                            state,
                            seq_rank=int(shift),
                            seq_block_sec=0,
                            seq_extra_sec=0,
                            scheduled_override=scheduled_seq,
                        ):
                            created += 1
                            assigned_this_target += 1
                            step_jitter = random.randint(int(min_s), int(max_s)) if int(max_s) >= int(min_s) else int(min_s)
                            block_cursor_sec += max(1, int(step_jitter))
                            last_used_shift = int(shift)
                    if assigned_this_target > 0:
                        used_target_blocks += 1
            else:
                # Sender-first mode: sender#1 -> all nicks, sender#2 -> all nicks.
                used_sender_blocks = 0
                for shift, acc in enumerate(ordered_accs):
                    if all(int(st["missing"]) <= 0 for st in target_states):
                        break
                    if campaign_id_for_tasks is not None:
                        account_daily_cap = per_sender_daily
                    else:
                        account_daily_cap = max(1, int(acc.daily_limit or DEFAULT_DAILY_LIMIT))
                    effective = int(acc.today_sent or 0) + int(reserved_by_account.get(acc.id, 0))
                    if effective >= account_daily_cap:
                        continue

                    sender_capacity = max(0, account_daily_cap - effective)
                    if sender_capacity <= 0:
                        continue

                    block_start_offset = max(0, elapsed_now_sec + int(used_sender_blocks * auto_switch_gap_sec))
                    block_cursor_sec = 0
                    assigned_this_sender = 0

                    for state in target_states:
                        if assigned_this_sender >= sender_capacity:
                            break
                        if int(state["missing"]) <= 0:
                            continue
                        if int(acc.id) in state["assigned"]:
                            continue
                        if int(acc.id) in state["skipped"]:
                            continue

                        offset = min(max(0, total_window_sec - 1), max(0, block_start_offset + block_cursor_sec))
                        scheduled_seq = _map_offset_to_intervals(intervals, int(offset))
                        if scheduled_seq < now:
                            scheduled_seq = now + timedelta(seconds=1)

                        if _assign_task(
                            acc,
                            state,
                            seq_rank=int(shift),
                            seq_block_sec=0,
                            seq_extra_sec=0,
                            scheduled_override=scheduled_seq,
                        ):
                            created += 1
                            assigned_this_sender += 1
                            step_jitter = random.randint(int(min_s), int(max_s)) if int(max_s) >= int(min_s) else int(min_s)
                            block_cursor_sec += max(1, int(step_jitter))
                            last_used_shift = int(shift)

                    if assigned_this_sender > 0:
                        used_sender_blocks += 1
            next_cursor = (start_idx + last_used_shift + 1) % len(valid_accs)
            set_setting(db, sender_cursor_key, str(int(next_cursor)))
    else:
        for state in target_states:
            t = state["target"]
            assigned_accounts = state["assigned"]
            required = int(state["required"])
            for _ in range(int(state["missing"])):
                excluded_ids = set(int(x) for x in assigned_accounts) | set(int(x) for x in state["skipped"])
                acc = pick_best_account_with_reservations_excluded(db, reserved_by_account, excluded_ids)
                if not acc or int(acc.id) in assigned_accounts:
                    log_event_in_db(
                        db,
                        "warning",
                        f"⚠️ Не хватило уникальных аккаунтов для цели {t.username} "
                        f"(цель #{(campaign_id_for_tasks if campaign_id_for_tasks is not None else 'current')}): "
                        f"нужно {required}, назначено {len(assigned_accounts)}",
                    )
                    break
                if _assign_task(acc, state):
                    created += 1

    for state in target_states:
        if int(state["missing"]) > 0:
            t = state["target"]
            log_event_in_db(
                db,
                "warning",
                f"⚠️ Не хватило уникальных аккаунтов для цели {t.username} "
                f"(цель #{(campaign_id_for_tasks if campaign_id_for_tasks is not None else 'current')}): "
                f"нужно {int(state['required'])}, назначено {len(state['assigned'])}",
            )
    db.commit()
    log_event_in_db(db, "info", f"⚙️ create_tasks_for_new_targets: готово, создано {created}")
    return created


def campaign_sent_today_for_account(db, account_id: int, campaign_id: int, now_utc: datetime) -> int:
    day_start, day_end = local_day_bounds_utc_naive(db, now_utc)
    return int(
        db.query(func.count(Task.id))
        .filter(
            Task.task_type == "send_request",
            Task.status == TaskStatus.DONE.value,
            Task.account_id == int(account_id),
            Task.campaign_id == int(campaign_id),
            Task.completed_at >= day_start,
            Task.completed_at < day_end,
        )
        .scalar()
        or 0
    )


def rebuild_campaign_send_queue(db, campaign_id: int, create_limit: int = 5000) -> tuple[int, int]:
    """
    Remove stale queued/postponed send tasks for a goal and rebuild missing tasks
    under current goal settings.
    """
    dropped_queue = int(
        db.query(Task)
        .filter(
            task_campaign_filter(db, campaign_id),
            Task.task_type == "send_request",
            Task.status.in_(QUEUED_OR_POSTPONED_TASK_STATUSES),
        )
        .delete(synchronize_session=False)
        or 0
    )
    db.commit()
    if not is_new_send_requests_enabled(db):
        return dropped_queue, 0
    effective_campaign_id = int(campaign_id) if int(campaign_id or 0) > 0 else None
    created_tasks = int(create_tasks_for_new_targets(db, limit=int(create_limit), campaign_id=effective_campaign_id) or 0)
    return dropped_queue, created_tasks


def process_tasks_job():
    """Обработать очередь задач"""

    def _inner(db):
        if not get_setting_bool(db, "processing_enabled", True):
            return 0
        processed = 0
        now = utc_now()
        tasks_query = db.query(Task).filter(
            Task.status.in_(QUEUED_OR_POSTPONED_TASK_STATUSES),
            Task.scheduled_for <= now
        ).order_by(
            case((Task.task_type == "send_request", 0), else_=1),
            Task.scheduled_for.asc(),
        )

        # Для PostgreSQL используем SKIP LOCKED, чтобы несколько worker-процессов
        # могли безопасно разбирать очередь без гонок и дублей.
        if DB_URL.startswith("postgresql"):
            tasks_query = tasks_query.with_for_update(skip_locked=True)

        tasks = tasks_query.limit(MAX_TASKS_PER_TICK).all()

        for task in tasks:
            acc = db.query(Account).filter(Account.id == task.account_id).first()
            tgt = db.query(Target).filter(Target.id == task.target_id).first()
            camp_id = int(task.campaign_id or getattr(tgt, "campaign_id", 0) or 0)
            camp = get_campaign_or_default(db, camp_id)
            # Compatibility path for old rows without campaign_id.
            no_campaign_link_mode = (
                camp_id <= 0
                and getattr(tgt, "campaign_id", None) in (None, 0)
                and task.campaign_id in (None, 0)
            )
            if not acc or not tgt:
                task.status = TaskStatus.FAILED.value
                task.last_error = "Account/target missing"
                task.completed_at = now
                processed += 1
                continue
            if camp is None and not no_campaign_link_mode:
                task.status = TaskStatus.FAILED.value
                task.last_error = "campaign_missing"
                task.completed_at = now
                processed += 1
                continue
            if camp is not None:
                task.campaign_id = int(camp.id)
            if camp is not None and not camp.enabled:
                task.status = TaskStatus.POSTPONED.value
                task.scheduled_for = now + timedelta(minutes=30)
                task.last_error = "campaign_disabled"
                continue

            if not acc.enabled or acc.status != AccountStatus.ACTIVE.value:
                task.status = TaskStatus.POSTPONED.value
                task.scheduled_for = now + timedelta(minutes=30)
                continue

            # В рабочем режиме используем только аккаунты с device_auth.
            if not (acc.epic_account_id and acc.device_id and acc.device_secret):
                task.status = TaskStatus.FAILED.value
                task.completed_at = now
                task.last_error = "missing_device_auth"
                acc.status = AccountStatus.MANUAL.value
                acc.last_error = "missing_device_auth"
                if tgt.status == TargetStatus.NEW.value:
                    tgt.status = TargetStatus.PENDING.value
                processed += 1
                continue

            if acc.warmup_until and now < acc.warmup_until:
                task.status = TaskStatus.POSTPONED.value
                task.scheduled_for = acc.warmup_until + timedelta(minutes=1)
                continue

            windows = json.loads(acc.active_windows_json or "[]")
            if not is_in_window_for_account(db, windows, now):
                task.status = TaskStatus.POSTPONED.value
                task.scheduled_for = now + timedelta(minutes=10)
                continue
            if camp is not None:
                try:
                    camp_windows = json.loads(camp.active_windows_json or "[]")
                except Exception:
                    camp_windows = []
            else:
                camp_windows = []
            if camp_windows and not is_in_window_for_account(db, camp_windows, now):
                task.status = TaskStatus.POSTPONED.value
                task.scheduled_for = now + timedelta(minutes=10)
                task.last_error = "campaign_out_of_window"
                continue

            # SEND_REQUEST
            if task.task_type == "send_request":
                req = target_required_senders(db, tgt)
                source_tag = str(task.last_error or "").strip().lower()
                is_recheck_resend = source_tag == "recheck_resend"
                is_manual_forced_cycle = source_tag == "manual_forced_cycle"
                if not is_recheck_resend:
                    new_sends_enabled = is_new_send_requests_enabled(db)
                    recheck_only_mode = is_recheck_only_mode_enabled(db)
                    if not new_sends_enabled or recheck_only_mode:
                        task.status = TaskStatus.POSTPONED.value
                        task.last_error = "new_send_requests_disabled"
                        task.scheduled_for = now + timedelta(minutes=30)
                        processed += 1
                        continue
                older_active_pair = db.query(Task.id).filter(
                    Task.task_type == "send_request",
                    Task.target_id == int(tgt.id),
                    Task.account_id == int(acc.id),
                    Task.status.in_(ACTIVE_SEND_TASK_STATUSES),
                    Task.id != int(task.id),
                ).order_by(Task.id.asc()).first()
                if older_active_pair and int(older_active_pair[0]) < int(task.id):
                    task.status = TaskStatus.CANCELLED.value
                    task.completed_at = now
                    task.last_error = "duplicate_active_pair_task"
                    processed += 1
                    continue
                if not (is_recheck_resend or is_manual_forced_cycle):
                    done_sender_ids = _done_sender_ids_for_target(db, int(tgt.id), camp, now)
                    if len(done_sender_ids) >= int(req):
                        task.status = TaskStatus.CANCELLED.value
                        task.completed_at = now
                        task.last_error = "target_sender_cap_reached"
                        processed += 1
                        continue
                    if int(acc.id) in done_sender_ids:
                        task.status = TaskStatus.CANCELLED.value
                        task.completed_at = now
                        task.last_error = "sender_already_used_for_target"
                        processed += 1
                        continue

                if not DRY_RUN and not SEND_REQUESTS_ENABLED:
                    # Never send real friend requests unless explicitly enabled.
                    task.status = TaskStatus.POSTPONED.value
                    task.last_error = "send_requests_disabled"
                    task.scheduled_for = now + timedelta(hours=6)
                    processed += 1
                    continue

                if camp is not None:
                    pace_ok, pace_next_at, pace_reason = _campaign_pacing_gate(db, camp, now)
                    if not pace_ok:
                        task.status = TaskStatus.POSTPONED.value
                        task.last_error = pace_reason
                        task.scheduled_for = pace_next_at or (now + timedelta(seconds=60))
                        continue

                if camp is not None:
                    campaign_daily_limit = _campaign_effective_daily_limit(db, camp, now)
                    campaign_today_sent = campaign_sent_today_for_account(db, acc.id, int(camp.id), now)
                else:
                    campaign_daily_limit = max(1, int(acc.daily_limit or DEFAULT_DAILY_LIMIT))
                    campaign_today_sent = 0
                if acc.today_sent >= acc.daily_limit or campaign_today_sent >= campaign_daily_limit:
                    # Если аккаунт упёрся в дневной лимит, пробуем сразу выполнить
                    # задачу другим аккаунтом (иначе "ранние" задачи на лимитных
                    # аккаунтах будут постоянно отъедать MAX_TASKS_PER_TICK и
                    # блокировать прогресс очереди).
                    alt = pick_best_account(db)
                    if not alt or alt.id == acc.id:
                        task.status = TaskStatus.POSTPONED.value
                        # Нет доступных аккаунтов прямо сейчас: отложим до следующего сброса,
                        # чтобы не перетирать MAX_TASKS_PER_TICK на одних и тех же задачах.
                        reset_at = next_daily_reset_utc(now)
                        task.scheduled_for = reset_at + timedelta(seconds=random.randint(0, 300))
                        continue

                    task.account_id = alt.id
                    acc = alt

                if not DRY_RUN:
                    send_api_cost = max(1, get_setting_int(db, "send_api_cost", DEFAULT_SEND_API_COST))
                    ok_rate, next_at, reason = enforce_api_rate_limit(db, acc, now, send_api_cost)
                    if not ok_rate:
                        task.status = TaskStatus.POSTPONED.value
                        task.last_error = f"api_rate_limit_{reason}"
                        task.scheduled_for = next_at or (now + timedelta(seconds=60))
                        continue

                task.status = TaskStatus.RUNNING.value
                task.started_at = now
                db.commit()

                proxy_url = get_proxy_for_account(db, acc.id)

                if DRY_RUN:
                    # Эмуляция отправки заявки (без запросов в Epic).
                    acc.today_sent += 1
                    acc.total_sent += 1
                    acc.last_activity_at = now
                    acc.last_error = None
                    tgt.sent_count += 1
                    tgt.last_attempt_at = now
                    tgt.attempt_count += 1
                    req = target_required_senders(db, tgt)
                    tgt.status = TargetStatus.SENT.value if int(tgt.sent_count or 0) >= req else TargetStatus.PENDING.value
                    task.status = TaskStatus.DONE.value
                    task.completed_at = now
                    db.add(
                        Task(
                            task_type="check_status",
                            status=TaskStatus.QUEUED.value,
                            campaign_id=(int(camp.id) if camp is not None else None),
                            account_id=acc.id,
                            target_id=tgt.id,
                            scheduled_for=now + timedelta(seconds=DRY_RUN_CHECK_DELAY_SEC),
                            max_attempts=5,
                        )
                    )
                    processed += 1
                    continue

                # Pre-check статуса: если уже pending/accepted, не шлём повторный POST в Epic.
                pre_status = check_friend_status_with_device(
                    login=acc.login,
                    password=acc.password,
                    target_username=tgt.username,
                    proxy_url=proxy_url,
                    epic_account_id=acc.epic_account_id,
                    device_id=acc.device_id,
                    device_secret=acc.device_secret,
                )
                if pre_status.ok and pre_status.code == "accepted":
                    task.status = TaskStatus.CANCELLED.value
                    task.completed_at = now
                    task.last_error = "precheck_accepted_skip"
                    # Already friends: this sender must not count as new coverage.
                    # Try to replace by another eligible sender.
                    done_sender_ids = _done_sender_ids_for_target(db, int(tgt.id), camp, now)
                    replaced = False
                    if len(done_sender_ids) < int(req):
                        replaced = _enqueue_replacement_send_task(
                            db,
                            target=tgt,
                            camp=camp,
                            now=now,
                            excluded_ids={int(acc.id)},
                            source_reason="replacement_after_precheck_accepted",
                        )
                    task.last_error = "precheck_accepted_skip_requeued" if replaced else "precheck_accepted_skip"
                    processed += 1
                    continue

                if pre_status.ok and pre_status.code == "pending":
                    task.status = TaskStatus.CANCELLED.value
                    task.completed_at = now
                    # Request is already pending for this sender: skip from coverage and replace sender.
                    done_sender_ids = _done_sender_ids_for_target(db, int(tgt.id), camp, now)
                    replaced = False
                    if len(done_sender_ids) < int(req):
                        replaced = _enqueue_replacement_send_task(
                            db,
                            target=tgt,
                            camp=camp,
                            now=now,
                            excluded_ids={int(acc.id)},
                            source_reason="replacement_after_precheck_pending",
                        )
                    task.last_error = "precheck_pending_skip_requeued" if replaced else "precheck_pending_skip"
                    processed += 1
                    continue

                result = send_friend_request_with_device(
                    login=acc.login,
                    password=acc.password,
                    target_username=tgt.username,
                    proxy_url=proxy_url,
                    epic_account_id=acc.epic_account_id,
                    device_id=acc.device_id,
                    device_secret=acc.device_secret,
                )

                if result.ok and result.code == "request_sent":
                    result_data = getattr(result, "data", None) or {}
                    was_idempotent = bool(result_data.get("note") == "idempotent_success")
                    if was_idempotent:
                        # Epic returned idempotent success (already friends / already pending).
                        # Do not count this as a new send for coverage; replace sender if needed.
                        task.status = TaskStatus.CANCELLED.value
                        task.completed_at = now
                        req = target_required_senders(db, tgt)
                        replaced = False
                        if not is_recheck_resend:
                            done_sender_ids = _done_sender_ids_for_target(db, int(tgt.id), camp, now)
                            if len(done_sender_ids) < int(req):
                                replaced = _enqueue_replacement_send_task(
                                    db,
                                    target=tgt,
                                    camp=camp,
                                    now=now,
                                    excluded_ids={int(acc.id)},
                                    source_reason="replacement_after_idempotent_send",
                                )
                        task.last_error = (
                            "idempotent_request_skip_requeued" if replaced else "idempotent_request_skip"
                        )
                        acc.last_activity_at = now
                        acc.last_error = None
                    else:
                        acc.today_sent += 1
                        acc.total_sent += 1
                        acc.last_activity_at = now
                        acc.last_error = None
                        tgt.sent_count += 1
                        tgt.last_attempt_at = now
                        tgt.attempt_count += 1
                        req = target_required_senders(db, tgt)
                        tgt.status = TargetStatus.SENT.value if int(tgt.sent_count or 0) >= req else TargetStatus.PENDING.value
                        task.status = TaskStatus.DONE.value
                        task.completed_at = now

                        check_task = Task(
                            task_type="check_status",
                            status=TaskStatus.QUEUED.value,
                            campaign_id=(int(camp.id) if camp is not None else None),
                            account_id=acc.id,
                            target_id=tgt.id,
                            scheduled_for=now + timedelta(hours=2),
                            max_attempts=5,
                        )
                        db.add(check_task)

                        if acc.proxy_id:
                            proxy = db.query(Proxy).filter(Proxy.id == acc.proxy_id).first()
                            if proxy:
                                proxy.success_count += 1
                else:
                    acc.total_failed += 1
                    acc.last_activity_at = now
                    acc.last_error = f"{result.code}: {result.message}"
                    task.attempt_number += 1
                    task.last_error = acc.last_error

                    if result.code == "rate_limited":
                        task.status = TaskStatus.POSTPONED.value
                        task.scheduled_for = now + timedelta(minutes=30)
                    elif result.code == "password_grant_blocked":
                        # Epic blocked password grant; this account must be re-authorized via login-link/device_auth.
                        acc.status = AccountStatus.MANUAL.value
                        acc.last_error = "password_grant_blocked_use_device_auth"
                        task.status = TaskStatus.FAILED.value
                        task.completed_at = now
                        task.last_error = acc.last_error
                        if tgt.status == TargetStatus.NEW.value:
                            tgt.status = TargetStatus.PENDING.value
                    elif result.code == "auth_failed" or "auth" in (result.code or "").lower():
                        acc.status = AccountStatus.MANUAL.value
                        task.status = TaskStatus.FAILED.value
                        task.completed_at = now
                        if tgt.status == TargetStatus.NEW.value:
                            tgt.status = TargetStatus.PENDING.value
                    elif task.attempt_number < task.max_attempts:
                        task.status = TaskStatus.POSTPONED.value
                        task.scheduled_for = now + timedelta(minutes=5 * task.attempt_number)
                    else:
                        task.status = TaskStatus.FAILED.value
                        task.completed_at = now
                        if tgt.status == TargetStatus.NEW.value:
                            tgt.status = TargetStatus.PENDING.value

                    if acc.proxy_id:
                        proxy = db.query(Proxy).filter(Proxy.id == acc.proxy_id).first()
                        if proxy:
                            proxy.failed_count += 1

                processed += 1

            # CHECK_STATUS
            elif task.task_type == "check_status":
                # Hard priority for outgoing requests:
                # while campaign has overdue send tasks, defer status checks.
                due_send_campaign = (
                    db.query(func.count(Task.id))
                    .filter(
                        Task.task_type == "send_request",
                        Task.status.in_(QUEUED_OR_POSTPONED_TASK_STATUSES),
                        Task.scheduled_for <= now,
                        task_campaign_filter(db, int(camp.id) if camp is not None else 0),
                    )
                    .scalar()
                    or 0
                )
                if int(due_send_campaign) > 0:
                    task.status = TaskStatus.POSTPONED.value
                    task.last_error = "deferred_for_send_priority"
                    task.scheduled_for = now + timedelta(minutes=10)
                    continue

                # If this sender has overdue send tasks, defer status checks so
                # API quota is spent on outgoing requests first.
                due_send_same_acc = (
                    db.query(func.count(Task.id))
                    .filter(
                        Task.task_type == "send_request",
                        Task.status.in_(QUEUED_OR_POSTPONED_TASK_STATUSES),
                        Task.account_id == int(acc.id),
                        Task.scheduled_for <= now,
                        task_campaign_filter(db, int(camp.id) if camp is not None else 0),
                    )
                    .scalar()
                    or 0
                )
                if int(due_send_same_acc) > 0:
                    task.status = TaskStatus.POSTPONED.value
                    task.last_error = "deferred_for_send_backlog"
                    task.scheduled_for = now + timedelta(minutes=10)
                    continue

                if not DRY_RUN:
                    check_api_cost = max(1, get_setting_int(db, "check_api_cost", DEFAULT_CHECK_API_COST))
                    ok_rate, next_at, reason = enforce_api_rate_limit(db, acc, now, check_api_cost)
                    if not ok_rate:
                        task.status = TaskStatus.POSTPONED.value
                        task.last_error = f"api_rate_limit_{reason}"
                        task.scheduled_for = next_at or (now + timedelta(seconds=60))
                        continue

                task.status = TaskStatus.RUNNING.value
                task.started_at = now
                db.commit()

                proxy_url = get_proxy_for_account(db, acc.id)

                if DRY_RUN:
                    # Детерминированно (по имени) решаем accepted/rejected.
                    # Это делает прогон повторяемым и позволяет тестировать ретраи/лимиты.
                    rate = max(0.0, min(1.0, DRY_RUN_ACCEPT_RATE))
                    h = sum(ord(c) for c in (tgt.username or "")) % 100
                    accepted = h < int(rate * 100)
                    if accepted:
                        tgt.status = TargetStatus.ACCEPTED.value
                        if not _pair_was_accepted(db, int(tgt.id), int(acc.id)):
                            tgt.accepted_count += 1
                        acc.total_accepted += 1
                        task.last_error = "friend_status:accepted"
                    else:
                        tgt.status = TargetStatus.REJECTED.value
                        task.last_error = "friend_status:rejected"
                    task.status = TaskStatus.DONE.value
                    task.completed_at = now
                    processed += 1
                    continue

                result = check_friend_status_with_device(
                    login=acc.login,
                    password=acc.password,
                    target_username=tgt.username,
                    proxy_url=proxy_url,
                    epic_account_id=acc.epic_account_id,
                    device_id=acc.device_id,
                    device_secret=acc.device_secret,
                )

                if result.ok and result.code in ("accepted", "rejected", "pending"):
                    if result.code == "accepted":
                        tgt.status = TargetStatus.ACCEPTED.value
                        if not _pair_was_accepted(db, int(tgt.id), int(acc.id)):
                            tgt.accepted_count += 1
                        acc.total_accepted += 1
                        task.last_error = "friend_status:accepted"
                    elif result.code == "pending":
                        req = target_required_senders(db, tgt)
                        tgt.status = TargetStatus.SENT.value if int(tgt.sent_count or 0) >= req else TargetStatus.PENDING.value
                        task.last_error = "friend_status:pending"
                    else:
                        # Friendship/request disappeared.
                        # If it was previously accepted, schedule resend from the same sender.
                        if int(tgt.accepted_count or 0) > 0:
                            has_active_send = db.query(Task.id).filter(
                                Task.task_type == "send_request",
                                Task.account_id == int(acc.id),
                                Task.target_id == int(tgt.id),
                                Task.status.in_(ACTIVE_SEND_TASK_STATUSES),
                            ).first()
                            if not has_active_send:
                                jitter_min = max(0, int(camp.jitter_min_sec or DEFAULT_SEND_JITTER_MIN_SEC)) if camp is not None else DEFAULT_SEND_JITTER_MIN_SEC
                                jitter_max = max(jitter_min, int(camp.jitter_max_sec or DEFAULT_SEND_JITTER_MAX_SEC)) if camp is not None else DEFAULT_SEND_JITTER_MAX_SEC
                                resend_task = Task(
                                    task_type="send_request",
                                    status=TaskStatus.QUEUED.value,
                                    campaign_id=(int(camp.id) if camp is not None else None),
                                    account_id=int(acc.id),
                                    target_id=int(tgt.id),
                                    scheduled_for=now + timedelta(seconds=random.randint(int(jitter_min), int(jitter_max))),
                                    max_attempts=5,
                                    last_error="recheck_resend",
                                )
                                db.add(resend_task)
                            tgt.status = TargetStatus.PENDING.value
                            tgt.last_error = "friendship_lost_requeue_send"
                        else:
                            tgt.status = TargetStatus.REJECTED.value
                        task.last_error = "friend_status:rejected"
                    task.status = TaskStatus.DONE.value
                    task.completed_at = now
                else:
                    task.attempt_number += 1
                    if task.attempt_number < task.max_attempts:
                        task.status = TaskStatus.POSTPONED.value
                        task.scheduled_for = now + timedelta(hours=6)
                    else:
                        task.status = TaskStatus.FAILED.value
                        task.completed_at = now

                processed += 1

        db.commit()
        return processed

    processed = db_exec(_inner)
    if processed > 0:
        log_event("info", f"✅ [{WORKER_INSTANCE}] Обработано: {processed} задач")

def verify_accounts_health_job():
    """Проверить здоровье аккаунтов"""
    if DRY_RUN:
        return

    def _inner(db):
        accs = db.query(Account).filter(
            Account.enabled == True,
            Account.status == AccountStatus.ACTIVE.value,
            Account.epic_account_id.isnot(None),
            Account.device_id.isnot(None),
            Account.device_secret.isnot(None),
        ).limit(10).all()
        updates = 0
        for a in accs:
            proxy_url = get_proxy_for_account(db, a.id)
            result = verify_account_health_with_device(
                login=a.login,
                password=a.password,
                proxy_url=proxy_url,
                epic_account_id=a.epic_account_id,
                device_id=a.device_id,
                device_secret=a.device_secret,
            )
            if not result.ok:
                # Reduce noise: do not log password grant issues here; mark account for manual re-auth instead.
                if result.code != "password_grant_blocked":
                    log_event_in_tx(
                        db,
                        "warning",
                        f"health_check_failed acc#{a.id} login={a.login} code={result.code} message={result.message}",
                    )
                if result.code == "account_banned":
                    a.status = AccountStatus.BANNED.value
                    a.last_error = f"Забанен: {result.message}"
                    updates += 1
                elif result.code == "password_grant_blocked":
                    a.status = AccountStatus.MANUAL.value
                    a.last_error = "password_grant_blocked_use_device_auth"
                else:
                    a.last_error = result.message
        db.commit()
        return updates

    updates = db_exec(_inner)
    if updates > 0:
        log_event("warning", f"⚠️ Забанено/заблокировано: {updates} аккаунтов")


def refresh_accounts_display_names_job(limit: int = 0) -> tuple[int, int, int]:
    """
    Обновить epic_display_name у аккаунтов через Epic API.
    limit=0 -> без лимита.
    Возвращает: (checked, updated, failed)
    """
    checked = 0
    updated = 0
    failed = 0

    def _load_accounts(db):
        q = db.query(Account).filter(
            Account.enabled == True,
            Account.status == AccountStatus.ACTIVE.value,
            Account.epic_account_id.isnot(None),
            Account.device_id.isnot(None),
            Account.device_secret.isnot(None),
        ).order_by(Account.id.asc())
        if int(limit or 0) > 0:
            q = q.limit(int(limit))
        return [
            {
                "id": int(a.id),
                "login": str(a.login),
                "password": str(a.password),
                "epic_account_id": str(a.epic_account_id or ""),
                "device_id": str(a.device_id or ""),
                "device_secret": str(a.device_secret or ""),
            }
            for a in q.all()
        ]

    accounts = db_exec(_load_accounts)
    for item in accounts:
        checked += 1
        proxy_url = db_exec(lambda db: get_proxy_for_account(db, int(item["id"])))
        result = verify_account_health_with_device(
            login=item["login"],
            password=item["password"],
            proxy_url=proxy_url,
            epic_account_id=item["epic_account_id"],
            device_id=item["device_id"],
            device_secret=item["device_secret"],
        )
        if not result.ok:
            failed += 1
            continue
        display_name = str((result.data or {}).get("display_name") or "").strip()
        if not display_name:
            continue

        def _save_name(db):
            acc = db.query(Account).filter(Account.id == int(item["id"])).first()
            if not acc:
                return False
            prev = str(acc.epic_display_name or "").strip()
            if prev == display_name:
                return False
            acc.epic_display_name = display_name
            db.commit()
            return True

        if db_exec(_save_name):
            updated += 1

    return int(checked), int(updated), int(failed)


def process_nickname_change_tasks_job():
    """Обработать очередь задач массовой смены ников."""

    def _inner(db):
        if not get_setting_bool(db, "processing_enabled", True):
            return 0
        now = utc_now()
        processed = 0
        q = db.query(NicknameChangeTask).filter(
            NicknameChangeTask.status.in_([
                NicknameChangeStatus.QUEUED.value,
                NicknameChangeStatus.POSTPONED.value,
            ]),
            NicknameChangeTask.scheduled_for <= now,
        ).order_by(NicknameChangeTask.scheduled_for.asc(), NicknameChangeTask.id.asc())
        if DB_URL.startswith("postgresql"):
            q = q.with_for_update(skip_locked=True)
        jobs = q.limit(max(1, min(20, MAX_TASKS_PER_TICK))).all()

        for job in jobs:
            acc = db.query(Account).filter(Account.id == int(job.account_id)).first()
            if not acc:
                job.status = NicknameChangeStatus.FAILED.value
                job.last_error = "account_missing"
                job.completed_at = now
                processed += 1
                continue
            if not acc.enabled or acc.status != AccountStatus.ACTIVE.value:
                job.status = NicknameChangeStatus.POSTPONED.value
                job.last_error = "account_not_active"
                job.scheduled_for = now + timedelta(minutes=30)
                processed += 1
                continue
            if not (acc.epic_account_id and acc.device_id and acc.device_secret):
                job.status = NicknameChangeStatus.FAILED.value
                job.last_error = "missing_device_auth"
                job.completed_at = now
                processed += 1
                continue

            ok_nick, nick_reason = validate_requested_nickname(job.requested_nick)
            if not ok_nick:
                job.status = NicknameChangeStatus.FAILED.value
                job.last_error = f"invalid_nickname_format: {nick_reason}"
                job.completed_at = now
                processed += 1
                continue

            if not DRY_RUN:
                api_cost = max(1, get_setting_int(db, "nickname_change_api_cost", 2))
                ok_rate, next_at, reason = enforce_api_rate_limit(db, acc, now, api_cost)
                if not ok_rate:
                    job.status = NicknameChangeStatus.POSTPONED.value
                    job.last_error = f"api_rate_limit_{reason}"
                    job.scheduled_for = next_at or (now + timedelta(seconds=60))
                    processed += 1
                    continue

            job.status = NicknameChangeStatus.RUNNING.value
            job.started_at = now
            db.commit()

            try:
                if DRY_RUN:
                    acc.epic_display_name = str(job.requested_nick)
                    acc.last_activity_at = now
                    acc.last_error = None
                    job.final_nick = str(job.requested_nick)
                    job.status = NicknameChangeStatus.DONE.value
                    job.completed_at = now
                    processed += 1
                    continue

                proxy_url = get_proxy_for_account(db, int(acc.id))
                result = change_display_name_with_device(
                    login=acc.login,
                    password=acc.password,
                    new_display_name=str(job.requested_nick),
                    proxy_url=proxy_url,
                    epic_account_id=acc.epic_account_id,
                    device_id=acc.device_id,
                    device_secret=acc.device_secret,
                )
                if result.ok:
                    applied = str(((result.data or {}).get("display_name") if result.data else "") or "").strip()
                    final_nick = applied or str(job.requested_nick)
                    acc.epic_display_name = final_nick
                    acc.last_activity_at = now
                    acc.last_error = None
                    job.final_nick = final_nick
                    job.status = NicknameChangeStatus.DONE.value
                    job.completed_at = now
                    job.last_error = None
                else:
                    job.attempt_number = int(job.attempt_number or 0) + 1
                    err = f"{result.code}: {result.message}"
                    job.last_error = err[:500]
                    acc.last_error = err[:500]
                    if result.code == "rate_limited":
                        job.status = NicknameChangeStatus.POSTPONED.value
                        job.scheduled_for = now + timedelta(minutes=30)
                    elif result.code == "nickname_taken":
                        variants = _nickname_fallback_candidates(str(job.requested_nick))
                        next_nick = variants[0] if variants else ""
                        if next_nick and int(job.attempt_number or 0) < int(job.max_attempts or 3):
                            job.requested_nick = next_nick
                            job.status = NicknameChangeStatus.POSTPONED.value
                            job.scheduled_for = now + timedelta(minutes=5)
                            job.last_error = f"nickname_taken_retry:{next_nick}"
                        else:
                            job.status = NicknameChangeStatus.FAILED.value
                            job.completed_at = now
                    elif result.code in {"nickname_cooldown", "invalid_nickname"}:
                        job.status = NicknameChangeStatus.FAILED.value
                        job.completed_at = now
                    elif result.code == "password_grant_blocked":
                        acc.status = AccountStatus.MANUAL.value
                        acc.last_error = "password_grant_blocked_use_device_auth"
                        job.status = NicknameChangeStatus.FAILED.value
                        job.completed_at = now
                    elif result.code == "auth_failed" or "auth" in str(result.code or "").lower():
                        acc.status = AccountStatus.MANUAL.value
                        job.status = NicknameChangeStatus.FAILED.value
                        job.completed_at = now
                    elif int(job.attempt_number or 0) < int(job.max_attempts or 3):
                        job.status = NicknameChangeStatus.POSTPONED.value
                        job.scheduled_for = now + timedelta(minutes=10 * int(job.attempt_number or 1))
                    else:
                        job.status = NicknameChangeStatus.FAILED.value
                        job.completed_at = now
                processed += 1
            except Exception as e:
                job.attempt_number = int(job.attempt_number or 0) + 1
                job.last_error = f"exception: {str(e)[:300]}"
                if int(job.attempt_number or 0) < int(job.max_attempts or 3):
                    job.status = NicknameChangeStatus.POSTPONED.value
                    job.scheduled_for = now + timedelta(minutes=10)
                else:
                    job.status = NicknameChangeStatus.FAILED.value
                    job.completed_at = now
                processed += 1

        db.commit()
        return int(processed)

    processed = db_exec(_inner)
    if processed > 0:
        log_event("info", f"📝 Обработано задач смены ников: {processed}")
    return int(processed or 0)


# ============================================================
# TELEGRAM UI (anti-spam)
# ============================================================

bot = telebot.TeleBot(BOT_TOKEN)

# chat_id -> {"screen_msg_id": int}
CHAT_UI = {}
CHAT_UI_LOCK = threading.Lock()

# (chat_id, user_id) -> prompt_msg_id
STEP_PROMPTS = {}
STEP_LOCK = threading.Lock()

# chat_id -> set(message_id) для коротких уведомлений notify()
TRANSIENT_MESSAGES = {}
TRANSIENT_LOCK = threading.Lock()

DEVICE_AUTH_TIMEOUT_SEC = int(os.getenv("DEVICE_AUTH_TIMEOUT_SEC", "900"))
PENDING_DEVICE = {}
PENDING_DEVICE_LOCK = threading.Lock()
REL_ACTION_JOBS = {}
REL_ACTION_JOBS_LOCK = threading.Lock()
ALLOW_SHOW_PASSWORD = os.getenv("ALLOW_SHOW_PASSWORD", "0").strip().lower() in {"1", "true", "yes"}
AUTH_OPERATOR_IDS_SETTING_KEY = "auth_operator_ids"


def _load_auth_operator_ids(db) -> set[int]:
    raw = str(get_setting(db, AUTH_OPERATOR_IDS_SETTING_KEY, "") or "")
    ids: set[int] = set()
    for part in raw.replace(";", ",").replace(" ", ",").split(","):
        p = part.strip()
        if not p:
            continue
        try:
            ids.add(int(p))
        except Exception:
            continue
    return ids


def _save_auth_operator_ids(db, ids: set[int]):
    normalized = ",".join(str(x) for x in sorted({int(v) for v in ids if int(v) > 0}))
    set_setting(db, AUTH_OPERATOR_IDS_SETTING_KEY, normalized)


def get_auth_operator_ids() -> set[int]:
    return db_exec(lambda db: _load_auth_operator_ids(db))


def is_auth_operator(user_id: int) -> bool:
    try:
        uid = int(user_id)
    except Exception:
        return False
    if is_admin(uid):
        return False
    return uid in get_auth_operator_ids()


def _can_use_bot(user_id: int) -> bool:
    return is_admin(user_id) or is_auth_operator(user_id)


def _can_use_callback(user_id: int, data: str) -> bool:
    if is_admin(user_id):
        return True
    if not is_auth_operator(user_id):
        return False
    d = str(data or "")
    if d in {"acc_device_auto", "acc_list"}:
        return True
    if d.startswith("act:"):
        code = d.split(":", 1)[1]
        return code in {
            "acc_list",
            "acc_device",
            "acc_prev",
            "acc_next",
            "acc_search",
            "back",
            "back_main",
            "main_accounts",
        }
    return d.startswith("acc_device_cancel:") or d.startswith("acc_device_show_login:") or d.startswith("acc_device_show_pass:")


def admin_only(fn):
    @wraps(fn)
    def wrapper(message):
        if not is_admin(message.from_user.id):
            _safe_reply(message, "❌ Доступ запрещён")
            return
        try:
            return fn(message)
        except Exception as e:
            log_event("error", f"ui_handler_error msg text={getattr(message,'text',None)}: {e}")
            _safe_send(message.chat.id, f"❌ Ошибка: {str(e)[:180]}", parse_mode=None)
            return
    return wrapper


def user_access_only(fn):
    @wraps(fn)
    def wrapper(message):
        if not _can_use_bot(message.from_user.id):
            _safe_reply(message, "❌ Доступ запрещён")
            return
        try:
            return fn(message)
        except Exception as e:
            log_event("error", f"ui_handler_error msg text={getattr(message,'text',None)}: {e}")
            _safe_send(message.chat.id, f"❌ Ошибка: {str(e)[:180]}", parse_mode=None)
            return
    return wrapper


def admin_only_call(fn):
    @wraps(fn)
    def wrapper(call):
        if not _can_use_callback(call.from_user.id, getattr(call, "data", "")):
            _safe_answer_callback(call, text="Нет доступа", show_alert=True)
            return
        return fn(call)
    return wrapper


def _is_inline_markup(markup) -> bool:
    return isinstance(markup, types.InlineKeyboardMarkup)


def _safe_answer_callback(call, text: str | None = None, show_alert: bool = False):
    try:
        if text:
            bot.answer_callback_query(call.id, text=text, show_alert=show_alert)
        else:
            bot.answer_callback_query(call.id)
    except Exception as e:
        logger.debug(f"answer_callback_query failed data={getattr(call, 'data', None)}: {e}")


def _safe_reply(message, text: str):
    try:
        bot.reply_to(message, text)
    except Exception as e:
        logger.debug(f"reply_to failed chat_id={getattr(getattr(message, 'chat', None), 'id', None)}: {e}")


def _safe_send(chat_id: int, text: str, **kwargs):
    try:
        bot.send_message(chat_id, text, **kwargs)
    except Exception as e:
        logger.debug(f"send_message failed chat_id={chat_id}: {e}")

def _safe_delete(chat_id: int, message_id: int):
    try:
        bot.delete_message(chat_id, message_id)
    except Exception as e:
        # This is very noisy in practice (race conditions, already deleted, etc.).
        # Keep it out of DB logs; debug is enough for troubleshooting.
        logger.debug(f"delete_message failed chat_id={chat_id} msg_id={message_id}: {e}")


def _safe_delete_user_message(message):
    # В личке Telegram обычно не даёт боту удалять чужие сообщения; в группе может дать (если права).
    try:
        bot.delete_message(message.chat.id, message.message_id)
    except Exception as e:
        logger.debug(f"delete_user_message failed chat_id={message.chat.id} msg_id={message.message_id}: {e}")


def _track_transient(chat_id: int, message_id: int):
    with TRANSIENT_LOCK:
        ids = TRANSIENT_MESSAGES.get(chat_id)
        if ids is None:
            ids = set()
            TRANSIENT_MESSAGES[chat_id] = ids
        ids.add(message_id)


def _untrack_transient(chat_id: int, message_id: int):
    with TRANSIENT_LOCK:
        ids = TRANSIENT_MESSAGES.get(chat_id)
        if not ids:
            return
        ids.discard(message_id)
        if not ids:
            TRANSIENT_MESSAGES.pop(chat_id, None)


def _delete_tracked_transient(chat_id: int, message_id: int):
    _safe_delete(chat_id, message_id)
    _untrack_transient(chat_id, message_id)


def cleanup_transient_messages(chat_id: int):
    """Удалить все активные notify-сообщения в чате."""
    with TRANSIENT_LOCK:
        ids = list(TRANSIENT_MESSAGES.get(chat_id) or [])
        TRANSIENT_MESSAGES.pop(chat_id, None)

    for mid in ids:
        _safe_delete(chat_id, mid)


def notify(chat_id: int, text: str, ttl_sec: int = 40, **kwargs):
    """Короткое уведомление, которое само удалится."""
    msg = bot.send_message(chat_id, text, **kwargs)
    _track_transient(chat_id, msg.message_id)

    timer = threading.Timer(ttl_sec, _delete_tracked_transient, args=(chat_id, msg.message_id))
    timer.daemon = True
    timer.start()
    return msg

def show_target_senders_page(chat_id: int, target_id: int, page: int = 1):
    selected_campaign_id = get_selected_campaign_id(chat_id)

    try:
        page_size = int(SENDERS_PAGE_SIZE)
    except Exception:
        page_size = 15

    def _sender_icon(target_status: str, sender_status: str, has_done_check: bool) -> str:
        if sender_status == "friend_status:accepted":
            return "✅"
        if sender_status == "friend_status:rejected":
            return "❌"
        if sender_status == "friend_status:pending":
            return "⏳"
        if has_done_check and target_status == TargetStatus.ACCEPTED.value:
            return "✅"
        if has_done_check and target_status == TargetStatus.REJECTED.value:
            return "❌"
        return "⏳"

    def _inner(db):
        filt = target_campaign_filter(db, selected_campaign_id)
        tgt = db.query(Target).filter(Target.id == int(target_id), filt).first()
        if not tgt:
            return None

        required = int(getattr(tgt, "required_senders", 0) or 0) or get_setting_int(
            db, "target_senders_count", DEFAULT_TARGET_SENDERS_COUNT
        )

        total = int(
            db.query(func.count(func.distinct(Task.account_id)))
            .filter(
                Task.target_id == tgt.id,
                Task.task_type == "send_request",
                Task.status == TaskStatus.DONE.value,
            )
            .scalar()
            or 0
        )

        pages = max(1, math.ceil(total / page_size)) if total else 1
        page_clamped = max(1, min(int(page), pages))

        rows = (
            db.query(
                Task.account_id.label("aid"),
                func.max(Task.completed_at).label("last_sent_at"),
                func.max(Task.id).label("last_task_id"),
            )
            .filter(
                Task.target_id == tgt.id,
                Task.task_type == "send_request",
                Task.status == TaskStatus.DONE.value,
            )
            .group_by(Task.account_id)
            .order_by(
                func.max(Task.completed_at).desc().nullslast(),
                func.max(Task.id).desc(),
            )
            .offset((page_clamped - 1) * page_size)
            .limit(page_size)
            .all()
        )
        account_ids = [int(r.aid) for r in rows]

        acc_map = {}
        if account_ids:
            acc_rows = (
                db.query(Account.id, Account.login, Account.epic_display_name)
                .filter(Account.id.in_(account_ids))
                .all()
            )
            for aid, login, disp in acc_rows:
                acc_map[int(aid)] = (str(login), (disp or "").strip())

        sender_status_map: dict[int, str] = {}
        check_done_ids: set[int] = set()
        if account_ids:
            check_rows = (
                db.query(Task.account_id, Task.last_error, Task.completed_at, Task.id)
                .filter(
                    Task.target_id == tgt.id,
                    Task.account_id.in_(account_ids),
                    Task.task_type == "check_status",
                    Task.status == TaskStatus.DONE.value,
                )
                .order_by(
                    Task.account_id.asc(),
                    Task.completed_at.desc().nullslast(),
                    Task.id.desc(),
                )
                .all()
            )
            seen_accounts = set()
            for account_id, last_error, _, _ in check_rows:
                aid = int(account_id)
                check_done_ids.add(aid)
                if aid in seen_accounts:
                    continue
                seen_accounts.add(aid)
                sender_status_map[aid] = str(last_error or "").strip().lower()

        items = []
        for r in rows:
            aid = int(r.aid)
            login, disp = acc_map.get(aid, ("?", ""))
            name = disp or login
            items.append((aid, name, login, sender_status_map.get(aid, ""), aid in check_done_ids))

        return {
            "tgt": tgt,
            "required": required,
            "total": total,
            "page": page_clamped,
            "pages": pages,
            "items": items,
        }

    data = db_exec(_inner)
    if not data:
        show_menu_status(chat_id, "targets", "❌ Цель не найдена.")
        return

    tgt = data["tgt"]
    required = int(data["required"])
    total = int(data["total"])
    page_clamped = int(data["page"])
    pages = int(data["pages"])
    items = data["items"]

    lines = [
        f"Цель #{tgt.id} {tgt.username}",
        f"Отправители ({total}/{required})",
        f"Страница: {page_clamped}/{pages} (по {page_size})",
        "Листай: ◀️ Страница / ▶️ Страница",
        "",
    ]

    if not items:
        lines.append("Отправителей пока нет.")
        text = "\n".join(lines)
        show_menu_status(chat_id, "targets", text)
        return

    start_idx = (page_clamped - 1) * page_size
    for i, (_, name, login, sender_status, has_done_check) in enumerate(items, start=1):
        icon = _sender_icon(str(tgt.status or ""), str(sender_status or ""), bool(has_done_check))
        lines.append(f"{start_idx + i}) {icon} {name} ({login})")

    text = "\n".join(lines)
    if len(text) > 3800:
        text = text[:3790] + "…"

    # сохраняем контекст (может пригодиться для доп. навигации)
    set_current_menu(chat_id, "goal_nicks")
    set_chat_ui_value(chat_id, "senders_target_id", int(tgt.id))
    set_chat_ui_value(chat_id, "senders_page", int(page_clamped))
    set_chat_ui_value(chat_id, "goal_page_context", "target_senders")

    # В проде показываем красивый экран с кнопками,
    # но в тестах/без токена может быть 401 — тогда fallback на show_menu_status.
    try:
        show_screen(chat_id, text, reply_markup=kb_goal_nicks_reply(), parse_mode=None, force_new=True)
    except Exception:
        show_menu_status(chat_id, "targets", text)

def show_screen(
    chat_id: int,
    text: str,
    reply_markup=None,
    parse_mode: str = None,
    disable_preview: bool = True,
    force_new: bool = False,
):
    """
    "Экран" на чат.
    По умолчанию стараемся редактировать, а не спамить новыми сообщениями.
    Для UX меню можно включить force_new=True: отправим новый экран "вниз" и удалим старый.
    """
    cleanup_transient_messages(chat_id)

    with CHAT_UI_LOCK:
        st = CHAT_UI.get(chat_id) or {}
        msg_id = st.get("screen_msg_id")

    if msg_id and not force_new:
        try:
            edit_markup = reply_markup if _is_inline_markup(reply_markup) else None
            bot.edit_message_text(
                text,
                chat_id=chat_id,
                message_id=msg_id,
                parse_mode=parse_mode,
                disable_web_page_preview=disable_preview,
                reply_markup=edit_markup,
            )
            return msg_id
        except Exception as e:
            s = str(e).lower()
            if "message is not modified" in s:
                return msg_id
            # Markdown parsing can fail on dynamic content; retry once in plain text.
            if "can't parse entities" in s and parse_mode is not None:
                try:
                    bot.edit_message_text(
                        text,
                        chat_id=chat_id,
                        message_id=msg_id,
                        parse_mode=None,
                        disable_web_page_preview=disable_preview,
                        reply_markup=edit_markup,
                    )
                    return msg_id
                except Exception as e2:
                    logger.warning(
                        "ui_edit_plain_fallback_failed chat_id=%s msg_id=%s err=%s",
                        chat_id,
                        msg_id,
                        str(e2)[:180],
                    )
            # Если редактирование не удалось, удалим старый экран, чтобы не плодить дубли.
            _safe_delete(chat_id, msg_id)

    try:
        msg = bot.send_message(
            chat_id,
            text,
            parse_mode=parse_mode,
            disable_web_page_preview=disable_preview,
            reply_markup=reply_markup,
        )
    except Exception as e:
        s = str(e).lower()
        if "can't parse entities" in s and parse_mode is not None:
            msg = bot.send_message(
                chat_id,
                text,
                parse_mode=None,
                disable_web_page_preview=disable_preview,
                reply_markup=reply_markup,
            )
        else:
            raise
    with CHAT_UI_LOCK:
        st = CHAT_UI.get(chat_id) or {}
        old_id = st.get("screen_msg_id")
        st["screen_msg_id"] = msg.message_id
        CHAT_UI[chat_id] = st
    # If we forced a new screen, keep the chat clean by deleting the previous screen message.
    if force_new and old_id and old_id != msg.message_id:
        _safe_delete(chat_id, old_id)
    return msg.message_id


def _set_step_prompt(chat_id: int, user_id: int, prompt_msg_id: int):
    with STEP_LOCK:
        STEP_PROMPTS[(chat_id, user_id)] = prompt_msg_id


def set_current_menu(chat_id: int, menu_key: str):
    with CHAT_UI_LOCK:
        st = CHAT_UI.get(chat_id) or {}
        st["menu_key"] = menu_key
        CHAT_UI[chat_id] = st


def get_current_menu(chat_id: int) -> str:
    with CHAT_UI_LOCK:
        st = CHAT_UI.get(chat_id) or {}
        return str(st.get("menu_key") or "")


def set_chat_ui_value(chat_id: int, key: str, value):
    with CHAT_UI_LOCK:
        st = CHAT_UI.get(chat_id) or {}
        st[key] = value
        CHAT_UI[chat_id] = st


def get_chat_ui_int(chat_id: int, key: str, default: int = 0) -> int:
    with CHAT_UI_LOCK:
        st = CHAT_UI.get(chat_id) or {}
        try:
            return int(st.get(key, default))
        except Exception:
            return int(default)


def get_chat_ui_value(chat_id: int, key: str, default=None):
    with CHAT_UI_LOCK:
        st = CHAT_UI.get(chat_id) or {}
        return st.get(key, default)


def get_selected_campaign_id(chat_id: int) -> int:
    cid = get_chat_ui_int(chat_id, "selected_campaign_id", 0)
    if cid > 0:
        exists = db_exec(lambda db: db.query(Campaign.id).filter(Campaign.id == int(cid)).first() is not None)
        if exists:
            return int(cid)

    def _pick_fallback(db):
        row_any = db.query(Campaign.id).order_by(Campaign.id.asc()).first()
        return int(row_any[0]) if row_any else 0

    fallback = db_exec(_pick_fallback)
    fallback = int(fallback or 0)
    set_chat_ui_value(chat_id, "selected_campaign_id", fallback)
    return fallback


def _pop_step_prompt(chat_id: int, user_id: int) -> Optional[int]:
    with STEP_LOCK:
        return STEP_PROMPTS.pop((chat_id, user_id), None)

def cancel_all_step_prompts(chat_id: int):
    """
    Удалить все активные step-prompt сообщения (ask_step) в чате.
    Это помогает не оставлять "введи ..." висеть, если пользователь перешел в другой раздел.
    """
    with STEP_LOCK:
        keys = [k for k in STEP_PROMPTS.keys() if k[0] == chat_id]
        prompt_ids = [STEP_PROMPTS.pop(k) for k in keys]

    for pid in prompt_ids:
        if pid:
            _safe_delete(chat_id, pid)


def ask_step(message, prompt_text: str, next_handler, parse_mode: str = None):
    """Задать вопрос и привязать next_step_handler, при этом чистить старый prompt."""
    chat_id = message.chat.id
    user_id = message.from_user.id
    cleanup_transient_messages(chat_id)

    old = _pop_step_prompt(chat_id, user_id)
    if old:
        _safe_delete(chat_id, old)

    p = bot.send_message(chat_id, prompt_text, parse_mode=parse_mode)
    _set_step_prompt(chat_id, user_id, p.message_id)
    bot.register_next_step_handler(p, next_handler)


def cleanup_step(message):
    """В начале step-хендлеров: удалить prompt и (попытаться) удалить ответ пользователя."""
    chat_id = message.chat.id
    user_id = message.from_user.id

    prompt_id = _pop_step_prompt(chat_id, user_id)
    if prompt_id:
        _safe_delete(chat_id, prompt_id)

    _safe_delete_user_message(message)
    cleanup_transient_messages(chat_id)


# ----------------------------
# Keyboards
# ----------------------------
def kb_senders_pager(target_id: int, page: int, pages: int) -> types.InlineKeyboardMarkup:
    m = types.InlineKeyboardMarkup(row_width=2)
    prev_p = max(1, page - 1)
    next_p = min(pages, page + 1)
    m.add(
        types.InlineKeyboardButton("⬅️", callback_data=f'tgt_senders:{target_id}:{prev_p}'),
        types.InlineKeyboardButton("➡️", callback_data=f'tgt_senders:{target_id}:{next_p}'),
    )
    return m

INLINE_ACTION_TEXT = {
    "main_accounts": "👥 Аккаунты",
    "main_targets": "🎯 Цели",
    "main_settings": "⚙️ Настройки",
    "main_manage": "🔧 Управление",
    "main_stats": "📊 Статистика",
    "main_diag": "⚠️ Диагностика",
    "back": "⬅️ Назад",
    "back_main": "🏠 Меню",
    "acc_import": "📥 Импорт файлов",
    "acc_list": "📋 Список аккаунтов",
    "acc_device": "🔐 Авторизация Epic (ссылка)",
    "acc_check": "✅ Проверить аккаунты",
    "acc_refresh_names": "🔄 Обновить ники Epic",
    "acc_nick_change_import": "📝 Массовая смена ников",
    "acc_nick_change_status": "📊 Статус смены ников",
    "acc_add": "➕ Добавить аккаунт",
    "acc_del": "➖ Удалить аккаунт",
    "acc_prev": "◀️ Аккаунты",
    "acc_next": "▶️ Аккаунты",
    "acc_search": "🔎 Поиск аккаунтов",
    "targets_manager": "🗂️ Менеджер целей",
    "targets_stats": "📊 Статистика целей",
    "targets_start_all": "▶️ Запустить все цели",
    "targets_stop_all": "⏸️ Остановить все цели",
    "targets_stop_operation": "⛔ Остановить операцию",
    "goal_add": "➕ Добавить цель",
    "goal_list": "📋 Список целей",
    "goal_select": "🎯 Выбрать цель",
    "goal_edit": "✏️ Редактировать цель",
    "goal_progress": "📊 Статистика цели",
    "goal_group_nicks": "👥 Ники",
    "goal_group_sending": "🚀 Отправка",
    "goal_group_ops": "🧹 Операции",
    "goal_import_nicks": "📥 Импорт ников",
    "goal_nicks": "📋 Ники цели",
    "goal_nick_statuses": "📄 Статусы по никам",
    "goal_page_prev": "◀️ Страница",
    "goal_page_next": "▶️ Страница",
    # Backward compatibility for old inline messages.
    "goal_nick_prev": "◀️ Страница",
    "goal_nick_next": "▶️ Страница",
    "goal_nick_search": "🔎 Поиск ников",
    "goal_add_nick": "➕ Добавить ник",
    "goal_del_nick": "➖ Удалить ник",
    "goal_distribute": "🚀 Распределить ники",
    "goal_senders_for_nick": "👀 Отправители по нику",
    # Backward compatibility aliases; unified page arrows are used in menu.
    "goal_senders_prev": "◀️ Страница",
    "goal_senders_next": "▶️ Страница",
    "goal_force_one": "⚡ Форс-цикл с аккаунта",
    "goal_force_random": "🎲 Форс-цикл (рандом)",
    "goal_start": "▶️ Запустить цель",
    "goal_stop": "⏸️ Остановить цель",
    "goal_revoke": "↩️ Отозвать заявки",
    "goal_remove_friends": "🗑️ Удалить из друзей",
    "goal_check_friends": "🔍 Проверить в друзьях",
    "goal_resend_missing": "🔁 Дослать отсутствующих",
    "goal_delete": "🗑️ Удалить цель",
    "goal_limit": "🔄 Лимит",
    "goal_jitter": "⏱️ Джиттер",
    "goal_windows": "🕐 Окна",
    "goal_target_limit": "🎯 На ник",
    "goal_recheck_limit": "🔁 Лимит повторных проверок",
    "goal_daily_repeat": "📅 Ежедневный повтор",
    "goal_send_algo": "🔀 Алгоритм отправки",
    "goal_sender_pick_mode": "🎲 Порядок отправителей",
    "goal_params": "📋 Параметры цели",
    "set_api_limits": "🛡️ API лимиты",
    "set_proxy": "📌 Прокси",
    "set_auth_access": "👤 Доступ auth",
    "set_new_sends": "🧯 Новые заявки",
    "set_recheck_only": "♻️ Только recheck",
    "auth_add": "➕ Добавить ID auth",
    "auth_del": "➖ Удалить ID auth",
    "auth_list": "📋 Список ID auth",
    "auth_clear": "🗑 Очистить ID auth",
    "manage_tick": "▶️ Тик (1 раз)",
    "manage_export": "📤 Экспорт",
    "manage_stop": "⏸️ Стоп обработки",
    "manage_start": "▶️ Старт обработки",
    "manage_status": "📍 Статус обработки",
    "proxy_add": "➕ Добавить прокси",
    "proxy_list": "📋 Список прокси",
    "proxy_del": "🗑️ Удалить прокси",
}


def _act(code: str) -> str:
    return f"act:{code}"


def _act_btn(code: str) -> types.InlineKeyboardButton:
    return types.InlineKeyboardButton(INLINE_ACTION_TEXT[code], callback_data=_act(code))


def kb_main_reply() -> types.InlineKeyboardMarkup:
    m = types.InlineKeyboardMarkup(row_width=2)
    m.add(_act_btn("main_accounts"), _act_btn("main_targets"))
    m.add(_act_btn("main_settings"), _act_btn("main_manage"))
    m.add(_act_btn("main_stats"), _act_btn("main_diag"))
    return m


def kb_auth_operator_reply() -> types.InlineKeyboardMarkup:
    m = types.InlineKeyboardMarkup(row_width=2)
    m.add(_act_btn("acc_list"), _act_btn("acc_device"))
    m.add(_act_btn("acc_prev"), _act_btn("acc_next"))
    m.add(_act_btn("acc_search"))
    m.add(_act_btn("back_main"))
    return m


def kb_accounts_reply() -> types.InlineKeyboardMarkup:
    m = types.InlineKeyboardMarkup(row_width=2)
    m.add(_act_btn("acc_import"), _act_btn("acc_list"))
    m.add(_act_btn("acc_device"), _act_btn("acc_check"))
    m.add(_act_btn("acc_refresh_names"))
    m.add(_act_btn("acc_nick_change_import"), _act_btn("acc_nick_change_status"))
    m.add(_act_btn("acc_add"), _act_btn("acc_del"))
    m.add(_act_btn("acc_prev"), _act_btn("acc_next"))
    m.add(_act_btn("acc_search"))
    m.add(_act_btn("back"))
    return m


def kb_targets_reply() -> types.InlineKeyboardMarkup:
    m = types.InlineKeyboardMarkup(row_width=2)
    m.add(_act_btn("targets_manager"), _act_btn("targets_stats"))
    m.add(_act_btn("targets_start_all"), _act_btn("targets_stop_all"))
    m.add(_act_btn("targets_stop_operation"))
    m.add(_act_btn("back"))
    return m


def kb_goal_manager_reply() -> types.InlineKeyboardMarkup:
    m = types.InlineKeyboardMarkup(row_width=2)
    m.add(_act_btn("goal_add"), _act_btn("goal_list"))
    m.add(_act_btn("goal_select"))
    m.add(_act_btn("back"))
    return m


def kb_goal_selected_reply() -> types.InlineKeyboardMarkup:
    m = types.InlineKeyboardMarkup(row_width=2)
    m.add(_act_btn("goal_edit"), _act_btn("goal_progress"))
    m.add(_act_btn("goal_group_nicks"), _act_btn("goal_group_sending"))
    m.add(_act_btn("goal_group_ops"), _act_btn("back"))
    return m


def kb_goal_nicks_reply() -> types.InlineKeyboardMarkup:
    m = types.InlineKeyboardMarkup(row_width=2)
    m.add(_act_btn("goal_import_nicks"), _act_btn("goal_nicks"))
    m.add(_act_btn("goal_nick_statuses"), _act_btn("goal_senders_for_nick"))
    m.add(_act_btn("goal_page_prev"), _act_btn("goal_page_next"))
    m.add(_act_btn("goal_nick_search"))
    m.add(_act_btn("goal_add_nick"), _act_btn("goal_del_nick"))
    m.add(_act_btn("back"))
    return m


def kb_goal_sending_reply() -> types.InlineKeyboardMarkup:
    m = types.InlineKeyboardMarkup(row_width=2)
    m.add(_act_btn("goal_distribute"))
    m.add(_act_btn("goal_force_one"), _act_btn("goal_force_random"))
    m.add(_act_btn("goal_start"), _act_btn("goal_stop"))
    m.add(_act_btn("back"))
    return m


def kb_goal_ops_reply() -> types.InlineKeyboardMarkup:
    m = types.InlineKeyboardMarkup(row_width=2)
    m.add(_act_btn("goal_check_friends"), _act_btn("goal_resend_missing"))
    m.add(_act_btn("goal_revoke"), _act_btn("goal_remove_friends"))
    m.add(_act_btn("targets_stop_operation"))
    m.add(_act_btn("goal_delete"))
    m.add(_act_btn("back"))
    return m


def kb_goal_edit_reply() -> types.InlineKeyboardMarkup:
    m = types.InlineKeyboardMarkup(row_width=2)
    m.add(_act_btn("goal_limit"), _act_btn("goal_jitter"))
    m.add(_act_btn("goal_windows"), _act_btn("goal_target_limit"))
    m.add(_act_btn("goal_recheck_limit"), _act_btn("goal_daily_repeat"))
    m.add(_act_btn("goal_send_algo"), _act_btn("goal_sender_pick_mode"))
    m.add(_act_btn("goal_params"))
    m.add(_act_btn("back"))
    return m

def kb_settings_reply() -> types.InlineKeyboardMarkup:
    m = types.InlineKeyboardMarkup(row_width=2)
    m.add(_act_btn("set_api_limits"), _act_btn("set_proxy"))
    m.add(_act_btn("set_new_sends"), _act_btn("set_recheck_only"))
    m.add(_act_btn("set_auth_access"))
    m.add(_act_btn("back"))
    return m


def kb_auth_access_reply() -> types.InlineKeyboardMarkup:
    m = types.InlineKeyboardMarkup(row_width=2)
    m.add(_act_btn("auth_add"), _act_btn("auth_del"))
    m.add(_act_btn("auth_list"), _act_btn("auth_clear"))
    m.add(_act_btn("back"))
    return m


def kb_manage_reply() -> types.InlineKeyboardMarkup:
    m = types.InlineKeyboardMarkup(row_width=2)
    m.add(_act_btn("manage_tick"), _act_btn("manage_export"))
    m.add(_act_btn("manage_stop"), _act_btn("manage_start"))
    m.add(_act_btn("manage_status"))
    m.add(_act_btn("back"))
    return m


def kb_proxy_reply() -> types.InlineKeyboardMarkup:
    m = types.InlineKeyboardMarkup(row_width=2)
    m.add(_act_btn("proxy_add"), _act_btn("proxy_list"))
    m.add(_act_btn("proxy_del"), _act_btn("back"))
    return m


# ----------------------------
# Menus (screens)
# ----------------------------

def show_stats_screen(chat_id: int):
    cancel_all_step_prompts(chat_id)
    def _inner(db):
        return {
            "acc_total": db.query(Account).count(),
            "acc_active": db.query(Account).filter(
                Account.enabled == True,
                Account.status == AccountStatus.ACTIVE.value
            ).count(),
            "acc_banned": db.query(Account).filter(
                Account.status == AccountStatus.BANNED.value
            ).count(),
            "tgt_total": db.query(Target).count(),
            "tgt_new": db.query(Target).filter(
                Target.status == TargetStatus.NEW.value
            ).count(),
            "tgt_accepted": db.query(Target).filter(
                Target.status == TargetStatus.ACCEPTED.value
            ).count(),
            "task_queued": db.query(Task).filter(
                Task.status.in_([TaskStatus.QUEUED.value])
            ).count(),
            "sent_today": db.query(func.sum(Account.today_sent)).scalar() or 0,
            "proxy_enabled": db.query(Proxy).filter(Proxy.enabled == True).count(),
        }

    stats = db_exec(_inner)
    text = (
        "📊 **Система:**\n\n"
        f"**Аккаунты:** {stats['acc_active']}/{stats['acc_total']} (❌ {stats['acc_banned']})\n"
        f"**Цели:** {stats['tgt_total']} (🆕 {stats['tgt_new']}, ✅ {stats['tgt_accepted']})\n"
        f"**Очередь:** {stats['task_queued']} задач\n"
        f"**Сегодня:** {stats['sent_today']} отправлено\n"
        f"**Прокси:** {stats['proxy_enabled']} активных\n"
    )
    show_screen(chat_id, text, reply_markup=kb_main_reply(), parse_mode="Markdown", force_new=True)

def show_main_menu(chat_id: int):
    # Главное меню (inline keyboard).
    cancel_all_step_prompts(chat_id)
    set_current_menu(chat_id, "home")
    show_screen(
        chat_id,
        "🏠 Главное меню\nВыбери раздел кнопками ниже.",
        reply_markup=kb_main_reply(),
        parse_mode=None,
        force_new=True,
    )


def show_auth_operator_menu(chat_id: int):
    cancel_all_step_prompts(chat_id)
    set_current_menu(chat_id, "auth_only")
    show_screen(
        chat_id,
        "🔐 Режим авторизации\nДоступны только: список аккаунтов и авторизация Epic по ссылке.",
        reply_markup=kb_auth_operator_reply(),
        parse_mode=None,
        force_new=True,
    )


def show_accounts_menu(chat_id: int):
    cancel_all_step_prompts(chat_id)
    set_current_menu(chat_id, "accounts")
    set_chat_ui_value(chat_id, "acc_page", 1)
    set_chat_ui_value(chat_id, "acc_query", "")
    show_screen(
        chat_id,
        "👥 Управление аккаунтами\nИмпорт, авторизация Epic, список, проверка.",
        reply_markup=kb_accounts_reply(),
        parse_mode=None,
        force_new=True,
    )


def show_nickname_change_import_menu(chat_id: int):
    cancel_all_step_prompts(chat_id)
    set_current_menu(chat_id, "nick_change_import")
    show_screen(
        chat_id,
        "📝 Массовая смена ников\n"
        "Пришли файл .xlsx/.txt/.csv с 2 колонками:\n"
        "1) login/email аккаунта\n"
        "2) новый ник\n\n"
        "Ограничения ника: 3-16, латиница/цифры/_.\n"
        "Если ник занят, бот попробует похожие варианты автоматически.\n\n"
        "Примеры строк:\n"
        "mail@example.com;NewNick123\n"
        "mail2@example.com,NextNick_1",
        reply_markup=kb_accounts_reply(),
        parse_mode=None,
        force_new=True,
    )


def show_targets_menu(chat_id: int):
    cancel_all_step_prompts(chat_id)
    set_current_menu(chat_id, "targets")
    set_chat_ui_value(chat_id, "tgt_page", 1)
    set_chat_ui_value(chat_id, "tgt_query", "")
    cid = get_selected_campaign_id(chat_id)
    label, active, total = db_exec(
        lambda db: (
            campaign_ui_label(db, cid),
            int(db.query(Campaign).filter(Campaign.enabled == True).count()),
            int(db.query(Campaign.id).count()),
        )
    )
    show_screen(
        chat_id,
        f"🎯 Цели\nВыбрана в интерфейсе: {label}\n"
        f"Активные цели: {active}/{total}\n"
        "Запуск/остановка и общий статус.",
        reply_markup=kb_targets_reply(),
        parse_mode=None,
        force_new=True,
    )


def show_goal_manager_menu(chat_id: int):
    cancel_all_step_prompts(chat_id)
    set_current_menu(chat_id, "goal_manager")
    cid = get_selected_campaign_id(chat_id)
    label = db_exec(lambda db: campaign_ui_label(db, cid))
    show_screen(
        chat_id,
        f"🗂️ Менеджер целей\nВыбрана в интерфейсе: {label}\n"
        "Создание, выбор и список целей.",
        reply_markup=kb_goal_manager_reply(),
        parse_mode=None,
        force_new=True,
    )


def show_selected_goal_menu(chat_id: int):
    cancel_all_step_prompts(chat_id)
    set_current_menu(chat_id, "goal_selected")
    cid = get_selected_campaign_id(chat_id)
    label = db_exec(lambda db: campaign_ui_label(db, cid))
    show_screen(
        chat_id,
        f"🎯 Выбрана цель: {label}\n"
        "Открой нужный блок: Ники, Отправка или Операции.",
        reply_markup=kb_goal_selected_reply(),
        parse_mode=None,
        force_new=True,
    )


def show_goal_nicks_menu(chat_id: int):
    cancel_all_step_prompts(chat_id)
    set_current_menu(chat_id, "goal_nicks")
    cid = get_selected_campaign_id(chat_id)
    label = db_exec(lambda db: campaign_ui_label(db, cid))
    show_screen(
        chat_id,
        f"👥 Ники цели\n{label}\n"
        "Импорт, список, статусы, поиск и редактирование ников.",
        reply_markup=kb_goal_nicks_reply(),
        parse_mode=None,
        force_new=True,
    )


def show_goal_sending_menu(chat_id: int):
    cancel_all_step_prompts(chat_id)
    set_current_menu(chat_id, "goal_sending")
    cid = get_selected_campaign_id(chat_id)
    label = db_exec(lambda db: campaign_ui_label(db, cid))
    show_screen(
        chat_id,
        f"🚀 Отправка по цели\n{label}\n"
        "Распределение, форс-цикл, запуск и остановка цели.",
        reply_markup=kb_goal_sending_reply(),
        parse_mode=None,
        force_new=True,
    )


def show_goal_ops_menu(chat_id: int):
    cancel_all_step_prompts(chat_id)
    set_current_menu(chat_id, "goal_ops")
    cid = get_selected_campaign_id(chat_id)
    label = db_exec(lambda db: campaign_ui_label(db, cid))
    show_screen(
        chat_id,
        f"🧹 Операции по цели\n{label}\n"
        "Отзыв заявок, удаление из друзей, остановка операции и удаление цели.",
        reply_markup=kb_goal_ops_reply(),
        parse_mode=None,
        force_new=True,
    )


def show_goal_edit_menu(chat_id: int):
    cancel_all_step_prompts(chat_id)
    set_current_menu(chat_id, "goal_edit")
    cid = get_selected_campaign_id(chat_id)
    label = db_exec(lambda db: campaign_ui_label(db, cid))
    show_screen(
        chat_id,
        f"✏️ Редактирование цели\n{label}\n"
        "Лимит на аккаунт, окна, джиттер, алгоритм отправки, лимит повторных проверок, ежедневный повтор.",
        reply_markup=kb_goal_edit_reply(),
        parse_mode=None,
        force_new=True,
    )


def show_settings_menu(chat_id: int):
    cancel_all_step_prompts(chat_id)
    set_current_menu(chat_id, "settings")
    sends_enabled, recheck_only = db_exec(
        lambda db: (
            is_new_send_requests_enabled(db),
            is_recheck_only_mode_enabled(db),
        )
    )
    show_screen(
        chat_id,
        "⚙️ Настройки\n"
        "Глобальные параметры: API лимиты, прокси, доступ auth.\n"
        f"Новые заявки: {'включены' if sends_enabled else 'выключены'}\n"
        f"Режим recheck-only: {'включен' if recheck_only else 'выключен'}",
        reply_markup=kb_settings_reply(),
        parse_mode=None,
        force_new=True,
    )


def show_auth_access_menu(chat_id: int):
    cancel_all_step_prompts(chat_id)
    set_current_menu(chat_id, "auth_access")

    ids = db_exec(lambda db: sorted(_load_auth_operator_ids(db)))
    preview = ", ".join(str(x) for x in ids) if ids else "пусто"
    show_screen(
        chat_id,
        "👤 Доступ auth\n"
        "Управление ID пользователей с ограниченным доступом "
        "(только список аккаунтов + авторизация Epic).\n\n"
        f"Текущие ID: {preview}",
        reply_markup=kb_auth_access_reply(),
        parse_mode=None,
        force_new=True,
    )


def show_manage_menu(chat_id: int):
    cancel_all_step_prompts(chat_id)
    set_current_menu(chat_id, "manage")
    show_screen(
        chat_id,
        "🔧 Управление\nРучной тик, экспорт, статус.",
        reply_markup=kb_manage_reply(),
        parse_mode=None,
        force_new=True,
    )

def show_diagnostics_screen(chat_id: int):
    cancel_all_step_prompts(chat_id)
    now = utc_now()

    def _inner(db):
        accs = db.query(Account).order_by(Account.id.asc()).all()
        tasks_queued = db.query(Task).filter(Task.status.in_(QUEUED_OR_POSTPONED_TASK_STATUSES)).count()
        tasks_send_active = db.query(Task).filter(
            Task.task_type == "send_request",
            Task.status.in_(ACTIVE_SEND_TASK_STATUSES),
        ).count()
        tasks_check_active = db.query(Task).filter(
            Task.task_type == "check_status",
            Task.status.in_(ACTIVE_SEND_TASK_STATUSES),
        ).count()

        targets_by_status = dict(db.query(Target.status, func.count(Target.id)).group_by(Target.status).all())

        # Last warnings/errors for quick debugging in-chat.
        last_logs = (
            db.query(LogEvent)
            .filter(LogEvent.level.in_(["warning", "error"]))
            .order_by(LogEvent.id.desc())
            .limit(10)
            .all()
        )

        return accs, tasks_queued, tasks_send_active, tasks_check_active, targets_by_status, last_logs

    accs, tasks_queued, tasks_send_active, tasks_check_active, targets_by_status, last_logs = db_exec(_inner)
    tz_name = db_exec(lambda db: get_setting(db, "runtime_timezone", DEFAULT_TIMEZONE))
    try:
        tz = ZoneInfo(tz_name)
    except Exception:
        tz = ZoneInfo(DEFAULT_TIMEZONE)

    total = len(accs)
    enabled = sum(1 for a in accs if a.enabled)
    active = sum(1 for a in accs if a.status == AccountStatus.ACTIVE.value and a.enabled)
    paused = sum(1 for a in accs if a.status == AccountStatus.PAUSED.value)
    banned = sum(1 for a in accs if a.status == AccountStatus.BANNED.value)
    manual = sum(1 for a in accs if a.status == AccountStatus.MANUAL.value)

    with_da = sum(1 for a in accs if a.epic_account_id and a.device_id and a.device_secret and a.enabled and a.status == AccountStatus.ACTIVE.value)
    missing_da = sum(1 for a in accs if (not (a.epic_account_id and a.device_id and a.device_secret)) and a.enabled and a.status == AccountStatus.ACTIVE.value)

    in_warmup = 0
    out_of_window = 0
    at_limit = 0
    eligible_now = 0
    bad_windows_json = 0

    for a in accs:
        if not a.enabled or a.status != AccountStatus.ACTIVE.value:
            continue
        if not (a.epic_account_id and a.device_id and a.device_secret):
            continue
        if a.warmup_until and now < a.warmup_until:
            in_warmup += 1
            continue

        try:
            windows = json.loads(a.active_windows_json or "[]")
        except Exception:
            bad_windows_json += 1
            windows = []

        local_now = now.replace(tzinfo=timezone.utc).astimezone(tz).replace(tzinfo=None)
        if not is_in_window_utc(windows, local_now):
            out_of_window += 1
            continue
        if int(a.today_sent or 0) >= int(a.daily_limit or 0):
            at_limit += 1
            continue
        eligible_now += 1

    def _t(st: str) -> int:
        return int(targets_by_status.get(st, 0) or 0)

    lines = []
    now_local = now.replace(tzinfo=timezone.utc).astimezone(tz)
    lines.append(f"⚠️ Диагностика (сейчас, {tz.key}): " + now_local.strftime("%Y-%m-%d %H:%M:%S"))
    lines.append("")
    lines.append(f"Аккаунты: всего={total} enabled={enabled}")
    lines.append(f"Статусы: активные={active} на_паузе={paused} бан={banned} ручные={manual}")
    lines.append(f"Авторизация Epic: с_ключами={with_da} без_ключей={missing_da}")
    lines.append("")
    lines.append("Готовность к отправке сейчас (активные + авторизованные):")
    lines.append(
        f"готовы_сейчас={eligible_now} прогрев={in_warmup} вне_окна={out_of_window} "
        f"в_лимите={at_limit} ошибки_окон={bad_windows_json}"
    )
    lines.append("")
    lines.append(f"Очередь: отложено_или_в_очереди={tasks_queued} отправка_активно={tasks_send_active} проверка_активно={tasks_check_active}")
    lines.append("")
    lines.append("Цели:")
    lines.append(
        " ".join(
            [
                f"new={_t(TargetStatus.NEW.value)}",
                f"в_ожидании={_t(TargetStatus.PENDING.value)}",
                f"отправлено={_t(TargetStatus.SENT.value)}",
                f"принято={_t(TargetStatus.ACCEPTED.value)}",
                f"отклонено={_t(TargetStatus.REJECTED.value)}",
                f"ошибка={_t(TargetStatus.FAILED.value)}",
            ]
        )
    )

    if last_logs:
        lines.append("")
        lines.append("Последние warning/error:")
        for ev in last_logs[:10]:
            ts = ev.created_at.strftime("%m-%d %H:%M:%S") if ev.created_at else "?"
            msg = (ev.message or "").replace("\n", " ")
            if len(msg) > 160:
                msg = msg[:157] + "..."
            lines.append(f"- {ts} {ev.level}: {msg}")

    text = "\n".join(lines)
    if len(text) > 3800:
        text = text[:3790] + "…"
    show_screen(chat_id, text, reply_markup=kb_main_reply(), parse_mode=None, force_new=True)


def show_menu_status(chat_id: int, menu_key: str, status_text: str):
    # Preserve deep goal context: status updates from goal screens must not kick user back.
    if menu_key == "targets":
        current = get_current_menu(chat_id)
        if current in {"goal_selected", "goal_nicks", "goal_sending", "goal_ops", "goal_edit"}:
            menu_key = current

    menus = {
        "accounts": ("👥 Управление аккаунтами", kb_accounts_reply),
        "nick_change_import": ("📝 Массовая смена ников", kb_accounts_reply),
        "targets": ("🎯 Цели", kb_targets_reply),
        "goal_manager": ("🗂️ Менеджер целей", kb_goal_manager_reply),
        "goal_selected": ("🎯 Текущая цель", kb_goal_selected_reply),
        "goal_nicks": ("👥 Ники цели", kb_goal_nicks_reply),
        "goal_sending": ("🚀 Отправка по цели", kb_goal_sending_reply),
        "goal_ops": ("🧹 Операции по цели", kb_goal_ops_reply),
        "goal_edit": ("✏️ Редактирование цели", kb_goal_edit_reply),
        "settings": ("⚙️ Настройки", kb_settings_reply),
        "auth_access": ("👤 Доступ auth", kb_auth_access_reply),
        "manage": ("🔧 Управление", kb_manage_reply),
        "proxy": ("📌 Прокси", kb_proxy_reply),
    }
    title, kb_fn = menus[menu_key]
    set_current_menu(chat_id, menu_key)
    body = f"{title}\n\n{status_text}" if status_text else title
    show_screen(chat_id, body, reply_markup=kb_fn(), parse_mode=None, force_new=True)


def _normalized_search_query(value: str) -> str:
    return (value or "").strip()


def _target_current_friends_map(db, target_ids: list[int]) -> dict[int, int]:
    """
    Current friend count per target based on the latest DONE check_status
    for each (target, sender) pair.
    """
    if not target_ids:
        return {}
    rows = (
        db.query(Task.target_id, Task.account_id, Task.last_error, Task.completed_at, Task.id)
        .filter(
            Task.target_id.in_([int(x) for x in target_ids]),
            Task.task_type == "check_status",
            Task.status == TaskStatus.DONE.value,
        )
        .order_by(
            Task.target_id.asc(),
            Task.account_id.asc(),
            Task.completed_at.desc().nullslast(),
            Task.id.desc(),
        )
        .all()
    )
    result: dict[int, int] = {}
    seen_pairs: set[tuple[int, int]] = set()
    for target_id, account_id, last_error, _, _ in rows:
        key = (int(target_id), int(account_id))
        if key in seen_pairs:
            continue
        seen_pairs.add(key)
        if str(last_error or "").strip().lower() == "friend_status:accepted":
            result[int(target_id)] = int(result.get(int(target_id), 0)) + 1
    return result


def _target_accepted_unique_map(db, target_ids: list[int]) -> dict[int, int]:
    """
    Unique accepted count per target by (target, sender) pair.
    Prevents overcount from repeated accepted checks for the same sender.
    """
    if not target_ids:
        return {}
    rows = (
        db.query(Task.target_id, Task.account_id, Task.task_type, Task.last_error, Task.completed_at, Task.id)
        .filter(
            Task.target_id.in_([int(x) for x in target_ids]),
            Task.status == TaskStatus.DONE.value,
            Task.task_type.in_(["check_status", "send_request"]),
        )
        .order_by(
            Task.target_id.asc(),
            Task.account_id.asc(),
            Task.completed_at.desc().nullslast(),
            Task.id.desc(),
        )
        .all()
    )
    result: dict[int, int] = {}
    seen_pairs: set[tuple[int, int]] = set()
    for target_id, account_id, task_type, last_error, _, _ in rows:
        pair = (int(target_id), int(account_id))
        if pair in seen_pairs:
            continue
        sender_status = str(last_error or "").strip().lower()
        is_accepted = (task_type == "check_status" and sender_status == "friend_status:accepted") or (
            task_type == "send_request" and sender_status == "already_accepted_before_send"
        )
        if not is_accepted:
            continue
        seen_pairs.add(pair)
        result[int(target_id)] = int(result.get(int(target_id), 0)) + 1
    return result


def _pair_was_accepted(db, target_id: int, account_id: int) -> bool:
    """
    Whether sender->target pair was already observed as accepted in DONE tasks.
    """
    row = (
        db.query(Task.id)
        .filter(
            Task.target_id == int(target_id),
            Task.account_id == int(account_id),
            Task.status == TaskStatus.DONE.value,
            or_(
                and_(Task.task_type == "check_status", Task.last_error == "friend_status:accepted"),
                and_(Task.task_type == "send_request", Task.last_error == "already_accepted_before_send"),
            ),
        )
        .first()
    )
    return bool(row)


def show_accounts_list(chat_id: int, page: int = 1, query: str = ""):
    query = _normalized_search_query(query)

    def _inner(db):
        q = db.query(Account)
        if query:
            q = q.filter(func.lower(Account.login).like(f"%{query.lower()}%"))
        total = q.count()
        pages = max(1, math.ceil(total / LIST_PAGE_SIZE)) if total else 1
        page_clamped = max(1, min(int(page), pages))
        rows = (
            q.order_by(Account.id.asc())
            .offset((page_clamped - 1) * LIST_PAGE_SIZE)
            .limit(LIST_PAGE_SIZE)
            .all()
        )
        return rows, total, pages, page_clamped

    accs, total, pages, page_clamped = db_exec(_inner)
    set_chat_ui_value(chat_id, "acc_page", page_clamped)
    set_chat_ui_value(chat_id, "acc_query", query)

    if total == 0:
        suffix = f" по запросу {md_inline_code(query)}" if query else ""
        show_screen(
            chat_id,
            f"❌ Аккаунтов не найдено{suffix}",
            reply_markup=kb_accounts_reply(),
            parse_mode="Markdown",
            force_new=True,
        )
        return

    lines = [f"👥 **Аккаунты** ({total})"]
    if query:
        lines.append(f"Фильтр: {md_inline_code(query)}")
    lines.append(f"Страница: {page_clamped}/{pages} (по {LIST_PAGE_SIZE})")
    lines.append("")
    start_idx = (page_clamped - 1) * LIST_PAGE_SIZE
    for i, a in enumerate(accs, start=1):
        icon = "✅" if a.status == AccountStatus.ACTIVE.value else "❌" if a.status == AccountStatus.BANNED.value else "⏸️"
        da_icon = "🔑" if a.device_id and a.device_secret and a.epic_account_id else "⚪️"
        sender_label = a.epic_display_name.strip() if getattr(a, "epic_display_name", None) else a.login
        lines.append(
            f"{icon} {da_icon} №{start_idx + i} #{a.id} {md_inline_code(sender_label)} "
            f"({md_inline_code(a.login)}) {a.today_sent}/{a.daily_limit}"
        )

    text = "\n".join(lines)
    if len(text) > 3800:
        text = text[:3790] + "…"
    show_screen(chat_id, text, reply_markup=kb_accounts_reply(), parse_mode="Markdown", force_new=True)


def show_targets_status(chat_id: int, page: int = 1, query: str = ""):
    query = _normalized_search_query(query)
    selected_campaign_id = get_selected_campaign_id(chat_id)

    def _inner(db):
        now = utc_now()
        day_start, day_end = local_day_bounds_utc_naive(db, now)
        base = db.query(Target).filter(target_campaign_filter(db, selected_campaign_id))
        if query:
            base = base.filter(func.lower(Target.username).like(f"%{query.lower()}%"))
        total = base.count()
        pages = max(1, math.ceil(total / TARGETS_PAGE_SIZE)) if total else 1
        page_clamped = max(1, min(int(page), pages))
        targets = (
            base.order_by(Target.id.asc())
            .offset((page_clamped - 1) * TARGETS_PAGE_SIZE)
            .limit(TARGETS_PAGE_SIZE)
            .all()
        )
        ids = [t.id for t in targets]
        friends_now_map = _target_current_friends_map(db, ids)
        senders_total_map = {}
        senders_today_map = {}
        sent_total_map = {}
        sent_today_map = {}
        if ids:
            rows_total = (
                db.query(Task.target_id, func.count(func.distinct(Task.account_id)))
                .filter(
                    Task.target_id.in_(ids),
                    Task.task_type == "send_request",
                    Task.status == TaskStatus.DONE.value,
                )
                .group_by(Task.target_id)
                .all()
            )
            rows_today = (
                db.query(Task.target_id, func.count(func.distinct(Task.account_id)))
                .filter(
                    Task.target_id.in_(ids),
                    Task.task_type == "send_request",
                    Task.status == TaskStatus.DONE.value,
                    Task.completed_at >= day_start,
                    Task.completed_at < day_end,
                )
                .group_by(Task.target_id)
                .all()
            )
            sent_total_rows = (
                db.query(Task.target_id, func.count(Task.id))
                .filter(
                    Task.target_id.in_(ids),
                    Task.task_type == "send_request",
                    Task.status == TaskStatus.DONE.value,
                )
                .group_by(Task.target_id)
                .all()
            )
            sent_today_rows = (
                db.query(Task.target_id, func.count(Task.id))
                .filter(
                    Task.target_id.in_(ids),
                    Task.task_type == "send_request",
                    Task.status == TaskStatus.DONE.value,
                    Task.completed_at >= day_start,
                    Task.completed_at < day_end,
                )
                .group_by(Task.target_id)
                .all()
            )
            senders_total_map = {int(tid): int(cnt or 0) for tid, cnt in rows_total}
            senders_today_map = {int(tid): int(cnt or 0) for tid, cnt in rows_today}
            sent_total_map = {int(tid): int(cnt or 0) for tid, cnt in sent_total_rows}
            sent_today_map = {int(tid): int(cnt or 0) for tid, cnt in sent_today_rows}
        by_status = dict(
            base.with_entities(Target.status, func.count(Target.id))
            .group_by(Target.status)
            .all()
        )
        default_required = max(1, get_setting_int(db, "target_senders_count", DEFAULT_TARGET_SENDERS_COUNT))
        camp = get_campaign_or_default(db, selected_campaign_id)
        camp_name = camp.name if camp else f"#{selected_campaign_id}"
        return (
            targets,
            total,
            pages,
            page_clamped,
            senders_total_map,
            senders_today_map,
            sent_total_map,
            sent_today_map,
            friends_now_map,
            by_status,
            default_required,
            camp_name,
        )

    (
        targets,
        total,
        pages,
        page_clamped,
        senders_total_map,
        senders_today_map,
        sent_total_map,
        sent_today_map,
        friends_now_map,
        by_status,
        default_required,
        camp_name,
    ) = db_exec(_inner)
    set_chat_ui_value(chat_id, "tgt_page", page_clamped)
    set_chat_ui_value(chat_id, "tgt_query", query)
    set_chat_ui_value(chat_id, "tgt_view_mode", "targets")
    set_chat_ui_value(chat_id, "goal_page_context", "targets_list")
    set_current_menu(chat_id, "goal_nicks")

    if total == 0:
        suffix = f" по запросу {md_inline_code(query)}" if query else ""
        show_screen(chat_id, f"❌ Целей не найдено{suffix}", reply_markup=kb_goal_nicks_reply(), parse_mode="Markdown")
        return

    icons = {
        TargetStatus.NEW.value: "🆕",
        TargetStatus.PENDING.value: "⏳",
        TargetStatus.SENT.value: "📨",
        TargetStatus.ACCEPTED.value: "✅",
        TargetStatus.REJECTED.value: "❌",
        TargetStatus.FAILED.value: "🚫",
    }
    lines = [f"🎯 **Ники цели {md_inline_code(camp_name)}** ({total})"]
    if query:
        lines.append(f"Фильтр: {md_inline_code(query)}")
    lines.append(f"Страница: {page_clamped}/{pages} (по {TARGETS_PAGE_SIZE})")
    lines.append("Листай: ◀️ Страница / ▶️ Страница")
    lines.append("")
    lines.append("Статусы:")
    for st, cnt in sorted(by_status.items()):
        lines.append(f"• {icons.get(st, '❓')} {target_status_ru(st)}: {int(cnt)}")
    lines.append("")
    start_idx = (page_clamped - 1) * TARGETS_PAGE_SIZE
    for i, t in enumerate(targets, start=1):
        senders_total = int(senders_total_map.get(int(t.id), 0))
        senders_today = int(senders_today_map.get(int(t.id), 0))
        sent_total = max(int(t.sent_count or 0), int(sent_total_map.get(int(t.id), 0)))
        sent_today = int(sent_today_map.get(int(t.id), 0))
        req = int(getattr(t, "required_senders", 0) or 0) or default_required
        lines.append(f"• N{start_idx + i} (id:{t.id}) {md_inline_code(t.username)}")
        lines.append(f"  Статус: {icons.get(t.status, '❓')} {target_status_ru(t.status)}")
        lines.append(
            f"  Отправлено (сегодня/всего): {sent_today}/{sent_total} | В друзьях: {int(friends_now_map.get(int(t.id), 0))}"
        )
        lines.append(f"  Отправители (сегодня/лимит): {senders_today}/{req} | Уникально всего: {senders_total}")
        lines.append("")
    text = "\n".join(lines)
    if len(text) > 3800:
        text = text[:3790] + "…"
    show_screen(chat_id, text, reply_markup=kb_goal_nicks_reply(), parse_mode="Markdown")


def show_targets_receiver_stats(chat_id: int, page: int = 1, query: str = ""):
    query = _normalized_search_query(query)
    selected_campaign_id = get_selected_campaign_id(chat_id)

    def _inner(db):
        base = db.query(Target).filter(target_campaign_filter(db, selected_campaign_id))
        if query:
            base = base.filter(func.lower(Target.username).like(f"%{query.lower()}%"))
        total = base.count()
        pages = max(1, math.ceil(total / TARGETS_PAGE_SIZE)) if total else 1
        page_clamped = max(1, min(int(page), pages))
        targets = (
            base.order_by(Target.id.asc())
            .offset((page_clamped - 1) * TARGETS_PAGE_SIZE)
            .limit(TARGETS_PAGE_SIZE)
            .all()
        )
        ids = [int(t.id) for t in targets]
        friends_now_map = _target_current_friends_map(db, ids)
        accepted_unique_map = _target_accepted_unique_map(db, ids)
        sent_done_map = {}
        send_q_map = {}
        check_q_map = {}
        revoke_map = {}
        remove_map = {}
        if ids:
            sent_done_rows = (
                db.query(Task.target_id, func.count(Task.id))
                .filter(
                    Task.target_id.in_(ids),
                    Task.task_type == "send_request",
                    Task.status == TaskStatus.DONE.value,
                )
                .group_by(Task.target_id)
                .all()
            )
            send_q_rows = (
                db.query(Task.target_id, func.count(Task.id))
                .filter(
                    Task.target_id.in_(ids),
                    Task.task_type == "send_request",
                    Task.status.in_(ACTIVE_SEND_TASK_STATUSES),
                )
                .group_by(Task.target_id)
                .all()
            )
            check_q_rows = (
                db.query(Task.target_id, func.count(Task.id))
                .filter(
                    Task.target_id.in_(ids),
                    Task.task_type == "check_status",
                    Task.status.in_(ACTIVE_SEND_TASK_STATUSES),
                )
                .group_by(Task.target_id)
                .all()
            )
            sent_done_map = {int(tid): int(cnt or 0) for tid, cnt in sent_done_rows}
            send_q_map = {int(tid): int(cnt or 0) for tid, cnt in send_q_rows}
            check_q_map = {int(tid): int(cnt or 0) for tid, cnt in check_q_rows}
            revoke_rows = (
                db.query(Task.target_id, func.count(Task.id))
                .filter(
                    Task.target_id.in_(ids),
                    Task.task_type == "revoke_request",
                    Task.status == TaskStatus.DONE.value,
                )
                .group_by(Task.target_id)
                .all()
            )
            remove_rows = (
                db.query(Task.target_id, func.count(Task.id))
                .filter(
                    Task.target_id.in_(ids),
                    Task.task_type == "remove_friend",
                    Task.status == TaskStatus.DONE.value,
                )
                .group_by(Task.target_id)
                .all()
            )
            revoke_map = {int(tid): int(cnt or 0) for tid, cnt in revoke_rows}
            remove_map = {int(tid): int(cnt or 0) for tid, cnt in remove_rows}
        camp = get_campaign_or_default(db, selected_campaign_id)
        camp_name = camp.name if camp else f"#{selected_campaign_id}"
        return (
            targets,
            total,
            pages,
            page_clamped,
            send_q_map,
            check_q_map,
            revoke_map,
            remove_map,
            friends_now_map,
            accepted_unique_map,
            sent_done_map,
            camp_name,
        )

    (
        targets,
        total,
        pages,
        page_clamped,
        send_q_map,
        check_q_map,
        revoke_map,
        remove_map,
        friends_now_map,
        accepted_unique_map,
        sent_done_map,
        camp_name,
    ) = db_exec(_inner)
    set_chat_ui_value(chat_id, "tgt_page", page_clamped)
    set_chat_ui_value(chat_id, "tgt_query", query)
    set_chat_ui_value(chat_id, "tgt_view_mode", "receiver_stats")
    set_chat_ui_value(chat_id, "goal_page_context", "targets_receiver_stats")
    set_current_menu(chat_id, "goal_nicks")

    if total == 0:
        suffix = f" по запросу {md_inline_code(query)}" if query else ""
        show_screen(chat_id, f"❌ Ников не найдено{suffix}", reply_markup=kb_goal_nicks_reply(), parse_mode="Markdown")
        return

    lines = [f"📄 **Статусы по никам {md_inline_code(camp_name)}** ({total})"]
    if query:
        lines.append(f"Фильтр: {md_inline_code(query)}")
    lines.append(f"Страница: {page_clamped}/{pages} (по {TARGETS_PAGE_SIZE})")
    lines.append("Листай: ◀️ Страница / ▶️ Страница")
    lines.append("")
    for t in targets:
        rejected = 1 if t.status == TargetStatus.REJECTED.value else 0
        accepted_unique = int(accepted_unique_map.get(int(t.id), 0))
        sent_total = max(
            int(t.sent_count or 0),
            int(sent_done_map.get(int(t.id), 0)),
            accepted_unique,
        )
        send_q = int(send_q_map.get(int(t.id), 0))
        check_q = int(check_q_map.get(int(t.id), 0))
        revoked = int(revoke_map.get(int(t.id), 0))
        removed = int(remove_map.get(int(t.id), 0))
        lines.append(f"• #{t.id} {md_inline_code(t.username)}")
        lines.append(f"  Статус: {target_status_ru(t.status)}")
        lines.append(
            f"  Отпр/Принято(уник)/В друзьях/Откл: "
            f"{sent_total}/{accepted_unique}/{int(friends_now_map.get(int(t.id), 0))}/{rejected}"
        )
        lines.append(f"  Очередь send/check: {send_q}/{check_q}")
        lines.append(f"  Отозвано/Удалено: {revoked}/{removed}")
        lines.append("")
    text = "\n".join(lines)
    if len(text) > 3800:
        text = text[:3790] + "…"
    show_screen(chat_id, text, reply_markup=kb_goal_nicks_reply(), parse_mode="Markdown")


def show_campaign_progress(chat_id: int, with_campaign_info: bool = False):
    selected_campaign_id = get_selected_campaign_id(chat_id)

    def _inner(db):
        camp = get_campaign_or_default(db, selected_campaign_id)
        now = utc_now()
        day_start, day_end = local_day_bounds_utc_naive(db, utc_now())
        targets = (
            db.query(Target.id, Target.status, Target.required_senders)
            .filter(target_campaign_filter(db, selected_campaign_id))
            .all()
        )
        status_counts = dict(
            db.query(Target.status, func.count(Target.id))
            .filter(target_campaign_filter(db, selected_campaign_id))
            .group_by(Target.status)
            .all()
        )
        send_rows = (
            db.query(Task.target_id, func.count(func.distinct(Task.account_id)))
            .filter(
                task_campaign_filter(db, selected_campaign_id),
                Task.task_type == "send_request",
                Task.status == TaskStatus.DONE.value,
            )
            .group_by(Task.target_id)
            .all()
        )
        send_rows_today = (
            db.query(Task.target_id, func.count(func.distinct(Task.account_id)))
            .filter(
                task_campaign_filter(db, selected_campaign_id),
                Task.task_type == "send_request",
                Task.status == TaskStatus.DONE.value,
                Task.completed_at >= day_start,
                Task.completed_at < day_end,
            )
            .group_by(Task.target_id)
            .all()
        )
        senders_map_total = {int(tid): int(cnt or 0) for tid, cnt in send_rows}
        senders_map_today = {int(tid): int(cnt or 0) for tid, cnt in send_rows_today}
        repeat_daily = bool(camp.daily_repeat_enabled) if camp is not None else False
        senders_map = senders_map_today if repeat_daily else senders_map_total
        total_send_done = (
            db.query(func.count(Task.id))
            .filter(
                task_campaign_filter(db, selected_campaign_id),
                Task.task_type == "send_request",
                Task.status == TaskStatus.DONE.value,
            )
            .scalar()
            or 0
        )
        total_send_done_today = (
            db.query(func.count(Task.id))
            .filter(
                task_campaign_filter(db, selected_campaign_id),
                Task.task_type == "send_request",
                Task.status == TaskStatus.DONE.value,
                Task.completed_at >= day_start,
                Task.completed_at < day_end,
            )
            .scalar()
            or 0
        )
        senders_total = (
            db.query(func.count(func.distinct(Task.account_id)))
            .filter(
                task_campaign_filter(db, selected_campaign_id),
                Task.task_type == "send_request",
                Task.status == TaskStatus.DONE.value,
            )
            .scalar()
            or 0
        )
        senders_today = (
            db.query(func.count(func.distinct(Task.account_id)))
            .filter(
                task_campaign_filter(db, selected_campaign_id),
                Task.task_type == "send_request",
                Task.status == TaskStatus.DONE.value,
                Task.completed_at >= day_start,
                Task.completed_at < day_end,
            )
            .scalar()
            or 0
        )
        send_queue = (
            db.query(func.count(Task.id))
            .filter(
                task_campaign_filter(db, selected_campaign_id),
                Task.task_type == "send_request",
                Task.status.in_(ACTIVE_SEND_TASK_STATUSES),
            )
            .scalar()
            or 0
        )
        check_queue = (
            db.query(func.count(Task.id))
            .filter(
                task_campaign_filter(db, selected_campaign_id),
                Task.task_type == "check_status",
                Task.status.in_(ACTIVE_SEND_TASK_STATUSES),
            )
            .scalar()
            or 0
        )
        accepted_today = (
            db.query(func.count(Task.id))
            .join(Target, Target.id == Task.target_id)
            .filter(
                task_campaign_filter(db, selected_campaign_id),
                Task.task_type == "check_status",
                Task.status == TaskStatus.DONE.value,
                Task.completed_at >= day_start,
                Task.completed_at < day_end,
                Target.status == TargetStatus.ACCEPTED.value,
            )
            .scalar()
            or 0
        )
        rejected_today = (
            db.query(func.count(Task.id))
            .join(Target, Target.id == Task.target_id)
            .filter(
                task_campaign_filter(db, selected_campaign_id),
                Task.task_type == "check_status",
                Task.status == TaskStatus.DONE.value,
                Task.completed_at >= day_start,
                Task.completed_at < day_end,
                Target.status == TargetStatus.REJECTED.value,
            )
            .scalar()
            or 0
        )
        awaiting_response = int(status_counts.get(TargetStatus.SENT.value, 0) or 0)
        waiting_delivery = int(status_counts.get(TargetStatus.PENDING.value, 0) or 0) + int(status_counts.get(TargetStatus.NEW.value, 0) or 0)

        ready_accounts = db.query(Account).filter(
            Account.enabled == True,
            Account.status == AccountStatus.ACTIVE.value,
            Account.epic_account_id.isnot(None),
            Account.device_id.isnot(None),
            Account.device_secret.isnot(None),
        ).all()
        ready_account_ids = [int(a.id) for a in ready_accounts]
        ready_count = len(ready_account_ids)
        used_sender_rows = (
            db.query(Task.account_id)
            .filter(
                task_campaign_filter(db, selected_campaign_id),
                Task.task_type == "send_request",
                Task.status == TaskStatus.DONE.value,
            )
            .distinct()
            .all()
        )
        used_sender_ids = {int(x[0]) for x in used_sender_rows}
        used_ready_count = len(set(ready_account_ids) & used_sender_ids)
        unused_ready_count = max(0, int(ready_count) - int(used_ready_count))
        at_limit_count = 0
        if camp is not None:
            eff_limit = _campaign_effective_daily_limit(db, camp, now)
            for a in ready_accounts:
                sent = campaign_sent_today_for_account(db, int(a.id), int(camp.id), now)
                if sent >= eff_limit:
                    at_limit_count += 1
        else:
            for a in ready_accounts:
                if int(a.today_sent or 0) >= int(a.daily_limit or 0):
                    at_limit_count += 1
        default_required = max(1, get_setting_int(db, "target_senders_count", DEFAULT_TARGET_SENDERS_COUNT))
        max_required_per_target = 0
        targets_with_sender_deficit = 0
        max_done_per_target = 0
        targets_pool_exhausted = 0
        for _, _, required_senders in targets:
            req = int(required_senders or 0) or default_required
            max_required_per_target = max(max_required_per_target, int(req))
            if int(req) > int(ready_count):
                targets_with_sender_deficit += 1
        for tid, _, _ in targets:
            done = int(senders_map.get(int(tid), 0))
            max_done_per_target = max(max_done_per_target, done)
            if int(ready_count) > 0 and done >= int(ready_count):
                targets_pool_exhausted += 1
        pool_layers_left = max(0, int(ready_count) - int(max_done_per_target))
        campaign_id = int(camp.id) if camp else 0
        campaign_num = campaign_ui_num(db, campaign_id) if camp else 0
        camp_name = camp.name if camp else "Текущая"
        return (
            campaign_id,
            campaign_num,
            targets,
            status_counts,
            senders_map,
            int(total_send_done),
            int(total_send_done_today),
            int(senders_total),
            int(senders_today),
            int(send_queue),
            int(check_queue),
            int(accepted_today),
            int(rejected_today),
            int(awaiting_response),
            int(waiting_delivery),
            int(ready_count),
            int(used_ready_count),
            int(unused_ready_count),
            int(at_limit_count),
            int(max_required_per_target),
            int(targets_with_sender_deficit),
            int(max_done_per_target),
            int(pool_layers_left),
            int(targets_pool_exhausted),
            default_required,
            camp_name,
            bool(repeat_daily),
        )

    (
        campaign_id,
        campaign_num,
        targets,
        status_counts,
        senders_map,
        total_send_done,
        total_send_done_today,
        senders_total,
        senders_today,
        send_queue,
        check_queue,
        accepted_today,
        rejected_today,
        awaiting_response,
        waiting_delivery,
        ready_count,
        used_ready_count,
        unused_ready_count,
        at_limit_count,
        max_required_per_target,
        targets_with_sender_deficit,
        max_done_per_target,
        pool_layers_left,
        targets_pool_exhausted,
        default_required,
        camp_name,
        repeat_daily,
    ) = db_exec(_inner)
    total_targets = len(targets)
    if total_targets == 0:
        if campaign_id > 0:
            title = f"📈 Прогресс цели №{campaign_num} {camp_name} (id:{campaign_id})"
        else:
            title = "📈 Прогресс текущей цели"
        empty_lines = [title, "Ники получателя пока не загружены."]
        if with_campaign_info:
            with SessionLocal() as db:
                c = db.query(Campaign).filter(Campaign.id == int(campaign_id)).first()
                if c:
                    empty_lines.append("")
                    empty_lines.append(_format_campaign_info(db, c))
        show_screen(
            chat_id,
            "\n".join(empty_lines),
            reply_markup=kb_goal_selected_reply(),
            parse_mode=None,
            force_new=True,
        )
        return

    required_total = 0
    covered_total = 0
    fully_covered = 0
    for tid, _, required_senders in targets:
        req = int(required_senders or 0) or default_required
        done = int(senders_map.get(int(tid), 0))
        required_total += req
        covered_total += min(done, req)
        if done >= req:
            fully_covered += 1

    remaining = max(0, required_total - covered_total)
    sender_deficit_per_nick = max(0, int(max_required_per_target) - int(ready_count))
    title = (
        f"📈 Прогресс цели №{campaign_num} {camp_name} (id:{campaign_id})"
        if campaign_id > 0
        else "📈 Прогресс текущей цели"
    )
    lines = [
        title,
        f"Целей: {total_targets}",
        f"Полностью покрыты: {fully_covered}/{total_targets}",
        f"Покрытие отправителей {'(сегодня)' if repeat_daily else '(всего)'}: {covered_total}/{required_total}",
        f"Осталось отправок для покрытия: {remaining}",
        "",
        "На ники получателя:",
        f"• Отправлено сегодня: {total_send_done_today}",
        f"• Отправлено всего: {total_send_done}",
        f"• Принято сегодня: {accepted_today}",
        f"• Отклонено сегодня: {rejected_today}",
        f"• Ожидают ответа (после отправки): {awaiting_response}",
        f"• Ожидают первичной отправки: {waiting_delivery}",
        f"• В очереди отправки: {send_queue}",
        f"• В очереди проверки: {check_queue}",
        "",
        "С аккаунтов-отправителей:",
        f"• Отправляли сегодня: {senders_today}",
        f"• Отправляли всего: {senders_total}",
        f"• Готовы к работе (active+device_auth): {ready_count}",
        f"• Уже использовались в цели: {used_ready_count}",
        f"• Ещё не использовались в цели: {unused_ready_count}",
        f"• Уперлись в дневной лимит: {at_limit_count}",
        "",
        "Ёмкость по цели:",
        f"• Макс. нужно отправителей на ник: {max_required_per_target}",
        f"• Ников с нехваткой отправителей: {targets_with_sender_deficit}",
        f"• Дефицит отправителей на ник (макс): {sender_deficit_per_nick}",
        "",
        "Запас отправителей (по текущему прогрессу):",
        f"• Самый покрытый ник: {max_done_per_target}/{ready_count} отправителей",
        f"• Осталось до исчерпания пула: {pool_layers_left}",
        f"• Ников уже уперлись в пул: {targets_pool_exhausted}",
        "",
        "Статусы целей:",
        f"🆕 новая={int(status_counts.get(TargetStatus.NEW.value, 0) or 0)}",
        f"⏳ в ожидании={int(status_counts.get(TargetStatus.PENDING.value, 0) or 0)}",
        f"📨 отправлено={int(status_counts.get(TargetStatus.SENT.value, 0) or 0)}",
        f"✅ принято={int(status_counts.get(TargetStatus.ACCEPTED.value, 0) or 0)}",
        f"❌ отклонено={int(status_counts.get(TargetStatus.REJECTED.value, 0) or 0)}",
        f"🚫 ошибка={int(status_counts.get(TargetStatus.FAILED.value, 0) or 0)}",
    ]
    if sender_deficit_per_nick > 0:
        lines.append(
            f"⚠️ Для полного покрытия добавь минимум {sender_deficit_per_nick} аккаунт(ов)-отправителей."
        )
    if with_campaign_info:
        with SessionLocal() as db:
            c = db.query(Campaign).filter(Campaign.id == int(campaign_id)).first()
            if c:
                lines.append("")
                lines.append(_format_campaign_info(db, c))
    show_screen(chat_id, "\n".join(lines), reply_markup=kb_goal_selected_reply(), parse_mode=None, force_new=True)


# ============================================================
# START + message handlers
# ============================================================

@bot.message_handler(commands=["start"])
@user_access_only
def cmd_start(message):
    if is_admin(message.from_user.id):
        show_main_menu(message.chat.id)
    else:
        show_auth_operator_menu(message.chat.id)


@bot.message_handler(commands=["stop_rel"])
@admin_only
def cmd_stop_relationship_action(message):
    running = stop_relationship_action(message.chat.id)
    if running:
        show_menu_status(message.chat.id, "targets", "⏳ Остановка операции принята. Завершу на ближайшем шаге.")
    else:
        show_menu_status(message.chat.id, "targets", "ℹ️ Сейчас нет активной операции отзыва/удаления.")


@bot.message_handler(commands=["clean"])
@admin_only
def cmd_clean(message):
    """
    Best-effort UI cleanup:
    - delete current screen message (bot)
    - delete transient notify messages (bot)
    Note: user messages in private chats cannot be deleted by the bot (Telegram limitation).
    """
    chat_id = message.chat.id
    _safe_delete_user_message(message)

    cleanup_transient_messages(chat_id)

    # delete current screen
    with CHAT_UI_LOCK:
        st = CHAT_UI.get(chat_id) or {}
        screen_id = st.get("screen_msg_id")
        st["screen_msg_id"] = None
        CHAT_UI[chat_id] = st
    if screen_id:
        _safe_delete(chat_id, screen_id)

    show_main_menu(chat_id)


@bot.callback_query_handler(func=lambda call: str(getattr(call, "data", "")).startswith("act:"))
@admin_only_call
def cb_inline_action_nav(call):
    _safe_answer_callback(call)
    data = str(call.data or "")
    code = data.split(":", 1)[1] if ":" in data else ""
    text = INLINE_ACTION_TEXT.get(code)
    if not text:
        return

    fake_message = py_types.SimpleNamespace(
        chat=call.message.chat,
        from_user=call.from_user,
        text=text,
        message_id=getattr(call.message, "message_id", None),
    )
    cmd_reply_nav(fake_message)


def _reply_call_from_message(message, data: str):
    return py_types.SimpleNamespace(id="reply", data=data, from_user=message.from_user, message=message)


def _invoke_reply_callback(message, data: str, handler):
    handler(_reply_call_from_message(message, data))


def _invoke_reply_callback_by_text(message, text: str, mapping: dict[str, tuple[str, object]]) -> bool:
    item = mapping.get(str(text or "").strip())
    if not item:
        return False
    data, handler = item
    _invoke_reply_callback(message, data, handler)
    return True


def _invoke_show_menu_by_text(chat_id: int, text: str, mapping: dict[str, object]) -> bool:
    fn = mapping.get(str(text or "").strip())
    if not fn:
        return False
    fn(chat_id)
    return True


def _handle_accounts_pager_and_search(message, text: str, chat_id: int) -> bool:
    if text == "◀️ Аккаунты":
        page = get_chat_ui_int(chat_id, "acc_page", 1) - 1
        query = str(get_chat_ui_value(chat_id, "acc_query", "") or "")
        show_accounts_list(chat_id, page=page, query=query)
        return True
    if text == "▶️ Аккаунты":
        page = get_chat_ui_int(chat_id, "acc_page", 1) + 1
        query = str(get_chat_ui_value(chat_id, "acc_query", "") or "")
        show_accounts_list(chat_id, page=page, query=query)
        return True
    if text == "🔎 Поиск аккаунтов":
        ask_step(
            message,
            "Введи часть логина (пусто = сброс):",
            handle_accounts_search_query,
        )
        return True
    return False


def _handle_auth_operator_nav(message, text: str, chat_id: int, callback_text_map: dict[str, tuple[str, object]]) -> bool:
    # Auth-operator mode: only auth-related account actions.
    if text in {"🏠 Меню", "⬅️ Назад"}:
        show_auth_operator_menu(chat_id)
        return True
    if _invoke_reply_callback_by_text(message, text, callback_text_map):
        return True
    if _handle_accounts_pager_and_search(message, text, chat_id):
        return True
    show_auth_operator_menu(chat_id)
    return True


def _handle_goal_edit_shortcuts(message, text: str, chat_id: int, current_menu: str) -> bool:
    if current_menu != "goal_edit":
        return False
    if text == "🔄 Лимит":
        cid = get_selected_campaign_id(chat_id)
        camp_info = db_exec(
            lambda db: db.query(Campaign.name, Campaign.daily_limit_per_account).filter(Campaign.id == cid).first()
        )
        current_limit = int((camp_info[1] if camp_info else 0) or 0)
        camp_name = (camp_info[0] if camp_info else f"#{cid}")
        ask_step(
            message,
            f"Текущая цель: {camp_name} (#{cid})\n"
            f"Сейчас лимит: {current_limit}/сутки на 1 аккаунт\n\n"
            "Введи новый лимит (>=1):",
            handle_set_goal_daily_limit,
        )
        return True
    if text == "⏱️ Джиттер":
        cid = get_selected_campaign_id(chat_id)
        camp_info = db_exec(
            lambda db: db.query(Campaign.name, Campaign.jitter_min_sec, Campaign.jitter_max_sec).filter(Campaign.id == cid).first()
        )
        camp_name = (camp_info[0] if camp_info else f"#{cid}")
        cur_min = int((camp_info[1] if camp_info else 0) or 0)
        cur_max = int((camp_info[2] if camp_info else 0) or 0)
        ask_step(
            message,
            f"Текущая цель: {camp_name} (#{cid})\n"
            f"Сейчас джиттер: {cur_min}-{cur_max} сек\n\n"
            "Введи новый джиттер: min_sec max_sec",
            handle_set_goal_jitter,
        )
        return True
    if text == "🕐 Окна":
        cid = get_selected_campaign_id(chat_id)
        camp_info = db_exec(
            lambda db: db.query(Campaign.name, Campaign.active_windows_json).filter(Campaign.id == cid).first()
        )
        camp_name = (camp_info[0] if camp_info else f"#{cid}")
        cur_windows = windows_human(camp_info[1] if camp_info else "[]")
        ask_step(
            message,
            f"Текущая цель: {camp_name} (#{cid})\n"
            f"Сейчас окна: {cur_windows}\n\n"
            "Введи окна: `24/7` или несколько строк.\n"
            "Пример будни/выходные:\n"
            "`days=1,2,3,4,5 from=12:00 to=20:00`\n"
            "`days=6,7 from=10:00 to=18:00`",
            handle_set_goal_windows,
        )
        return True
    if text == "🔀 Алгоритм отправки":
        cid = get_selected_campaign_id(chat_id)
        mode_ru = db_exec(
            lambda db: "Сначала отправитель -> все ники"
            if get_campaign_send_mode(db, cid) == "sender_first"
            else "Сначала ник -> все отправители"
        )
        ask_step(
            message,
            f"Текущая цель: #{cid}\n"
            f"Текущий алгоритм: {mode_ru}\n\n"
            "Выбери режим:\n"
            "1 — сначала отправитель, потом следующий\n"
            "2 — сначала ник, потом следующий",
            handle_set_goal_send_mode,
        )
        return True
    if text == "🎲 Порядок отправителей":
        cid = get_selected_campaign_id(chat_id)
        mode_ru = db_exec(
            lambda db: "По порядку (ID аккаунтов)"
            if get_campaign_sender_pick_mode(db, cid) == "ordered"
            else "Случайный порядок"
        )
        ask_step(
            message,
            f"Текущая цель: #{cid}\n"
            f"Текущий порядок отправителей: {mode_ru}\n\n"
            "Выбери режим:\n"
            "1 — по порядку аккаунтов\n"
            "2 — случайный порядок",
            handle_set_goal_sender_pick_mode,
        )
        return True
    if text == "📋 Параметры цели":
        show_selected_goal_params(chat_id)
        return True
    return False


def _handle_targets_goals_actions(message, text: str, chat_id: int) -> bool:
    if text == "👥 Ники":
        show_goal_nicks_menu(chat_id)
        return True
    if text == "🚀 Отправка":
        show_goal_sending_menu(chat_id)
        return True
    if text == "🧹 Операции":
        show_goal_ops_menu(chat_id)
        return True
    if text == "▶️ Запустить все цели":
        total, changed = set_all_goals_enabled(True)
        active_now = db_exec(lambda db: int(db.query(Campaign).filter(Campaign.enabled == True).count()))
        show_menu_status(
            chat_id,
            "targets",
            f"✅ Все цели запущены.\nАктивные цели: {active_now}/{total}\nИзменено: {changed}",
        )
        return True
    if text == "⏸️ Остановить все цели":
        total, changed = set_all_goals_enabled(False)
        active_now = db_exec(lambda db: int(db.query(Campaign).filter(Campaign.enabled == True).count()))
        show_menu_status(
            chat_id,
            "targets",
            f"✅ Все цели остановлены.\nАктивные цели: {active_now}/{total}\nИзменено: {changed}",
        )
        return True
    if text == "➕ Добавить цель":
        ask_step(
            message,
            "Создание цели.\nВведи имя:",
            handle_campaign_create_name,
        )
        return True
    if text == "🎯 Выбрать цель":
        ask_step(
            message,
            "Введи ID цели из списка:",
            handle_campaign_select,
        )
        return True
    if text == "📊 Статистика цели":
        show_campaign_progress(chat_id, with_campaign_info=True)
        return True
    if text == "🗑️ Удалить цель":
        ask_step(
            message,
            "Введи ID цели для удаления (или текущий ID).\nВнимание: будут удалены её ники и задачи.",
            handle_delete_goal_single,
        )
        return True
    if text == "⏸️ Остановить цель":
        cid = get_selected_campaign_id(chat_id)
        ok, msg = _set_campaign_enabled(cid, False)
        show_menu_status(chat_id, "targets", ("✅ " + msg) if ok else ("❌ " + msg))
        return True
    if text == "▶️ Запустить цель":
        cid = get_selected_campaign_id(chat_id)
        ok, msg = _set_campaign_enabled(cid, True)
        show_menu_status(chat_id, "targets", ("✅ " + msg) if ok else ("❌ " + msg))
        return True
    if text == "↩️ Отозвать заявки":
        ask_step(
            message,
            "Подтверждение: это отзовёт исходящие заявки по текущей цели.\n"
            "Введи: ОТОЗВАТЬ",
            handle_revoke_requests_confirm,
        )
        return True
    if text == "🔍 Проверить в друзьях":
        queued, skipped_active, skipped_no_auth = enqueue_goal_friend_presence_checks(chat_id)
        show_menu_status(
            chat_id,
            "targets",
            "✅ Проверка наличия в друзьях запланирована.\n"
            f"Создано check-задач: {queued}\n"
            f"Пропущено (уже есть активная check): {skipped_active}\n"
            f"Пропущено (нет auth/неактивные): {skipped_no_auth}",
        )
        return True
    if text == "🔁 Дослать отсутствующих":
        queued, skipped_connected, skipped_other = enqueue_goal_resend_missing(chat_id)
        show_menu_status(
            chat_id,
            "targets",
            "✅ Досыл отсутствующих запланирован.\n"
            f"Создано resend-задач: {queued}\n"
            f"Пропущено (уже accepted/pending): {skipped_connected}\n"
            f"Пропущено (активная задача/нет auth): {skipped_other}",
        )
        return True
    if text == "🗑️ Удалить из друзей":
        ask_step(
            message,
            "Подтверждение: это удалит из друзей по текущей цели.\n"
            "Введи: УДАЛИТЬ",
            handle_remove_friends_confirm,
        )
        return True
    if text == "⛔ Остановить операцию":
        running = stop_relationship_action(chat_id)
        if running:
            show_menu_status(chat_id, "targets", "⏳ Остановка операции принята. Завершу на ближайшем шаге.")
        else:
            show_menu_status(chat_id, "targets", "ℹ️ Сейчас нет активной операции отзыва/удаления.")
        return True
    if text == "📄 Статусы по никам":
        query = str(get_chat_ui_value(chat_id, "tgt_query", "") or "")
        page = get_chat_ui_int(chat_id, "tgt_page", 1)
        show_targets_receiver_stats(chat_id, page=page, query=query)
        return True
    if text in {"◀️ Ники", "▶️ Ники", "◀️ Отправители", "▶️ Отправители", "◀️ Страница", "▶️ Страница"}:
        delta = -1 if text.startswith("◀️") else 1
        page_ctx = str(get_chat_ui_value(chat_id, "goal_page_context", "") or "")
        if page_ctx == "target_senders":
            target_id = int(get_chat_ui_int(chat_id, "senders_target_id", 0))
            if target_id <= 0:
                show_menu_status(chat_id, "targets", "ℹ️ Сначала открой «👀 Отправители по нику».")
                return True
            page = get_chat_ui_int(chat_id, "senders_page", 1) + delta
            show_target_senders_page(chat_id, target_id=target_id, page=page)
            return True

        page = get_chat_ui_int(chat_id, "tgt_page", 1) + delta
        query = str(get_chat_ui_value(chat_id, "tgt_query", "") or "")
        mode = str(get_chat_ui_value(chat_id, "tgt_view_mode", "targets") or "targets")
        if page_ctx == "targets_receiver_stats" or mode == "receiver_stats":
            show_targets_receiver_stats(chat_id, page=page, query=query)
        else:
            show_targets_status(chat_id, page=page, query=query)
        return True
    if text == "🔎 Поиск ников":
        ask_step(
            message,
            "Введи часть ника получателя (пусто = сброс):",
            handle_targets_search_query,
        )
        return True
    if text == "🧹 Очистить все цели":
        ask_step(
            message,
            "⚠️ Это удалит ВСЕ ники текущей цели и связанные с ними задачи.\n"
            "Для подтверждения введи: ОЧИСТИТЬ",
            handle_clear_all_targets_confirm,
        )
        return True
    if text == "➕ Добавить ник":
        ask_step(
            message,
            "Добавление ника получателя.\nВведи ник:",
            handle_add_target_single,
        )
        return True
    if text == "➖ Удалить ник":
        ask_step(
            message,
            "Введи ID ника или ник.\nМожно списком: по одному в строке или через , ;",
            handle_delete_target_single,
        )
        return True
    if text == "👀 Отправители по нику":
        ask_step(
            message,
            "Покажу отправителей для ника получателя.\nВведи ID или ник:",
            handle_show_target_senders,
        )
        return True
    if text == "⚡ Форс-цикл с аккаунта":
        ask_step(
            message,
            "Введи ID аккаунта-отправителя для одного цикла по всем никам текущей цели:",
            handle_force_cycle_account,
        )
        return True
    if text == "🎲 Форс-цикл (рандом)":
        handle_force_cycle_random(message)
        return True
    return False


def _handle_settings_and_manage_actions(message, text: str, chat_id: int) -> bool:
    if text == "🛡️ API лимиты":
        cur = db_exec(
            lambda db: (
                get_setting_int(db, "min_request_interval_sec", DEFAULT_MIN_REQUEST_INTERVAL_SEC),
                get_setting_int(db, "hourly_api_limit", DEFAULT_HOURLY_API_LIMIT),
                get_setting_int(db, "daily_api_limit", DEFAULT_DAILY_API_LIMIT),
            )
        )
        ask_step(
            message,
            "Текущие API-лимиты на 1 аккаунт:\n"
            f"• Не чаще 1 запроса в {cur[0]} сек\n"
            f"• До {cur[1]} запросов в час\n"
            f"• До {cur[2]} запросов в сутки\n\n"
            "Введи новые значения в формате:\n"
            "`min_interval_sec hourly_limit daily_limit`\n"
            "Пример: `40 40 500`",
            handle_set_api_limits,
            parse_mode="Markdown",
        )
        return True
    if text == "👤 Доступ auth":
        show_auth_access_menu(chat_id)
        return True
    if text == "🧯 Новые заявки":
        def _toggle(db):
            cur = is_new_send_requests_enabled(db)
            set_new_send_requests_enabled(db, not cur)
            return bool(cur), bool(not cur), bool(is_recheck_only_mode_enabled(db))

        was_enabled, now_enabled, recheck_only = db_exec(_toggle)
        show_menu_status(
            chat_id,
            "settings",
            "✅ Режим новых заявок обновлён.\n"
            f"Было: {'включено' if was_enabled else 'выключено'}\n"
            f"Стало: {'включено' if now_enabled else 'выключено'}\n"
            f"Recheck-only: {'включен' if recheck_only else 'выключен'}",
        )
        return True
    if text == "♻️ Только recheck":
        def _toggle(db):
            cur = is_recheck_only_mode_enabled(db)
            new_val = not cur
            set_recheck_only_mode_enabled(db, new_val)
            # recheck-only подразумевает отключение новых заявок.
            if new_val:
                set_new_send_requests_enabled(db, False)
            return bool(cur), bool(new_val), bool(is_new_send_requests_enabled(db))

        was_enabled, now_enabled, sends_enabled = db_exec(_toggle)
        show_menu_status(
            chat_id,
            "settings",
            "✅ Режим recheck-only обновлён.\n"
            f"Было: {'включен' if was_enabled else 'выключен'}\n"
            f"Стало: {'включен' if now_enabled else 'выключен'}\n"
            f"Новые заявки: {'включены' if sends_enabled else 'выключены'}",
        )
        return True
    if text == "➕ Добавить ID auth":
        ask_step(
            message,
            "Введи Telegram ID для добавления в доступ auth:",
            handle_auth_operator_add_id,
        )
        return True
    if text == "➖ Удалить ID auth":
        ask_step(
            message,
            "Введи Telegram ID для удаления из доступа auth:",
            handle_auth_operator_remove_id,
        )
        return True
    if text == "📋 Список ID auth":
        show_auth_access_menu(chat_id)
        return True
    if text == "🗑 Очистить ID auth":
        ask_step(
            message,
            "Подтверждение очистки списка auth-ID.\n"
            "Введи `ОЧИСТИТЬ`, чтобы удалить все ID auth-операторов.",
            handle_auth_operator_clear_all_confirm,
            parse_mode="Markdown",
        )
        return True
    if text == "📌 Прокси":
        set_current_menu(chat_id, "proxy")
        show_screen(chat_id, "📌 Прокси", reply_markup=kb_proxy_reply(), parse_mode=None, force_new=True)
        return True
    if text == "⏸️ Стоп обработки":
        set_processing_enabled(False)
        show_menu_status(chat_id, "manage", "⏸️ Обработка остановлена.")
        return True
    if text == "▶️ Старт обработки":
        set_processing_enabled(True)
        show_menu_status(chat_id, "manage", "▶️ Обработка запущена.")
        return True
    if text == "📍 Статус обработки":
        show_processing_status(chat_id)
        return True
    return False


def _handle_accounts_actions(message, text: str, chat_id: int) -> bool:
    if text == "📥 Импорт файлов":
        show_menu_status(chat_id, "accounts", "📥 Пришли .xlsx/.txt/.csv файлом в чат.")
        return True
    if text == "📝 Массовая смена ников":
        show_nickname_change_import_menu(chat_id)
        return True
    if text == "📊 Статус смены ников":
        show_nickname_change_status(chat_id)
        return True
    if text == "➕ Добавить аккаунт":
        ask_step(
            message,
            "Добавление аккаунта-отправителя.\nФормат: login:password",
            handle_add_account_single,
        )
        return True
    if text == "➖ Удалить аккаунт":
        ask_step(
            message,
            "Введи ID или login для удаления.\nМожно списком: по одному в строке или через , ;",
            handle_delete_account_single,
        )
        return True
    return False


def _show_menu_by_current_context(chat_id: int, full_admin: bool, current_menu: str):
    if not full_admin:
        show_auth_operator_menu(chat_id)
        return
    fn_map = {
        "main": show_main_menu,
        "accounts": show_accounts_menu,
        "nick_change_import": show_nickname_change_import_menu,
        "targets": show_targets_menu,
        "goal_manager": show_goal_manager_menu,
        "goal_selected": show_selected_goal_menu,
        "goal_nicks": show_goal_nicks_menu,
        "goal_sending": show_goal_sending_menu,
        "goal_ops": show_goal_ops_menu,
        "goal_edit": show_goal_edit_menu,
        "settings": show_settings_menu,
        "auth_access": show_auth_access_menu,
        "manage": show_manage_menu,
        "proxy": lambda cid: show_screen(cid, "📌 Прокси", reply_markup=kb_proxy_reply(), parse_mode=None, force_new=True),
    }
    fn = fn_map.get(current_menu, show_main_menu)
    fn(chat_id)


def _dispatch_admin_nav(
    message,
    text: str,
    chat_id: int,
    current_menu: str,
    callback_text_map: dict[str, tuple[str, object]],
    show_menu_text_map: dict[str, object],
    back_by_menu_map: dict[str, object],
) -> bool:
    if _handle_accounts_pager_and_search(message, text, chat_id):
        return True

    if text == "⬅️ Назад":
        fn = back_by_menu_map.get(current_menu, show_main_menu)
        fn(chat_id)
        return True

    if text == "🏠 Меню":
        show_main_menu(chat_id)
        return True
    if _invoke_show_menu_by_text(chat_id, text, show_menu_text_map):
        return True

    if _handle_accounts_actions(message, text, chat_id):
        return True
    if _invoke_reply_callback_by_text(message, text, callback_text_map):
        return True
    if _handle_targets_goals_actions(message, text, chat_id):
        return True
    if _handle_goal_edit_shortcuts(message, text, chat_id, current_menu):
        return True
    if _handle_settings_and_manage_actions(message, text, chat_id):
        return True

    return False


def _nav_callback_text_map() -> dict[str, tuple[str, object]]:
    return {
        "📋 Список аккаунтов": ("acc_list", cb_acc_list),
        "✅ Проверить аккаунты": ("acc_verify", cb_acc_verify),
        "🔄 Обновить ники Epic": ("acc_refresh_names", cb_acc_refresh_names),
        "🔐 Авторизация Epic (ссылка)": ("acc_device_auto", cb_acc_device_auto),
        "📥 Импорт ников": ("tgt_import", cb_tgt_import),
        "📋 Ники цели": ("tgt_status", cb_tgt_status),
        "🚀 Распределить ники": ("tgt_distribute", cb_tgt_distribute),
        "🎯 На ник": ("set_target_senders", cb_set_target_senders),
        "🔁 Лимит повторных проверок": ("set_recheck_limit", cb_set_recheck_limit),
        "📅 Ежедневный повтор": ("set_daily_repeat", cb_set_daily_repeat),
        "➕ Добавить прокси": ("proxy_add", cb_proxy_add),
        "📋 Список прокси": ("proxy_list", cb_proxy_list),
        "🗑️ Удалить прокси": ("proxy_delete", cb_proxy_delete),
        "▶️ Тик (1 раз)": ("manage_tick", cb_manage_tick),
        "📤 Экспорт": ("manage_export", cb_manage_export),
    }


def _nav_show_menu_text_map() -> dict[str, object]:
    return {
        "📊 Статистика": show_stats_screen,
        "👥 Аккаунты": show_accounts_menu,
        "🎯 Цели": show_targets_menu,
        "⚙️ Настройки": show_settings_menu,
        "🔧 Управление": show_manage_menu,
        "⚠️ Диагностика": show_diagnostics_screen,
        "🗂️ Менеджер целей": show_goal_manager_menu,
        "📊 Статистика целей": show_goals_info_all,
        "📋 Список целей": show_campaigns_list,
        "✏️ Редактировать цель": show_goal_edit_menu,
    }


def _nav_back_by_menu_map() -> dict[str, object]:
    return {
        "accounts": show_main_menu,
        "nick_change_import": show_accounts_menu,
        "targets": show_main_menu,
        "goal_manager": show_targets_menu,
        "goal_selected": show_goal_manager_menu,
        "goal_nicks": show_selected_goal_menu,
        "goal_sending": show_selected_goal_menu,
        "goal_ops": show_selected_goal_menu,
        "goal_edit": show_selected_goal_menu,
        "settings": show_main_menu,
        "auth_access": show_settings_menu,
        "manage": show_main_menu,
        "proxy": show_settings_menu,
    }


@bot.message_handler(func=lambda m: bool(getattr(m, "text", None)))
@user_access_only
def cmd_reply_nav(message):
    text = (message.text or "").strip()
    chat_id = message.chat.id
    current_menu = get_current_menu(chat_id)
    full_admin = is_admin(message.from_user.id)
    callback_text_map = _nav_callback_text_map()
    show_menu_text_map = _nav_show_menu_text_map()
    back_by_menu_map = _nav_back_by_menu_map()

    if not full_admin:
        if _handle_auth_operator_nav(message, text, chat_id, callback_text_map):
            return

    if _dispatch_admin_nav(
        message,
        text,
        chat_id,
        current_menu,
        callback_text_map,
        show_menu_text_map,
        back_by_menu_map,
    ):
        return

    # Fallback: unknown text -> return user to current menu context.
    _show_menu_by_current_context(chat_id, full_admin, current_menu)


def handle_accounts_search_query(message):
    cleanup_step(message)
    query = (message.text or "").strip()
    show_accounts_list(message.chat.id, page=1, query=query)


def handle_targets_search_query(message):
    cleanup_step(message)
    query = (message.text or "").strip()
    chat_id = int(message.chat.id)
    mode = str(get_chat_ui_value(chat_id, "tgt_view_mode", "targets") or "targets")
    if mode == "receiver_stats":
        show_targets_receiver_stats(chat_id, page=1, query=query)
    else:
        show_targets_status(chat_id, page=1, query=query)


CAMPAIGN_WIZARD_STATE = {}


def show_campaigns_list(chat_id: int):
    def _inner(db):
        rows = db.query(Campaign).order_by(Campaign.id.asc()).all()
        items = []
        for idx, c in enumerate(rows, start=1):
            targets_cnt = db.query(Target).filter(Target.campaign_id == c.id).count()
            items.append((idx, c, int(targets_cnt)))
        return items

    items = db_exec(_inner)
    if not items:
        show_screen(chat_id, "Целей нет.", reply_markup=kb_goal_manager_reply(), parse_mode=None, force_new=True)
        return
    lines = ["📊 Статистика целей:"]
    for idx, c, targets_cnt in items:
        st = "🟢 активна" if c.enabled else "⏸️ остановлена"
        lines.append(
            f"№{idx} {c.name} (id:{c.id}) — {st}, целей={targets_cnt}, "
            f"лимит/акк={c.daily_limit_per_account}, на цель={c.target_senders_count}"
        )
    show_screen(chat_id, "\n".join(lines), reply_markup=kb_goal_manager_reply(), parse_mode=None, force_new=True)


def show_goals_info_all(chat_id: int):
    # Reuse goals list view as compact global stats by goals.
    show_campaigns_list(chat_id)


def set_all_goals_enabled(enabled: bool) -> tuple[int, int]:
    def _inner(db):
        rows = db.query(Campaign).all()
        changed = 0
        for c in rows:
            if bool(c.enabled) != bool(enabled):
                c.enabled = bool(enabled)
                c.updated_at = utc_now()
                changed += 1
        db.commit()
        return len(rows), changed
    return db_exec(_inner)


def _format_campaign_info(db, c: Campaign) -> str:
    targets_cnt = db.query(Target).filter(Target.campaign_id == c.id).count()
    num = campaign_ui_num(db, int(c.id))
    mode = get_campaign_send_mode(db, int(c.id))
    mode_ru = "Сначала отправитель -> все ники" if mode == "sender_first" else "Сначала ник -> все отправители"
    sender_pick_mode = get_campaign_sender_pick_mode(db, int(c.id))
    sender_pick_ru = "По порядку (ID аккаунтов)" if sender_pick_mode == "ordered" else "Случайный порядок"
    return (
        f"📄 Цель №{num}: {c.name} (id:{c.id})\n"
        f"Статус: {'активна' if c.enabled else 'остановлена'}\n"
        f"Целей: {targets_cnt}\n"
        f"Лимит/акк/сутки: {c.daily_limit_per_account}\n"
        f"Отправителей на цель: {c.target_senders_count}\n"
        f"Окна: {windows_human(c.active_windows_json or '[]')}\n"
        f"Джиттер: {c.jitter_min_sec}-{c.jitter_max_sec} сек\n"
        f"Алгоритм: {mode_ru}\n"
        f"Порядок отправителей: {sender_pick_ru}\n"
        f"Лимит повторных проверок: {c.recheck_daily_limit}\n"
        f"Ежедневный повтор: {'включен' if c.daily_repeat_enabled else 'выключен'}"
    )


def handle_campaign_create_name(message):
    cleanup_step(message)
    name = (message.text or "").strip()
    if not name:
        show_screen(message.chat.id, "❌ Имя не может быть пустым.", reply_markup=kb_goal_manager_reply(), parse_mode=None, force_new=True)
        return
    key = (message.chat.id, message.from_user.id)
    CAMPAIGN_WIZARD_STATE[key] = {"name": name}
    ask_step(
        message,
        "Шаг 2/6. Лимит на 1 аккаунт-отправитель в сутки:",
        handle_campaign_create_daily_limit,
    )


def handle_campaign_create_daily_limit(message):
    cleanup_step(message)
    key = (message.chat.id, message.from_user.id)
    st = CAMPAIGN_WIZARD_STATE.get(key) or {}
    try:
        val = int((message.text or "").strip())
        if val < 1:
            raise ValueError()
    except Exception:
        show_screen(message.chat.id, "❌ Нужен положительный лимит.", reply_markup=kb_goal_manager_reply(), parse_mode=None, force_new=True)
        CAMPAIGN_WIZARD_STATE.pop(key, None)
        return
    st["daily_limit_per_account"] = val
    CAMPAIGN_WIZARD_STATE[key] = st
    ask_step(
        message,
        "Шаг 3/6. Сколько аккаунтов-отправителей на 1 ник получателя:",
        handle_campaign_create_senders,
    )


def handle_campaign_create_senders(message):
    cleanup_step(message)
    key = (message.chat.id, message.from_user.id)
    st = CAMPAIGN_WIZARD_STATE.get(key) or {}
    try:
        val = int((message.text or "").strip())
        if val < 1:
            raise ValueError()
    except Exception:
        show_screen(message.chat.id, "❌ Нужны число >= 1.", reply_markup=kb_goal_manager_reply(), parse_mode=None, force_new=True)
        CAMPAIGN_WIZARD_STATE.pop(key, None)
        return
    st["target_senders_count"] = val
    CAMPAIGN_WIZARD_STATE[key] = st
    ask_step(
        message,
        "Шаг 4/6. Окна работы цели.\n"
        "Введи `24/7` или несколько строк вида:\n"
        "`days=1,2,3,4,5 from=12:00 to=20:00`\n"
        "`days=6,7 from=10:00 to=18:00`",
        handle_campaign_create_windows,
    )


def handle_campaign_create_windows(message):
    cleanup_step(message)
    key = (message.chat.id, message.from_user.id)
    st = CAMPAIGN_WIZARD_STATE.get(key) or {}
    raw = (message.text or "").strip()
    if raw.lower() in {"24/7", "24x7", "always"}:
        st["active_windows_json"] = "[]"
    else:
        try:
            st["active_windows_json"] = json.dumps(parse_windows_text(raw), ensure_ascii=False)
        except Exception as e:
            show_screen(message.chat.id, f"❌ Ошибка окна: {e}", reply_markup=kb_goal_manager_reply(), parse_mode=None, force_new=True)
            CAMPAIGN_WIZARD_STATE.pop(key, None)
            return
    CAMPAIGN_WIZARD_STATE[key] = st
    ask_step(
        message,
        "Шаг 5/6. Лимит повторных проверок в сутки:",
        handle_campaign_create_recheck,
    )


def handle_campaign_create_recheck(message):
    cleanup_step(message)
    key = (message.chat.id, message.from_user.id)
    st = CAMPAIGN_WIZARD_STATE.get(key) or {}
    try:
        val = int((message.text or "").strip())
        if val < 0:
            raise ValueError()
    except Exception:
        show_screen(message.chat.id, "❌ Нужен лимит повторных проверок >= 0.", reply_markup=kb_goal_manager_reply(), parse_mode=None, force_new=True)
        CAMPAIGN_WIZARD_STATE.pop(key, None)
        return
    st["recheck_daily_limit"] = val
    CAMPAIGN_WIZARD_STATE[key] = st
    ask_step(
        message,
        "Шаг 6/6. Ежедневный повтор цели.\n1 = да, 0 = нет.",
        handle_campaign_create_repeat,
    )


def handle_campaign_create_repeat(message):
    cleanup_step(message)
    key = (message.chat.id, message.from_user.id)
    st = CAMPAIGN_WIZARD_STATE.pop(key, None) or {}
    raw = (message.text or "").strip().lower()
    if raw not in {"0", "1", "yes", "no", "true", "false"}:
        show_screen(message.chat.id, "❌ Введи 1 или 0.", reply_markup=kb_goal_manager_reply(), parse_mode=None, force_new=True)
        return
    st["daily_repeat_enabled"] = raw in {"1", "yes", "true"}

    def _inner(db):
        if db.query(Campaign).filter(Campaign.name == st["name"]).first():
            return None, "Цель с таким именем уже есть"
        c = Campaign(
            name=st["name"],
            enabled=False,
            daily_limit_per_account=int(st["daily_limit_per_account"]),
            target_senders_count=int(st["target_senders_count"]),
            active_windows_json=st.get("active_windows_json", "[]"),
            jitter_min_sec=DEFAULT_SEND_JITTER_MIN_SEC,
            jitter_max_sec=DEFAULT_SEND_JITTER_MAX_SEC,
            recheck_daily_limit=int(st["recheck_daily_limit"]),
            daily_repeat_enabled=bool(st["daily_repeat_enabled"]),
            updated_at=utc_now(),
        )
        db.add(c)
        db.commit()
        db.refresh(c)
        set_campaign_send_mode(db, int(c.id), "sender_first")
        set_campaign_sender_pick_mode(db, int(c.id), "ordered")
        return int(c.id), None

    cid, err = db_exec(_inner)
    if err:
        show_screen(message.chat.id, f"❌ {err}", reply_markup=kb_goal_manager_reply(), parse_mode=None, force_new=True)
        return
    set_chat_ui_value(message.chat.id, "selected_campaign_id", int(cid))
    show_menu_status(
        message.chat.id,
        "goal_selected",
        f"✅ Цель создана и остановлена.\nТеперь она выбрана в интерфейсе: id:{cid}.",
    )


def handle_campaign_select(message):
    cleanup_step(message)
    try:
        cid = int((message.text or "").strip())
    except Exception:
        show_screen(message.chat.id, "❌ Нужен числовой ID.", reply_markup=kb_goal_manager_reply(), parse_mode=None, force_new=True)
        return
    def _inner(db):
        return db.query(Campaign).filter(Campaign.id == cid).first()
    c = db_exec(_inner)
    if not c:
        show_screen(message.chat.id, "❌ Цель не найдена.", reply_markup=kb_goal_manager_reply(), parse_mode=None, force_new=True)
        return
    set_chat_ui_value(message.chat.id, "selected_campaign_id", int(c.id))
    show_menu_status(message.chat.id, "goal_selected", f"✅ Выбрана цель #{c.id}: {c.name}")


def _set_campaign_enabled(cid: int, enabled: bool) -> tuple[bool, str]:
    def _inner(db):
        c = db.query(Campaign).filter(Campaign.id == cid).first()
        if not c:
            return False, "Цель не найдена"
        c.enabled = bool(enabled)
        c.updated_at = utc_now()
        db.commit()
        return True, f"Цель #{c.id} {c.name}: {'запущена' if enabled else 'остановлена'}"
    return db_exec(_inner)


def handle_set_goal_daily_limit(message):
    cleanup_step(message)
    try:
        limit = int((message.text or "").strip())
        if limit < 1:
            raise ValueError()
    except Exception:
        show_menu_status(message.chat.id, "targets", "❌ Нужен лимит >= 1")
        return

    cid = get_selected_campaign_id(message.chat.id)

    def _inner(db):
        camp = get_campaign_or_default(db, cid)
        if not camp:
            return False, 0, "", 0, 0
        prev = int(camp.daily_limit_per_account or 0)
        camp.daily_limit_per_account = int(limit)
        # Apply new per-goal daily limit immediately (do not wait next day cache rollover).
        set_setting(db, f"camp_daily_limit_day_{int(camp.id)}", "")
        set_setting(db, f"camp_daily_limit_val_{int(camp.id)}", str(int(limit)))
        camp.updated_at = utc_now()
        db.commit()
        dropped_queue, created_tasks = rebuild_campaign_send_queue(db, int(camp.id), create_limit=5000)
        return True, prev, camp.name, dropped_queue, created_tasks

    ok, prev, camp_name, dropped_queue, created_tasks = db_exec(_inner)
    if ok:
        show_menu_status(
            message.chat.id,
            "targets",
            f"✅ Лимит цели «{camp_name}» обновлён:\n"
            f"Было: {prev}/сутки\n"
            f"Стало: {limit}/сутки на 1 аккаунт\n"
            f"Очищено старых задач: {dropped_queue}\n"
            f"Добавлено недостающих задач: {created_tasks}",
        )
    else:
        show_menu_status(message.chat.id, "targets", "❌ Цель не найдена")


def handle_set_goal_jitter(message):
    cleanup_step(message)
    try:
        mn, mx = map(int, (message.text or "").strip().split())
        if mn < 0 or mx < 0:
            raise ValueError()
        if mx < mn:
            mx = mn
    except Exception:
        show_menu_status(message.chat.id, "targets", "❌ Формат: min_sec max_sec")
        return

    cid = get_selected_campaign_id(message.chat.id)

    def _inner(db):
        camp = get_campaign_or_default(db, cid)
        if not camp:
            return False, 0, 0, "", 0, 0
        prev_min = int(camp.jitter_min_sec or 0)
        prev_max = int(camp.jitter_max_sec or 0)
        camp.jitter_min_sec = int(mn)
        camp.jitter_max_sec = int(mx)
        camp.updated_at = utc_now()
        db.commit()
        dropped_queue, created_tasks = rebuild_campaign_send_queue(db, int(camp.id), create_limit=5000)
        return True, prev_min, prev_max, camp.name, dropped_queue, created_tasks

    ok, prev_min, prev_max, camp_name, dropped_queue, created_tasks = db_exec(_inner)
    if ok:
        show_menu_status(
            message.chat.id,
            "targets",
            f"✅ Джиттер цели «{camp_name}» обновлён:\n"
            f"Было: {prev_min}-{prev_max} сек\n"
            f"Стало: {mn}-{mx} сек\n"
            f"Очищено старых задач: {dropped_queue}\n"
            f"Добавлено недостающих задач: {created_tasks}",
        )
    else:
        show_menu_status(message.chat.id, "targets", "❌ Цель не найдена")


def handle_set_goal_windows(message):
    cleanup_step(message)
    raw = (message.text or "").strip()
    try:
        if raw.lower() in {"24/7", "24x7", "always"}:
            windows_json = "[]"
        else:
            windows_json = json.dumps(parse_windows_text(raw), ensure_ascii=False)
    except Exception as e:
        show_menu_status(message.chat.id, "targets", f"❌ Ошибка окна: {e}")
        return

    cid = get_selected_campaign_id(message.chat.id)

    def _inner(db):
        camp = get_campaign_or_default(db, cid)
        if not camp:
            return False, "[]", "", 0, 0
        prev = camp.active_windows_json or "[]"
        camp.active_windows_json = windows_json
        camp.updated_at = utc_now()
        db.commit()
        dropped_queue, created_tasks = rebuild_campaign_send_queue(db, int(camp.id), create_limit=5000)
        return True, prev, camp.name, dropped_queue, created_tasks

    ok, prev, camp_name, dropped_queue, created_tasks = db_exec(_inner)
    if ok:
        show_menu_status(
            message.chat.id,
            "targets",
            f"✅ Окна цели «{camp_name}» обновлены:\n"
            f"Было: {windows_human(prev)}\n"
            f"Стало: {windows_human(windows_json)}\n"
            f"Очищено старых задач: {dropped_queue}\n"
            f"Добавлено недостающих задач: {created_tasks}",
        )
    else:
        show_menu_status(message.chat.id, "targets", "❌ Цель не найдена")


def show_selected_goal_params(chat_id: int):
    cid = get_selected_campaign_id(chat_id)

    def _inner(db):
        c = get_campaign_or_default(db, cid)
        if not c:
            return None
        return _format_campaign_info(db, c)

    info = db_exec(_inner)
    if not info:
        show_menu_status(chat_id, "targets", "❌ Цель не найдена.")
        return
    show_menu_status(chat_id, "targets", info)


def handle_set_goal_send_mode(message):
    cleanup_step(message)
    raw = str(message.text or "").strip().lower()
    if raw in {"1", "sender_first", "отправитель"}:
        mode = "sender_first"
        mode_ru = "Сначала отправитель -> все ники"
    elif raw in {"2", "target_first", "ник"}:
        mode = "target_first"
        mode_ru = "Сначала ник -> все отправители"
    else:
        show_menu_status(message.chat.id, "targets", "❌ Введи 1 или 2.")
        return

    cid = get_selected_campaign_id(message.chat.id)
    def _inner(db):
        set_campaign_send_mode(db, cid, mode)
        return rebuild_campaign_send_queue(db, int(cid), create_limit=5000)
    dropped_queue, created_tasks = db_exec(_inner)
    show_menu_status(
        message.chat.id,
        "targets",
        f"✅ Алгоритм цели обновлён:\n{mode_ru}\n"
        f"Очищено старых задач: {dropped_queue}\n"
        f"Добавлено недостающих задач: {created_tasks}",
    )


def handle_set_goal_sender_pick_mode(message):
    cleanup_step(message)
    raw = str(message.text or "").strip().lower()
    if raw in {"1", "ordered", "порядок", "по порядку"}:
        mode = "ordered"
        mode_ru = "По порядку (ID аккаунтов)"
    elif raw in {"2", "random", "случайно", "рандом"}:
        mode = "random"
        mode_ru = "Случайный порядок"
    else:
        show_menu_status(message.chat.id, "targets", "❌ Введи 1 или 2.")
        return

    cid = get_selected_campaign_id(message.chat.id)
    def _inner(db):
        set_campaign_sender_pick_mode(db, cid, mode)
        return rebuild_campaign_send_queue(db, int(cid), create_limit=5000)
    dropped_queue, created_tasks = db_exec(_inner)
    show_menu_status(
        message.chat.id,
        "targets",
        f"✅ Порядок отправителей обновлён:\n{mode_ru}\n"
        f"Очищено старых задач: {dropped_queue}\n"
        f"Добавлено недостающих задач: {created_tasks}",
    )


def _create_manual_force_cycle_for_account(db, camp: Campaign, acc: Account) -> dict:
    targets = (
        db.query(Target)
        .filter(target_campaign_filter(db, int(camp.id)))
        .order_by(Target.id.asc())
        .all()
    )
    if not targets:
        return {"ok": False, "msg": "В цели нет ников."}

    now = utc_now()
    jitter_min = max(0, int(camp.jitter_min_sec or DEFAULT_SEND_JITTER_MIN_SEC))
    jitter_max = max(jitter_min, int(camp.jitter_max_sec or DEFAULT_SEND_JITTER_MAX_SEC))
    scheduled = now
    created = 0
    skipped_active = 0
    skipped_done = 0
    skipped_connected = 0
    replaced_connected = 0

    for t in targets:
        has_active = db.query(Task.id).filter(
            Task.task_type == "send_request",
            Task.campaign_id == int(camp.id),
            Task.account_id == int(acc.id),
            Task.target_id == int(t.id),
            Task.status.in_(ACTIVE_SEND_TASK_STATUSES),
        ).first()
        if has_active:
            skipped_active += 1
            continue

        already_sent = db.query(Task.id).filter(
            Task.task_type == "send_request",
            Task.campaign_id == int(camp.id),
            Task.account_id == int(acc.id),
            Task.target_id == int(t.id),
            Task.status == TaskStatus.DONE.value,
        ).first()
        if already_sent:
            skipped_done += 1
            continue

        known_connected = db.query(Task.id).filter(
            Task.campaign_id == int(camp.id),
            Task.account_id == int(acc.id),
            Task.target_id == int(t.id),
            or_(
                and_(
                    Task.task_type == "check_status",
                    Task.status == TaskStatus.DONE.value,
                    Task.last_error.in_(["friend_status:accepted", "friend_status:pending"]),
                ),
                and_(
                    Task.task_type == "send_request",
                    Task.last_error.in_(PRECHECK_SKIP_REASONS),
                ),
            ),
        ).first()
        if known_connected:
            req = target_required_senders(db, t)
            done_sender_ids = _done_sender_ids_for_target(db, int(t.id), camp, now)
            replaced = False
            if len(done_sender_ids) < int(req):
                replaced = _enqueue_replacement_send_task(
                    db,
                    target=t,
                    camp=camp,
                    now=now,
                    excluded_ids={int(acc.id)},
                    source_reason="replacement_after_manual_connected",
                )
            if replaced:
                replaced_connected += 1
            else:
                skipped_connected += 1
            continue

        db.add(
            Task(
                task_type="send_request",
                status=TaskStatus.QUEUED.value,
                campaign_id=int(camp.id),
                account_id=int(acc.id),
                target_id=int(t.id),
                scheduled_for=scheduled,
                max_attempts=int(t.max_attempts or 3),
                last_error="manual_forced_cycle",
            )
        )
        created += 1
        scheduled = scheduled + timedelta(seconds=random.randint(int(jitter_min), int(jitter_max)))

    db.commit()
    return {
        "ok": True,
        "created": int(created),
        "skipped_active": int(skipped_active),
        "skipped_done": int(skipped_done),
        "skipped_connected": int(skipped_connected),
        "replaced_connected": int(replaced_connected),
        "targets": int(len(targets)),
        "camp_name": str(camp.name),
    }


def handle_force_cycle_account(message):
    cleanup_step(message)
    try:
        account_id = int(str(message.text or "").strip())
    except Exception:
        show_menu_status(message.chat.id, "targets", "❌ Нужен числовой ID аккаунта.")
        return

    cid = get_selected_campaign_id(message.chat.id)

    def _inner(db):
        camp = get_campaign_or_default(db, cid)
        if not camp:
            return {"ok": False, "msg": "Цель не найдена."}
        acc = db.query(Account).filter(Account.id == int(account_id)).first()
        if not acc:
            return {"ok": False, "msg": f"Аккаунт #{account_id} не найден."}
        if not acc.enabled or acc.status != AccountStatus.ACTIVE.value:
            return {"ok": False, "msg": f"Аккаунт #{account_id} неактивен."}
        if not (acc.epic_account_id and acc.device_id and acc.device_secret):
            return {"ok": False, "msg": f"У аккаунта #{account_id} нет device auth."}

        result = _create_manual_force_cycle_for_account(db, camp, acc)
        if result.get("ok"):
            result["account_id"] = int(acc.id)
        return result

    res = db_exec(_inner)
    if not res.get("ok"):
        show_menu_status(message.chat.id, "targets", f"❌ {res.get('msg', 'Ошибка форс-цикла.')}")
        return
    show_menu_status(
        message.chat.id,
        "targets",
        f"✅ Форс-цикл запланирован.\n"
        f"Цель: {res['camp_name']}\n"
        f"Аккаунт: #{account_id}\n"
        f"Всего ников: {res['targets']}\n"
        f"Создано задач: {res['created']}\n"
        f"Пропущено (уже в очереди): {res['skipped_active']}\n"
        f"Пропущено (уже отправлял): {res.get('skipped_done', 0)}\n"
        f"Заменено на других отправителей: {res.get('replaced_connected', 0)}\n"
        f"Пропущено (уже в друзьях/ожидании): {res.get('skipped_connected', 0)}",
    )


def handle_force_cycle_random(message):
    chat_id = message.chat.id
    cid = get_selected_campaign_id(chat_id)

    def _inner(db):
        camp = get_campaign_or_default(db, cid)
        if not camp:
            return {"ok": False, "msg": "Цель не найдена."}

        candidate_ids = []
        accs = (
            db.query(Account)
            .filter(
                Account.enabled == True,
                Account.status == AccountStatus.ACTIVE.value,
                Account.epic_account_id.isnot(None),
                Account.device_id.isnot(None),
                Account.device_secret.isnot(None),
            )
            .order_by(Account.id.asc())
            .all()
        )
        for acc in accs:
            has_pending_any = db.query(Task.id).filter(
                Task.task_type == "send_request",
                Task.campaign_id == int(camp.id),
                Task.account_id == int(acc.id),
                Task.status.in_(ACTIVE_SEND_TASK_STATUSES),
            ).first()
            if has_pending_any:
                continue
            candidate_ids.append(int(acc.id))

        if not candidate_ids:
            return {"ok": False, "msg": "Нет доступных аккаунтов для рандомного форс-цикла."}

        random.shuffle(candidate_ids)
        chosen = None
        for aid in candidate_ids:
            acc = db.query(Account).filter(Account.id == int(aid)).first()
            result = _create_manual_force_cycle_for_account(db, camp, acc)
            if result.get("ok") and int(result.get("created", 0)) > 0:
                result["account_id"] = int(aid)
                return result
            if chosen is None:
                chosen = result

        if chosen and chosen.get("ok"):
            chosen["account_id"] = int(candidate_ids[0])
            return chosen
        return {"ok": False, "msg": "Не удалось запланировать форс-цикл."}

    res = db_exec(_inner)
    if not res.get("ok"):
        show_menu_status(chat_id, "targets", f"❌ {res.get('msg', 'Ошибка форс-цикла.')}")
        return
    show_menu_status(
        chat_id,
        "targets",
        f"✅ Форс-цикл (рандом) запланирован.\n"
        f"Цель: {res['camp_name']}\n"
        f"Выбран аккаунт: #{res['account_id']}\n"
        f"Всего ников: {res['targets']}\n"
        f"Создано задач: {res['created']}\n"
        f"Пропущено (уже в очереди): {res['skipped_active']}\n"
        f"Пропущено (уже отправлял): {res.get('skipped_done', 0)}\n"
        f"Заменено на других отправителей: {res.get('replaced_connected', 0)}\n"
        f"Пропущено (уже в друзьях/ожидании): {res.get('skipped_connected', 0)}",
    )


def handle_delete_goal_single(message):
    cleanup_step(message)
    raw = (message.text or "").strip()
    if not raw:
        show_menu_status(message.chat.id, "targets", "❌ Нужен ID цели.")
        return
    try:
        cid = int(raw)
    except Exception:
        show_menu_status(message.chat.id, "targets", "❌ Нужен числовой ID цели.")
        return

    def _inner(db):
        c = db.query(Campaign).filter(Campaign.id == cid).first()
        if not c:
            return False, "Цель не найдена"
        tgt_ids = [int(x[0]) for x in db.query(Target.id).filter(Target.campaign_id == int(c.id)).all()]
        if tgt_ids:
            db.query(Task).filter(Task.target_id.in_(tgt_ids)).delete(synchronize_session=False)
            db.query(Target).filter(Target.id.in_(tgt_ids)).delete(synchronize_session=False)
        db.delete(c)
        db.commit()
        return True, f"Цель #{cid} удалена"

    ok, text = db_exec(_inner)
    if ok:
        next_cid = db_exec(lambda db: (db.query(Campaign.id).order_by(Campaign.id.asc()).first() or (0,))[0])
        set_chat_ui_value(message.chat.id, "selected_campaign_id", int(next_cid or 0))
        show_menu_status(message.chat.id, "goal_manager", "✅ " + text)
    else:
        show_menu_status(message.chat.id, "targets", "❌ " + text)


def handle_clear_all_targets_confirm(message):
    cleanup_step(message)
    confirm = (message.text or "").strip().upper()
    if confirm != "ОЧИСТИТЬ":
        show_menu_status(message.chat.id, "targets", "❌ Отменено. Для очистки нужно ввести ровно: ОЧИСТИТЬ")
        return

    selected_campaign_id = get_selected_campaign_id(message.chat.id)

    def _inner(db):
        filt = target_campaign_filter(db, selected_campaign_id)
        target_ids = [
            x[0]
            for x in db.query(Target.id).filter(filt).all()
        ]
        deleted_tasks = 0
        if target_ids:
            deleted_tasks = (
                db.query(Task)
                .filter(Task.target_id.in_(target_ids))
                .delete(synchronize_session=False)
            )
        deleted_targets = (
            db.query(Target)
            .filter(filt)
            .delete(synchronize_session=False)
        )
        db.commit()
        return int(deleted_targets or 0), int(deleted_tasks or 0)

    deleted_targets, deleted_tasks = db_exec(_inner)
    show_menu_status(
        message.chat.id,
        "targets",
        f"✅ Полная очистка завершена:\nУдалено ников получателя: {deleted_targets}\nУдалено задач по ним: {deleted_tasks}",
    )



# ============================================================
# DEVICE AUTH (авто режим) + ОТМЕНА
# ============================================================

def _make_device_auth_markup(account_id: int, url: str) -> types.InlineKeyboardMarkup:
    m = types.InlineKeyboardMarkup(row_width=2)
    # URL button doesn't create user messages; opens in browser.
    if url:
        m.add(types.InlineKeyboardButton("🔗 Открыть ссылку", url=url))
    m.add(
        types.InlineKeyboardButton("📋 Логин", callback_data=f"acc_device_show_login:{account_id}"),
        types.InlineKeyboardButton("🔐 Пароль", callback_data=f"acc_device_show_pass:{account_id}"),
    )
    m.add(types.InlineKeyboardButton("✖️ Отмена ожидания", callback_data=f"acc_device_cancel:{account_id}"))
    return m


def _cancel_pending_device(account_id: int) -> bool:
    with PENDING_DEVICE_LOCK:
        pending = PENDING_DEVICE.get(account_id)
        if not pending:
            return False
        ev = pending.get("cancel")
        if isinstance(ev, threading.Event):
            ev.set()
        return True


def _normalize_identity_value(value) -> str:
    return str(value or "").strip().lower()


def _validate_device_auth_identity(db, selected_acc: Account, result) -> tuple[bool, str, int | None]:
    """
    Verify that OAuth login result belongs to the selected sender account.
    Returns: (ok, reason_code, matched_account_id)
    """
    oauth_email_norm = _normalize_identity_value(getattr(result, "email", ""))
    selected_login_norm = _normalize_identity_value(getattr(selected_acc, "login", ""))
    oauth_epic_id = str(getattr(result, "epic_account_id", "") or "").strip()
    selected_epic_id = str(getattr(selected_acc, "epic_account_id", "") or "").strip()

    if oauth_epic_id:
        existing_by_epic = (
            db.query(Account)
            .filter(Account.epic_account_id == oauth_epic_id, Account.id != selected_acc.id)
            .first()
        )
        if existing_by_epic:
            return False, "epic_account_already_bound", int(existing_by_epic.id)

    if oauth_email_norm:
        all_accounts = db.query(Account.id, Account.login).all()
        for row in all_accounts:
            row_login_norm = _normalize_identity_value(row.login)
            if row_login_norm == oauth_email_norm and int(row.id) != int(selected_acc.id):
                return False, "email_already_bound", int(row.id)

    if oauth_email_norm and selected_login_norm and oauth_email_norm != selected_login_norm:
        return False, "email_mismatch", None

    if selected_epic_id and oauth_epic_id and selected_epic_id != oauth_epic_id:
        return False, "epic_id_mismatch", None

    return True, "", None


def _start_device_auth_worker(chat_id: int, account_id: int, login: str):
    async def run():
        link_msg_id = None
        cancel_event = threading.Event()

        try:
            async with EpicDeviceAuthGenerator() as gen:
                url, device_code = await gen.create_login_link()

                link_msg = bot.send_message(
                    chat_id,
                    f"Аккаунт #{account_id} ({login})\n\n"
                    f"1) Открой ссылку и войди в Epic:\n{url}\n\n"
                    f"2) Если в этом браузере УЖЕ залогинен нужный аккаунт Epic, останется только Confirm.\n"
                    f"Если нет, Epic попросит логин/пароль.\n\n"
                    f"3) Я сам дождусь авторизации, сохраню device_auth и удалю это сообщение.\n"
                    f"Если передумал — нажми «Отмена ожидания».",
                    disable_web_page_preview=True,
                    reply_markup=_make_device_auth_markup(account_id, url),
                )
                link_msg_id = link_msg.message_id

                with PENDING_DEVICE_LOCK:
                    PENDING_DEVICE[account_id] = {
                        "chat_id": chat_id,
                        "link_msg_id": link_msg_id,
                        "login": login,
                        "cancel": cancel_event,
                    }

                login_task = asyncio.create_task(gen.complete_login(device_code))
                cancel_task = asyncio.create_task(asyncio.to_thread(cancel_event.wait))

                done, pending = await asyncio.wait(
                    {login_task, cancel_task},
                    timeout=DEVICE_AUTH_TIMEOUT_SEC,
                    return_when=asyncio.FIRST_COMPLETED,
                )

                if not done:
                    cancel_event.set()
                    for t in pending:
                        t.cancel()
                    if link_msg_id:
                        _safe_delete(chat_id, link_msg_id)
                    notify(chat_id, f"⌛️ Таймаут ожидания авторизации для #{account_id}.", ttl_sec=90)
                    return

                if cancel_task in done and cancel_event.is_set():
                    login_task.cancel()
                    if link_msg_id:
                        _safe_delete(chat_id, link_msg_id)
                    notify(chat_id, f"❌ Ожидание авторизации для #{account_id} отменено.", ttl_sec=45)
                    return

                cancel_event.set()
                cancel_task.cancel()
                result = login_task.result()

            # Optional JSON export of device_auth (DB is the source of truth).
            # Leave unset in production to avoid writing secrets to the project directory.
            device_auths_json_path = os.getenv("DEVICE_AUTHS_JSON_PATH", "").strip()
            if device_auths_json_path:
                append_device_auth_to_file(result.email, result.raw, path=device_auths_json_path)

            def inner_db(db):
                acc = db.query(Account).filter(Account.id == account_id).first()
                if not acc:
                    return {"ok": False, "code": "account_missing"}
                identity_ok, identity_code, matched_acc_id = _validate_device_auth_identity(db, acc, result)
                if not identity_ok:
                    acc.last_error = f"device_auth_identity_mismatch:{identity_code}"
                    db.commit()
                    return {
                        "ok": False,
                        "code": identity_code,
                        "matched_acc_id": matched_acc_id,
                        "selected_login": acc.login,
                    }
                acc.epic_account_id = result.epic_account_id
                acc.epic_display_name = (result.display_name or "").strip() or acc.epic_display_name
                acc.device_id = result.device_id
                acc.device_secret = result.device_secret
                acc.last_error = None
                if acc.status == AccountStatus.MANUAL.value:
                    acc.status = AccountStatus.ACTIVE.value
                db.commit()
                return {"ok": True, "code": "saved", "selected_login": acc.login}

            save_res = db_exec(inner_db)

            if link_msg_id:
                _safe_delete(chat_id, link_msg_id)

            if save_res.get("ok"):
                log_event("info", f"device_auth_saved acc#{account_id} login={login} email={result.email}")
                notify(chat_id, f"✅ Device auth сохранён для #{account_id} ({result.email}).", ttl_sec=180)
            elif save_res.get("code") in {
                "email_mismatch",
                "epic_id_mismatch",
                "email_already_bound",
                "epic_account_already_bound",
            }:
                code = save_res.get("code")
                matched = save_res.get("matched_acc_id")
                selected_login = save_res.get("selected_login") or login
                auth_email = (result.email or "").strip() or "неизвестно"
                auth_name = (result.display_name or "").strip() or "без ника"
                tail = ""
                if matched:
                    tail = f"\nЭтот Epic уже привязан к аккаунту #{matched}."
                log_event(
                    "warning",
                    f"device_auth_identity_mismatch acc#{account_id} login={selected_login} "
                    f"auth_email={auth_email} code={code} matched={matched}",
                )
                notify(
                    chat_id,
                    "❌ Авторизован другой Epic-аккаунт. Сохранение отменено.\n"
                    f"Выбранный аккаунт: #{account_id} ({selected_login})\n"
                    f"Фактически авторизован: {auth_email} ({auth_name}){tail}\n"
                    "Повтори авторизацию и войди под нужным аккаунтом Epic.",
                    ttl_sec=180,
                )
            else:
                log_event("warning", f"device_auth_saved_but_account_missing acc#{account_id} login={login} email={result.email}")
                notify(chat_id, f"⚠️ Device auth получен, но аккаунт #{account_id} не найден в БД.", ttl_sec=120)

        except Exception as e:
            log_event("error", f"device_auth_error acc#{account_id} login={login}: {e}")
            if link_msg_id:
                _safe_delete(chat_id, link_msg_id)
            notify(chat_id, f"❌ Ошибка device_auth для #{account_id}: {e}", ttl_sec=120)
        finally:
            with PENDING_DEVICE_LOCK:
                PENDING_DEVICE.pop(account_id, None)

    asyncio.run(run())


# ============================================================
# CALLBACK HANDLERS (всё стараемся показывать через show_screen)
# ============================================================
@bot.callback_query_handler(func=lambda call: call.data == "noop")
def cb_noop(call):
    _safe_answer_callback(call)

@bot.callback_query_handler(func=lambda call: str(getattr(call, "data", "") or "").startswith("tgt_senders:"))
@admin_only_call
def cb_tgt_senders_pager(call):
    _safe_answer_callback(call)
    try:
        _, tid, page = str(call.data).split(":", 2)
        target_id = int(tid)
        page = int(page)
    except Exception:
        notify(call.message.chat.id, "❌ Неверная команда пагинации.", ttl_sec=10)
        return

    show_target_senders_page(call.message.chat.id, target_id=target_id, page=page)

@bot.callback_query_handler(func=lambda call: call.data == "acc_import")
@admin_only_call
def cb_acc_import(call):
    _safe_answer_callback(call)
    show_menu_status(
        call.message.chat.id,
        "accounts",
        "📥 Пришли .xlsx/.txt/.csv файлом в чат. Импорт сработает автоматически.",
    )


@bot.callback_query_handler(func=lambda call: call.data == "acc_list")
@admin_only_call
def cb_acc_list(call):
    _safe_answer_callback(call)
    chat_id = call.message.chat.id
    show_accounts_list(
        chat_id,
        page=get_chat_ui_int(chat_id, "acc_page", 1),
        query=str(get_chat_ui_value(chat_id, "acc_query", "") or ""),
    )


@bot.callback_query_handler(func=lambda call: call.data == "acc_banned")
@admin_only_call
def cb_acc_banned(call):
    _safe_answer_callback(call)

    def _inner(db):
        return db.query(Account).filter(Account.status == AccountStatus.BANNED.value).order_by(Account.id.asc()).limit(50).all()

    accs = db_exec(_inner)
    if not accs:
        show_screen(call.message.chat.id, "✅ Забаненных нет", reply_markup=kb_accounts_reply(), parse_mode=None)
        return

    text = f"❌ **Забанено ({len(accs)}):**\n\n" + "\n".join(
        [f"#{a.id} {md_inline_code(a.login)}" for a in accs]
    )

    show_screen(call.message.chat.id, text, reply_markup=kb_accounts_reply(), parse_mode="Markdown")


@bot.callback_query_handler(func=lambda call: call.data == "acc_paused")
@admin_only_call
def cb_acc_paused(call):
    _safe_answer_callback(call)

    def _inner(db):
        return db.query(Account).filter(Account.status == AccountStatus.PAUSED.value).order_by(Account.id.asc()).limit(50).all()

    accs = db_exec(_inner)
    if not accs:
        show_screen(call.message.chat.id, "✅ На паузе нет", reply_markup=kb_accounts_reply(), parse_mode=None)
        return

    text = f"⏸️ **На паузе ({len(accs)}):**\n\n" + "\n".join(
        [f"#{a.id} {md_inline_code(a.login)}" for a in accs]
    )

    show_screen(call.message.chat.id, text, reply_markup=kb_accounts_reply(), parse_mode="Markdown")


@bot.callback_query_handler(func=lambda call: call.data == "acc_verify")
@admin_only_call
def cb_acc_verify(call):
    _safe_answer_callback(call)

    chat_id = call.message.chat.id
    show_menu_status(chat_id, "accounts", "⏳ Проверяю аккаунты...")

    def worker():
        try:
            verify_accounts_health_job()
            show_menu_status(chat_id, "accounts", "✅ Проверка завершена")
        except Exception as e:
            show_menu_status(chat_id, "accounts", f"❌ Ошибка проверки: {e}")

    threading.Thread(target=worker, daemon=True).start()


@bot.callback_query_handler(func=lambda call: call.data == "acc_refresh_names")
@admin_only_call
def cb_acc_refresh_names(call):
    _safe_answer_callback(call)

    chat_id = call.message.chat.id
    show_menu_status(chat_id, "accounts", "⏳ Обновляю ники Epic у авторизованных аккаунтов...")

    def worker():
        try:
            checked, updated, failed = refresh_accounts_display_names_job(limit=0)
            show_menu_status(
                chat_id,
                "accounts",
                "✅ Обновление ников завершено.\n"
                f"Проверено: {checked}\n"
                f"Обновлено: {updated}\n"
                f"Ошибок: {failed}",
            )
        except Exception as e:
            show_menu_status(chat_id, "accounts", f"❌ Ошибка обновления ников: {e}")

    threading.Thread(target=worker, daemon=True).start()

@bot.callback_query_handler(func=lambda call: call.data == "acc_device_auto")
@admin_only_call
def cb_acc_device_auto(call):
    _safe_answer_callback(call)

    # просим ID отдельным prompt-ом (и чистим его после ответа)
    dummy = py_types.SimpleNamespace(chat=call.message.chat, from_user=call.from_user)
    ask_step(
        dummy,
        "Авторизация Epic по ссылке.\nВведи ID аккаунта из списка:",
        handle_device_auto_id,
    )


def handle_device_auto_id(message):
    cleanup_step(message)

    try:
        acc_id = int((message.text or "").strip())
    except Exception:
        show_menu_status(message.chat.id, "accounts", "❌ Нужен числовой ID аккаунта.")
        return

    def inner_db(db):
        return db.query(Account).filter(Account.id == acc_id).first()

    acc = db_exec(inner_db)
    if not acc:
        show_menu_status(message.chat.id, "accounts", f"❌ Аккаунт с ID {acc_id} не найден.")
        return

    with PENDING_DEVICE_LOCK:
        if acc_id in PENDING_DEVICE:
            show_menu_status(message.chat.id, "accounts", f"⚠️ Для #{acc_id} уже идёт ожидание авторизации.")
            return

    show_menu_status(message.chat.id, "accounts", f"⏳ Генерирую ссылку для #{acc_id} ({acc.login})…")

    t = threading.Thread(
        target=_start_device_auth_worker,
        args=(message.chat.id, acc_id, acc.login),
        daemon=True,
    )
    t.start()


@bot.callback_query_handler(func=lambda call: call.data.startswith("acc_device_cancel:"))
@admin_only_call
def cb_acc_device_cancel(call):
    _safe_answer_callback(call)

    try:
        acc_id = int(call.data.split(":", 1)[1])
    except Exception:
        show_menu_status(call.message.chat.id, "accounts", "❌ Неверная команда.")
        return

    existed = _cancel_pending_device(acc_id)
    if not existed:
        show_menu_status(call.message.chat.id, "accounts", "ℹ️ Ожидание уже завершено")
        return

    # удаляем сообщение со ссылкой сразу (воркер тоже попробует — это ок)
    _safe_delete(call.message.chat.id, call.message.message_id)


@bot.callback_query_handler(func=lambda call: call.data.startswith("acc_device_show_login:"))
@admin_only_call
def cb_acc_device_show_login(call):
    _safe_answer_callback(call)

    try:
        acc_id = int(call.data.split(":", 1)[1])
    except Exception:
        notify(call.message.chat.id, "❌ Неверная команда.", ttl_sec=10, parse_mode=None)
        return

    def _inner(db):
        acc = db.query(Account).filter(Account.id == acc_id).first()
        return acc.login if acc else None

    login = db_exec(_inner)
    if not login:
        notify(call.message.chat.id, "❌ Аккаунт не найден.", ttl_sec=10, parse_mode=None)
        return

    notify(call.message.chat.id, f"Логин: {login}", ttl_sec=25, parse_mode=None)


@bot.callback_query_handler(func=lambda call: call.data.startswith("acc_device_show_pass:"))
@admin_only_call
def cb_acc_device_show_pass(call):
    _safe_answer_callback(call)

    if not ALLOW_SHOW_PASSWORD:
        notify(
            call.message.chat.id,
            "🔐 Показ пароля отключён. Включи ALLOW_SHOW_PASSWORD=1 в .env (на свой риск).",
            ttl_sec=25,
            parse_mode=None,
        )
        return

    try:
        acc_id = int(call.data.split(":", 1)[1])
    except Exception:
        notify(call.message.chat.id, "❌ Неверная команда.", ttl_sec=10, parse_mode=None)
        return

    def _inner(db):
        acc = db.query(Account).filter(Account.id == acc_id).first()
        return acc.password if acc else None

    pwd = db_exec(_inner)
    if not pwd:
        notify(call.message.chat.id, "❌ Аккаунт не найден.", ttl_sec=10, parse_mode=None)
        return

    notify(call.message.chat.id, f"Пароль: {pwd}", ttl_sec=18, parse_mode=None)


@bot.callback_query_handler(func=lambda call: call.data == "tgt_import")
@admin_only_call
def cb_tgt_import(call):
    _safe_answer_callback(call)

    dummy = py_types.SimpleNamespace(chat=call.message.chat, from_user=call.from_user)
    ask_step(
        dummy,
        "Импорт ников получателя в текущую цель.\nПо одному нику в строке:",
        handle_import_targets,
    )


def handle_import_targets(message):
    cleanup_step(message)

    names = safe_split_lines(message.text)
    if not names:
        show_menu_status(message.chat.id, "targets", "❌ Пустой список.")
        return

    selected_campaign_id = get_selected_campaign_id(message.chat.id)
    if selected_campaign_id <= 0:
        show_menu_status(message.chat.id, "targets", "❌ Нет цели. Сначала создай и выбери цель.")
        return

    def _inner(db):
        added = 0
        duplicate_in_payload = 0
        seen_names = set()
        unique_names = []
        for name in names:
            if name in seen_names:
                duplicate_in_payload += 1
                continue
            seen_names.add(name)
            unique_names.append(name)

        if not unique_names:
            return 0, duplicate_in_payload, 0, []

        camp = get_campaign_or_default(db, selected_campaign_id)
        if not camp:
            return 0, duplicate_in_payload, 0, []
        required = max(1, int(camp.target_senders_count or DEFAULT_TARGET_SENDERS_COUNT))
        effective_campaign_id = int(camp.id)
        existing_names = {
            x[0]
            for x in db.query(Target.username)
            .filter(
                Target.username.in_(unique_names),
                target_campaign_filter(db, effective_campaign_id),
            )
            .all()
        }
        already_in_db = 0
        for name in unique_names:
            if name in existing_names:
                already_in_db += 1
                continue
            db.add(
                Target(
                    username=name,
                    campaign_id=effective_campaign_id,
                    status=TargetStatus.NEW.value,
                    priority=random.randint(1, 100),
                    required_senders=required,
                )
            )
            added += 1
        db.commit()
        sample_existing = sorted(existing_names)[:10]
        return added, duplicate_in_payload, already_in_db, sample_existing

    added, duplicate_in_payload, already_in_db, sample_existing = db_exec(_inner)
    lines = [
        "✅ Импорт ников получателя завершён",
        f"Добавлено: {added}",
        f"Дубликаты в этом списке: {duplicate_in_payload}",
        f"Уже были в базе: {already_in_db}",
    ]
    if sample_existing:
        lines.append("Примеры уже существующих:")
        lines.extend([f"- {x}" for x in sample_existing[:5]])
    show_menu_status(message.chat.id, "targets", "\n".join(lines))


def handle_add_account_single(message):
    cleanup_step(message)
    parsed = _parse_account_line((message.text or "").strip())
    if not parsed:
        show_menu_status(message.chat.id, "accounts", "❌ Неверный формат. Используй: login:password.")
        return
    login, password = parsed

    def _inner(db):
        if db.query(Account).filter(Account.login == login).first():
            return False
        db.add(
            Account(
                login=login,
                password=password,
                daily_limit=DEFAULT_DAILY_LIMIT,
                active_windows_json="[]",
                warmup_until=utc_now() + timedelta(minutes=5),
            )
        )
        db.commit()
        return True

    ok = db_exec(_inner)
    show_menu_status(message.chat.id, "accounts", "✅ Аккаунт добавлен." if ok else "❌ Такой аккаунт уже существует.")


def handle_delete_account_single(message):
    cleanup_step(message)
    items = split_multi_values(message.text or "")
    if not items:
        show_menu_status(message.chat.id, "accounts", "❌ Пустой ввод.")
        return

    def _inner(db):
        deleted = 0
        missing = 0
        for raw in items:
            acc = None
            if raw.isdigit():
                acc = db.query(Account).filter(Account.id == int(raw)).first()
            if not acc:
                acc = db.query(Account).filter(Account.login == raw).first()
            if not acc:
                missing += 1
                continue
            db.query(Task).filter(Task.account_id == acc.id).delete(synchronize_session=False)
            db.delete(acc)
            deleted += 1
        db.commit()
        return deleted, missing

    deleted, missing = db_exec(_inner)
    if deleted == 0:
        show_menu_status(message.chat.id, "accounts", "❌ Аккаунты не найдены.")
        return
    show_menu_status(
        message.chat.id,
        "accounts",
        f"✅ Удалено аккаунтов: {deleted}" + (f"\nНе найдено: {missing}" if missing else ""),
    )


def handle_add_target_single(message):
    cleanup_step(message)
    name = (message.text or "").strip()
    if not name:
        show_menu_status(message.chat.id, "targets", "❌ Пустой ник.")
        return

    selected_campaign_id = get_selected_campaign_id(message.chat.id)
    if selected_campaign_id <= 0:
        show_menu_status(message.chat.id, "targets", "❌ Нет цели. Сначала создай и выбери цель.")
        return

    def _inner(db):
        camp = get_campaign_or_default(db, selected_campaign_id)
        if not camp:
            return None
        effective_campaign_id = int(camp.id)
        if db.query(Target).filter(
            Target.username == name,
            target_campaign_filter(db, effective_campaign_id),
        ).first():
            return False
        if camp:
            required = max(1, int(camp.target_senders_count or DEFAULT_TARGET_SENDERS_COUNT))
        else:
            required = max(1, get_setting_int(db, "target_senders_count", DEFAULT_TARGET_SENDERS_COUNT))
        db.add(
            Target(
                username=name,
                campaign_id=effective_campaign_id,
                status=TargetStatus.NEW.value,
                priority=random.randint(1, 100),
                required_senders=required,
            )
        )
        db.commit()
        return True

    ok = db_exec(_inner)
    if ok is None:
        show_menu_status(message.chat.id, "targets", "❌ Цель не найдена. Выбери цель заново.")
    else:
        show_menu_status(message.chat.id, "targets", "✅ Цель добавлена." if ok else "❌ Такая цель уже есть.")


def handle_delete_target_single(message):
    cleanup_step(message)
    items = split_multi_values(message.text or "")
    if not items:
        show_menu_status(message.chat.id, "targets", "❌ Пустой ввод.")
        return

    selected_campaign_id = get_selected_campaign_id(message.chat.id)

    def _inner(db):
        filt = target_campaign_filter(db, selected_campaign_id)
        deleted = 0
        missing = 0
        for raw in items:
            tgt = None
            if raw.isdigit():
                tgt = db.query(Target).filter(Target.id == int(raw), filt).first()
            if not tgt:
                tgt = db.query(Target).filter(Target.username == raw, filt).first()
            if not tgt:
                missing += 1
                continue
            db.query(Task).filter(Task.target_id == tgt.id).delete(synchronize_session=False)
            db.delete(tgt)
            deleted += 1
        db.commit()
        return deleted, missing

    deleted, missing = db_exec(_inner)
    if deleted == 0:
        show_menu_status(message.chat.id, "targets", "❌ Цели не найдены.")
        return
    show_menu_status(
        message.chat.id,
        "targets",
        f"✅ Удалено ников получателя: {deleted}" + (f"\nНе найдено: {missing}" if missing else ""),
    )


def handle_show_target_senders(message):
    cleanup_step(message)
    raw = (message.text or "").strip()
    if not raw:
        show_menu_status(message.chat.id, "targets", "❌ Пустой ввод.")
        return

    # Ввод: "<ник_or_id>" или "<ник_or_id> <page>"
    parts = raw.split()
    query = parts[0].strip()
    try:
        page = int(parts[1]) if len(parts) >= 2 else 1
    except Exception:
        page = 1
    if page < 1:
        page = 1

    selected_campaign_id = get_selected_campaign_id(message.chat.id)

    def _inner(db):
        filt = target_campaign_filter(db, selected_campaign_id)
        tgt = None
        if query.isdigit():
            tgt = db.query(Target).filter(Target.id == int(query), filt).first()
        if not tgt:
            tgt = db.query(Target).filter(Target.username == query, filt).first()
        return int(tgt.id) if tgt else 0

    target_id = int(db_exec(_inner) or 0)
    if target_id <= 0:
        show_menu_status(message.chat.id, "targets", "❌ Цель не найдена.")
        return

    # Вся пагинация/иконки/имена — внутри show_target_senders_page
    show_target_senders_page(message.chat.id, target_id=target_id, page=page)

def _goal_sender_target_pairs(db, campaign_id: int) -> list[tuple[int, int, str]]:
    rows = (
        db.query(Task.account_id, Target.id, Target.username)
        .join(Target, Target.id == Task.target_id)
        .filter(
            task_campaign_filter(db, campaign_id),
            Task.task_type == "send_request",
            Task.status == TaskStatus.DONE.value,
        )
        .distinct()
        .all()
    )
    return [
        (int(account_id), int(target_id), str(username))
        for account_id, target_id, username in rows
        if account_id and target_id and username
    ]


def enqueue_goal_friend_presence_checks(chat_id: int) -> tuple[int, int, int]:
    cid = get_selected_campaign_id(chat_id)

    def _inner(db):
        camp = get_campaign_or_default(db, cid)
        if not camp:
            return 0, 0, 0
        pairs = _goal_sender_target_pairs(db, int(camp.id))
        if not pairs:
            return 0, 0, 0
        now = utc_now()
        queued = 0
        skipped_active = 0
        skipped_no_auth = 0
        for account_id, target_id, _ in pairs:
            has_active = db.query(Task.id).filter(
                Task.task_type == "check_status",
                Task.campaign_id == int(camp.id),
                Task.account_id == int(account_id),
                Task.target_id == int(target_id),
                Task.status.in_(ACTIVE_SEND_TASK_STATUSES),
            ).first()
            if has_active:
                skipped_active += 1
                continue
            acc = db.query(Account).filter(Account.id == int(account_id)).first()
            if not acc or not acc.enabled or acc.status != AccountStatus.ACTIVE.value:
                skipped_no_auth += 1
                continue
            if not (acc.epic_account_id and acc.device_id and acc.device_secret):
                skipped_no_auth += 1
                continue
            min_s = max(0, int(camp.jitter_min_sec or DEFAULT_SEND_JITTER_MIN_SEC))
            max_s = max(min_s, int(camp.jitter_max_sec or DEFAULT_SEND_JITTER_MAX_SEC))
            db.add(
                Task(
                    task_type="check_status",
                    status=TaskStatus.QUEUED.value,
                    campaign_id=int(camp.id),
                    account_id=int(account_id),
                    target_id=int(target_id),
                    scheduled_for=now + timedelta(seconds=random.randint(min_s, max_s)),
                    max_attempts=5,
                    last_error="manual_friend_presence_check",
                )
            )
            queued += 1
        db.commit()
        return int(queued), int(skipped_active), int(skipped_no_auth)

    return db_exec(_inner)


def enqueue_goal_resend_missing(chat_id: int) -> tuple[int, int, int]:
    cid = get_selected_campaign_id(chat_id)

    def _inner(db):
        camp = get_campaign_or_default(db, cid)
        if not camp:
            return 0, 0, 0
        pairs = _goal_sender_target_pairs(db, int(camp.id))
        if not pairs:
            return 0, 0, 0
        pair_keys = {(int(aid), int(tid)) for aid, tid, _ in pairs}

        # Latest done check-status per pair.
        latest_status: dict[tuple[int, int], str] = {}
        check_rows = (
            db.query(Task.account_id, Task.target_id, Task.last_error)
            .filter(
                task_campaign_filter(db, int(camp.id)),
                Task.task_type == "check_status",
                Task.status == TaskStatus.DONE.value,
            )
            .order_by(Task.completed_at.desc().nullslast(), Task.id.desc())
            .all()
        )
        for account_id, target_id, last_error in check_rows:
            key = (int(account_id or 0), int(target_id or 0))
            if key not in pair_keys or key in latest_status:
                continue
            latest_status[key] = str(last_error or "").strip().lower()

        now = utc_now()
        queued = 0
        skipped_connected = 0
        skipped_active_or_invalid = 0

        for account_id, target_id, _ in pairs:
            key = (int(account_id), int(target_id))
            st = str(latest_status.get(key, "") or "")
            if st in {"friend_status:accepted", "friend_status:pending"}:
                skipped_connected += 1
                continue

            has_active_send = db.query(Task.id).filter(
                Task.task_type == "send_request",
                Task.campaign_id == int(camp.id),
                Task.account_id == int(account_id),
                Task.target_id == int(target_id),
                Task.status.in_(ACTIVE_SEND_TASK_STATUSES),
            ).first()
            if has_active_send:
                skipped_active_or_invalid += 1
                continue

            acc = db.query(Account).filter(Account.id == int(account_id)).first()
            if not acc or not acc.enabled or acc.status != AccountStatus.ACTIVE.value:
                skipped_active_or_invalid += 1
                continue
            if not (acc.epic_account_id and acc.device_id and acc.device_secret):
                skipped_active_or_invalid += 1
                continue

            min_s = max(0, int(camp.jitter_min_sec or DEFAULT_SEND_JITTER_MIN_SEC))
            max_s = max(min_s, int(camp.jitter_max_sec or DEFAULT_SEND_JITTER_MAX_SEC))
            db.add(
                Task(
                    task_type="send_request",
                    status=TaskStatus.QUEUED.value,
                    campaign_id=int(camp.id),
                    account_id=int(account_id),
                    target_id=int(target_id),
                    scheduled_for=now + timedelta(seconds=random.randint(min_s, max_s)),
                    max_attempts=5,
                    last_error="recheck_resend",
                )
            )
            queued += 1
        db.commit()
        return int(queued), int(skipped_connected), int(skipped_active_or_invalid)

    return db_exec(_inner)


def _run_goal_relationship_action(chat_id: int, action: str):
    cid = get_selected_campaign_id(chat_id)
    with REL_ACTION_JOBS_LOCK:
        existing = REL_ACTION_JOBS.get(int(chat_id))
        if existing and existing.get("thread") and existing["thread"].is_alive():
            show_menu_status(chat_id, "targets", "⚠️ Уже выполняется операция. Сначала останови её кнопкой «⛔ Остановить операцию».")
            return

    def _snapshot(db):
        camp = get_campaign_or_default(db, cid)
        camp_name = camp.name if camp else f"#{cid}"
        pairs = _goal_sender_target_pairs(db, cid)
        return camp_name, pairs

    camp_name, pairs = db_exec(_snapshot)
    if not pairs:
        show_menu_status(chat_id, "targets", "ℹ️ Нет отправленных заявок по цели для выполнения операции.")
        return

    show_menu_status(
        chat_id,
        "targets",
        f"⏳ Выполняю операцию по цели «{camp_name}».\nПары отправитель→ник: {len(pairs)}",
    )

    cancel_event = threading.Event()

    interval_sec = db_exec(lambda db: max(1, get_setting_int(db, "relationship_action_interval_sec", 40)))

    def worker():
        ok_count = 0
        skipped = 0
        skipped_no_auth = 0
        skipped_not_in_needed_status = 0
        failed = 0
        processed = 0
        total = len(pairs)
        pending_found = 0
        accepted_found = 0
        last_push = 0.0
        last_api_call_ts = 0.0
        stopped_by_user = False
        crashed = False
        op_name = "Отзыв заявок" if action == "revoke" else "Удаление из друзей"

        def _wait_api_slot() -> bool:
            nonlocal last_api_call_ts
            now_ts = time.monotonic()
            wait_sec = max(0.0, float(interval_sec) - (now_ts - last_api_call_ts))
            if wait_sec > 0:
                end_at = now_ts + wait_sec
                while time.monotonic() < end_at:
                    if cancel_event.is_set():
                        return False
                    time.sleep(0.2)
            if cancel_event.is_set():
                return False
            last_api_call_ts = time.monotonic()
            return True

        def _push_progress(force: bool = False):
            nonlocal last_push
            now_ts = time.monotonic()
            if not force and processed < total and processed % 15 != 0 and (now_ts - last_push) < 8:
                return
            last_push = now_ts
            left = max(0, total - processed)
            status_lines = [
                f"⏳ {op_name}: {processed}/{total}",
                f"Осталось проверить: {left}",
                f"Успешно: {ok_count}",
                f"Пропущено: {skipped} (без auth: {skipped_no_auth}, не подходит статус: {skipped_not_in_needed_status})",
                f"Ошибок: {failed}",
            ]
            if action == "revoke":
                status_lines.insert(2, f"Найдено pending: {pending_found}")
            else:
                status_lines.insert(2, f"Найдено в друзьях: {accepted_found}")
            show_menu_status(chat_id, "targets", "\n".join(status_lines))

        try:
            for account_id, target_id, username in pairs:
                if cancel_event.is_set():
                    stopped_by_user = True
                    break
                def _acc(db):
                    acc = db.query(Account).filter(Account.id == int(account_id)).first()
                    if not acc:
                        return None, None
                    return acc, get_proxy_for_account(db, acc.id)

                acc, proxy_url = db_exec(_acc)
                if not acc or not (acc.epic_account_id and acc.device_id and acc.device_secret):
                    skipped += 1
                    skipped_no_auth += 1
                    processed += 1
                    _push_progress()
                    continue

                if not _wait_api_slot():
                    stopped_by_user = True
                    break
                st = check_friend_status_with_device(
                    login=acc.login,
                    password=acc.password,
                    target_username=username,
                    proxy_url=proxy_url,
                    epic_account_id=acc.epic_account_id,
                    device_id=acc.device_id,
                    device_secret=acc.device_secret,
                )
                if not st.ok:
                    failed += 1
                    processed += 1
                    _push_progress()
                    continue

                if action == "revoke":
                    if st.code != "pending":
                        skipped += 1
                        skipped_not_in_needed_status += 1
                        processed += 1
                        _push_progress()
                        continue
                    pending_found += 1
                    if not _wait_api_slot():
                        stopped_by_user = True
                        break
                    res = cancel_friend_request_with_device(
                        login=acc.login,
                        password=acc.password,
                        target_username=username,
                        proxy_url=proxy_url,
                        epic_account_id=acc.epic_account_id,
                        device_id=acc.device_id,
                        device_secret=acc.device_secret,
                    )
                else:
                    if st.code != "accepted":
                        skipped += 1
                        skipped_not_in_needed_status += 1
                        processed += 1
                        _push_progress()
                        continue
                    accepted_found += 1
                    if not _wait_api_slot():
                        stopped_by_user = True
                        break
                    res = remove_friend_with_device(
                        login=acc.login,
                        password=acc.password,
                        target_username=username,
                        proxy_url=proxy_url,
                        epic_account_id=acc.epic_account_id,
                        device_id=acc.device_id,
                        device_secret=acc.device_secret,
                    )

                if res.ok:
                    ok_count += 1
                    def _save_action(db):
                        now_ts = utc_now()
                        db.add(
                            Task(
                                task_type="revoke_request" if action == "revoke" else "remove_friend",
                                status=TaskStatus.DONE.value,
                                campaign_id=int(cid) if int(cid or 0) > 0 else None,
                                account_id=int(account_id),
                                target_id=int(target_id),
                                scheduled_for=now_ts,
                                completed_at=now_ts,
                                attempt_number=1,
                            )
                        )
                        db.commit()

                    db_exec(_save_action)
                else:
                    failed += 1

                processed += 1
                _push_progress()
        except Exception as e:
            crashed = True
            failed += 1
            err = str(e)
            log_event("error", f"relationship_action_failed action={action} chat_id={chat_id} processed={processed}/{total} err={err}")
            show_menu_status(
                chat_id,
                "targets",
                f"❌ {op_name} прерван из-за ошибки.\n"
                f"Проверено: {processed}/{total}\n"
                f"Ошибка: {err[:300]}",
            )
        finally:
            if not crashed:
                _push_progress(force=True)
                if stopped_by_user:
                    show_menu_status(
                        chat_id,
                        "targets",
                        f"⛔ {op_name} остановлен пользователем.\n"
                        f"Проверено: {processed}/{total}\n"
                        f"Успешно: {ok_count}\n"
                        f"Пропущено: {skipped} (без auth: {skipped_no_auth}, не подходит статус: {skipped_not_in_needed_status})\n"
                        f"Ошибок: {failed}",
                    )
                else:
                    extra = (
                        f"Найдено pending: {pending_found}\n"
                        if action == "revoke"
                        else f"Найдено в друзьях: {accepted_found}\n"
                    )
                    show_menu_status(
                        chat_id,
                        "targets",
                        f"✅ {op_name} завершен.\n"
                        f"Проверено: {processed}/{total}\n"
                        f"{extra}"
                        f"Успешно: {ok_count}\n"
                        f"Пропущено: {skipped} (без auth: {skipped_no_auth}, не подходит статус: {skipped_not_in_needed_status})\n"
                        f"Ошибок: {failed}",
                    )
            with REL_ACTION_JOBS_LOCK:
                REL_ACTION_JOBS.pop(int(chat_id), None)

    th = threading.Thread(target=worker, daemon=True)
    with REL_ACTION_JOBS_LOCK:
        REL_ACTION_JOBS[int(chat_id)] = {"event": cancel_event, "action": action, "thread": th}
    th.start()


def stop_relationship_action(chat_id: int) -> bool:
    with REL_ACTION_JOBS_LOCK:
        job = REL_ACTION_JOBS.get(int(chat_id))
        ev = (job or {}).get("event")
        running = bool(job and job.get("thread") and job["thread"].is_alive())
        if ev is not None:
            ev.set()
    return running


def handle_revoke_requests_confirm(message):
    cleanup_step(message)
    if str(message.text or "").strip().upper() != "ОТОЗВАТЬ":
        show_menu_status(message.chat.id, "targets", "❌ Отменено.")
        return
    _run_goal_relationship_action(message.chat.id, "revoke")


def handle_remove_friends_confirm(message):
    cleanup_step(message)
    if str(message.text or "").strip().upper() != "УДАЛИТЬ":
        show_menu_status(message.chat.id, "targets", "❌ Отменено.")
        return
    _run_goal_relationship_action(message.chat.id, "remove")


@bot.callback_query_handler(func=lambda call: call.data == "tgt_status")
@admin_only_call
def cb_tgt_status(call):
    _safe_answer_callback(call)
    chat_id = call.message.chat.id
    show_targets_status(
        chat_id,
        page=get_chat_ui_int(chat_id, "tgt_page", 1),
        query=str(get_chat_ui_value(chat_id, "tgt_query", "") or ""),
    )


@bot.callback_query_handler(func=lambda call: call.data == "tgt_distribute")
@admin_only_call
def cb_tgt_distribute(call):
    _safe_answer_callback(call)

    chat_id = call.message.chat.id
    selected_campaign_id = get_selected_campaign_id(chat_id)
    enabled = db_exec(lambda db: get_setting_bool(db, "processing_enabled", True))
    if not enabled:
        show_menu_status(chat_id, "targets", "⏸️ Обработка остановлена. Сначала включи «▶️ Старт обработки».")
        return
    new_sends_enabled = db_exec(lambda db: is_new_send_requests_enabled(db))
    if not new_sends_enabled:
        show_menu_status(
            chat_id,
            "targets",
            "🧯 Новые заявки отключены в настройках. Распределение новых send-задач недоступно.",
        )
        return
    show_menu_status(chat_id, "targets", f"⏳ Распределяю ники получателя в цели #{selected_campaign_id}...")

    def _distribute_zero_reason(db, campaign_id: int) -> str:
        tgt_total = int(
            db.query(func.count(Target.id))
            .filter(target_campaign_filter(db, campaign_id))
            .scalar()
            or 0
        )
        status_counts = dict(
            db.query(Target.status, func.count(Target.id))
            .filter(target_campaign_filter(db, campaign_id))
            .group_by(Target.status)
            .all()
        )
        send_q = int(
            db.query(func.count(Task.id))
            .filter(
                task_campaign_filter(db, campaign_id),
                Task.task_type == "send_request",
                Task.status.in_(ACTIVE_SEND_TASK_STATUSES),
            )
            .scalar()
            or 0
        )
        check_q = int(
            db.query(func.count(Task.id))
            .filter(
                task_campaign_filter(db, campaign_id),
                Task.task_type == "check_status",
                Task.status.in_(ACTIVE_SEND_TASK_STATUSES),
            )
            .scalar()
            or 0
        )
        pending = int(status_counts.get(TargetStatus.PENDING.value, 0) or 0)
        accepted = int(status_counts.get(TargetStatus.ACCEPTED.value, 0) or 0)
        sent = int(status_counts.get(TargetStatus.SENT.value, 0) or 0)
        new = int(status_counts.get(TargetStatus.NEW.value, 0) or 0)
        return (
            "Новые задачи не добавлены:\n"
            f"• Ников в цели: {tgt_total}\n"
            f"• Статусы: new={new}, pending={pending}, sent={sent}, accepted={accepted}\n"
            f"• В очереди отправки: {send_q}\n"
            f"• В очереди проверки: {check_q}\n"
            "Если нужно ускорить именно обработку, используй «▶️ Старт обработки»."
        )

    def worker():
        try:
            created = db_exec(lambda db: create_tasks_for_new_targets(db, limit=1000, campaign_id=selected_campaign_id))
            if created > 0:
                show_menu_status(chat_id, "targets", f"✅ Распределение завершено.\nСоздано задач: {created}")
            else:
                reason_text = db_exec(lambda db: _distribute_zero_reason(db, selected_campaign_id))
                show_menu_status(chat_id, "targets", "✅ Распределение завершено.\nСоздано задач: 0\n" + reason_text)
        except Exception as e:
            show_menu_status(chat_id, "targets", f"❌ Ошибка распределения: {e}")

    threading.Thread(target=worker, daemon=True).start()


@bot.callback_query_handler(func=lambda call: call.data == "set_limit")
@admin_only_call
def cb_set_limit(call):
    _safe_answer_callback(call)

    def _inner(db):
        vals = [int(v[0] or 0) for v in db.query(Account.daily_limit).all()]
        if not vals:
            return 0, 0, 0
        return len(vals), min(vals), max(vals)

    total, min_limit, max_limit = db_exec(_inner)
    current = f"{min_limit}" if min_limit == max_limit else f"{min_limit}..{max_limit}"
    dummy = py_types.SimpleNamespace(chat=call.message.chat, from_user=call.from_user)
    ask_step(
        dummy,
        f"Текущий суточный лимит аккаунтов: {current}\n"
        f"Аккаунтов в базе: {total}\n\n"
        "Введи новый лимит (число >= 0):",
        handle_set_limit,
    )


def handle_set_limit(message):
    cleanup_step(message)

    try:
        limit = int((message.text or "").strip())
        if limit < 0:
            raise ValueError()
    except Exception:
        show_menu_status(message.chat.id, "settings", "❌ Неверное число.")
        return

    def _inner(db):
        prev_vals = [int(v[0] or 0) for v in db.query(Account.daily_limit).all()]
        for a in db.query(Account).all():
            a.daily_limit = limit
        db.commit()
        count = db.query(Account).count()
        if not prev_vals:
            return count, 0, 0
        return count, min(prev_vals), max(prev_vals)

    count, prev_min, prev_max = db_exec(_inner)
    prev_txt = f"{prev_min}" if prev_min == prev_max else f"{prev_min}..{prev_max}"
    show_menu_status(
        message.chat.id,
        "settings",
        f"✅ Суточный лимит аккаунтов обновлён:\nБыло: {prev_txt}\nСтало: {limit}\nПрименено к аккаунтам: {count}",
    )


@bot.callback_query_handler(func=lambda call: call.data == "set_jitter")
@admin_only_call
def cb_set_jitter(call):
    _safe_answer_callback(call)

    cur = db_exec(
        lambda db: (
            get_setting_int(db, "jitter_min_sec", DEFAULT_SEND_JITTER_MIN_SEC),
            get_setting_int(db, "jitter_max_sec", DEFAULT_SEND_JITTER_MAX_SEC),
        )
    )
    dummy = py_types.SimpleNamespace(chat=call.message.chat, from_user=call.from_user)
    ask_step(
        dummy,
        f"Текущий глобальный джиттер: {cur[0]}-{cur[1]} сек\n\n"
        "Введи: мин макс (в сек)\nПример: 300 900",
        handle_set_jitter,
    )


def handle_set_jitter(message):
    cleanup_step(message)

    try:
        min_s, max_s = map(int, (message.text or "").strip().split())
        if min_s < 0 or max_s < min_s:
            raise ValueError()
    except Exception:
        show_menu_status(message.chat.id, "settings", "❌ Неверный формат.")
        return

    prev = db_exec(
        lambda db: (
            get_setting_int(db, "jitter_min_sec", DEFAULT_SEND_JITTER_MIN_SEC),
            get_setting_int(db, "jitter_max_sec", DEFAULT_SEND_JITTER_MAX_SEC),
        )
    )
    db_exec(lambda db: (set_setting(db, "jitter_min_sec", str(min_s)), set_setting(db, "jitter_max_sec", str(max_s))))
    show_menu_status(
        message.chat.id,
        "settings",
        f"✅ Глобальный джиттер обновлён:\nБыло: {prev[0]}-{prev[1]} сек\nСтало: {min_s}-{max_s} сек",
    )


@bot.callback_query_handler(func=lambda call: call.data == "set_windows")
@admin_only_call
def cb_set_windows(call):
    _safe_answer_callback(call)

    dummy = py_types.SimpleNamespace(chat=call.message.chat, from_user=call.from_user)
    ask_step(
        dummy,
        "Формат:\nID_аккаунта\ndays=1,2,3 from=09:00 to=23:00\nПример:\n5\ndays=1,2,3,4,5 from=10:00 to=22:00\n"
        "Важно: окно интерпретируется по МСК (Europe/Moscow).",
        handle_set_windows,
    )


def handle_set_windows(message):
    cleanup_step(message)

    lines = safe_split_lines(message.text)
    if len(lines) < 2:
        show_menu_status(message.chat.id, "settings", "❌ Нужно минимум 2 строки.")
        return

    try:
        acc_id = int(lines[0])
        windows = parse_windows_text("\n".join(lines[1:]))
    except Exception as e:
        show_menu_status(message.chat.id, "settings", f"❌ Ошибка: {e}")
        return

    def _inner(db):
        acc = db.query(Account).filter(Account.id == acc_id).first()
        if not acc:
            return False
        acc.active_windows_json = json.dumps(windows)
        db.commit()
        return True

    ok = db_exec(_inner)
    show_menu_status(message.chat.id, "settings", "✅ Окна сохранены." if ok else "❌ Аккаунт не найден.")


@bot.callback_query_handler(func=lambda call: call.data == "set_target_senders")
@admin_only_call
def cb_set_target_senders(call):
    _safe_answer_callback(call)

    chat_id = call.message.chat.id
    selected_campaign_id = get_selected_campaign_id(chat_id)

    def _inner(db):
        camp = get_campaign_or_default(db, selected_campaign_id)
        current = int(camp.target_senders_count or 1) if camp else 1
        camp_name = camp.name if camp else f"#{selected_campaign_id}"
        return camp_name, current

    camp_name, current = db_exec(_inner)
    dummy = py_types.SimpleNamespace(chat=call.message.chat, from_user=call.from_user)
    ask_step(
        dummy,
        f"Цель: {camp_name} (#{selected_campaign_id})\n"
        f"Сейчас: {current} аккаунтов на 1 ник получателя\n\n"
        "Введи новое количество (целое >= 1):",
        handle_set_target_senders,
    )


def handle_set_target_senders(message):
    cleanup_step(message)
    try:
        count = int((message.text or "").strip())
        if count < 1:
            raise ValueError()
    except Exception:
        show_menu_status(message.chat.id, "targets", "❌ Неверное число.")
        return

    selected_campaign_id = get_selected_campaign_id(message.chat.id)

    def _inner(db):
        camp = get_campaign_or_default(db, selected_campaign_id)
        if camp:
            camp.target_senders_count = int(count)
            camp.updated_at = utc_now()
        set_setting(db, "target_senders_count", str(count))
        updated_targets = (
            db.query(Target)
            .filter(target_campaign_filter(db, selected_campaign_id))
            .update({Target.required_senders: int(count)}, synchronize_session=False)
        )
        db.commit()
        dropped_queue, created_tasks = rebuild_campaign_send_queue(db, int(selected_campaign_id), create_limit=5000)
        return int(updated_targets or 0), int(dropped_queue or 0), int(created_tasks or 0)

    updated_targets, dropped_queue, created_tasks = db_exec(_inner)
    show_menu_status(
        message.chat.id,
        "targets",
        f"✅ На 1 ник получателя: {count} аккаунтов\n"
        f"Цель: #{selected_campaign_id}\n"
        f"Обновлено ников получателя: {updated_targets}\n"
        f"Очищено старых задач: {dropped_queue}\n"
        f"Добавлено недостающих задач: {created_tasks}",
    )


@bot.callback_query_handler(func=lambda call: call.data == "set_timezone")
@admin_only_call
def cb_set_timezone(call):
    _safe_answer_callback(call)

    dummy = py_types.SimpleNamespace(chat=call.message.chat, from_user=call.from_user)
    ask_step(
        dummy,
        "Введи timezone (пример: Europe/Moscow):",
        handle_set_timezone,
    )


def handle_set_timezone(message):
    cleanup_step(message)
    tz_name = (message.text or "").strip()
    try:
        ZoneInfo(tz_name)
    except Exception:
        show_menu_status(message.chat.id, "settings", "❌ Неверный часовой пояс (пример: Europe/Moscow)")
        return

    db_exec(lambda db: set_setting(db, "runtime_timezone", tz_name))
    show_menu_status(message.chat.id, "settings", f"✅ Часовой пояс: {tz_name}")


def handle_set_api_limits(message):
    cleanup_step(message)
    try:
        parts = [int(x) for x in (message.text or "").strip().split()]
        if len(parts) != 3:
            raise ValueError()
        min_interval_sec, hourly_limit, daily_limit = parts
        if min_interval_sec < 0 or hourly_limit < 1 or daily_limit < 1:
            raise ValueError()
    except Exception:
        show_menu_status(
            message.chat.id,
            "settings",
            "❌ Неверный формат. Пример: `40 40 500`",
        )
        return

    def _inner(db):
        set_setting(db, "min_request_interval_sec", str(int(min_interval_sec)))
        set_setting(db, "max_request_interval_sec", str(int(min_interval_sec)))
        set_setting(db, "hourly_api_limit", str(int(hourly_limit)))
        set_setting(db, "daily_api_limit", str(int(daily_limit)))

    db_exec(_inner)
    show_menu_status(
        message.chat.id,
        "settings",
        "✅ API-лимиты обновлены:\n"
        f"• Интервал: {min_interval_sec} сек\n"
        f"• В час: {hourly_limit}\n"
        f"• В сутки: {daily_limit}",
    )


def _parse_positive_user_id(raw: str) -> int | None:
    try:
        uid = int(str(raw or "").strip())
        if uid <= 0:
            return None
        return uid
    except Exception:
        return None


def handle_auth_operator_add_id(message):
    cleanup_step(message)
    uid = _parse_positive_user_id(message.text)
    if not uid:
        show_menu_status(message.chat.id, "auth_access", "❌ Неверный ID. Введи положительное число.")
        return
    if is_admin(uid):
        show_menu_status(message.chat.id, "auth_access", "ℹ️ Этот ID уже админ, отдельный доступ не нужен.")
        return

    def _inner(db):
        ids = _load_auth_operator_ids(db)
        ids.add(int(uid))
        _save_auth_operator_ids(db, ids)
        return sorted(ids)

    ids = db_exec(_inner)
    show_menu_status(
        message.chat.id,
        "auth_access",
        f"✅ ID {uid} добавлен.\nТекущий список: " + (", ".join(str(x) for x in ids) if ids else "пусто"),
    )


def handle_auth_operator_remove_id(message):
    cleanup_step(message)
    uid = _parse_positive_user_id(message.text)
    if not uid:
        show_menu_status(message.chat.id, "auth_access", "❌ Неверный ID. Введи положительное число.")
        return

    def _inner(db):
        ids = _load_auth_operator_ids(db)
        ids.discard(int(uid))
        _save_auth_operator_ids(db, ids)
        return sorted(ids)

    ids = db_exec(_inner)
    show_menu_status(
        message.chat.id,
        "auth_access",
        f"✅ ID {uid} удалён (если был в списке).\nТекущий список: " + (", ".join(str(x) for x in ids) if ids else "пусто"),
    )


def handle_auth_operator_clear_all_confirm(message):
    cleanup_step(message)
    raw = str(message.text or "").strip().upper()
    if raw != "ОЧИСТИТЬ":
        show_menu_status(
            message.chat.id,
            "auth_access",
            "❌ Очистка отменена. Для подтверждения нужно ввести: ОЧИСТИТЬ",
        )
        return

    def _inner(db):
        _save_auth_operator_ids(db, set())

    db_exec(_inner)
    show_menu_status(message.chat.id, "auth_access", "✅ Список ID auth очищен.")


@bot.callback_query_handler(func=lambda call: call.data == "set_recheck_limit")
@admin_only_call
def cb_set_recheck_limit(call):
    _safe_answer_callback(call)

    chat_id = call.message.chat.id
    selected_campaign_id = get_selected_campaign_id(chat_id)

    def _inner(db):
        camp = get_campaign_or_default(db, selected_campaign_id)
        current = int(camp.recheck_daily_limit or DEFAULT_RECHECK_DAILY_LIMIT) if camp else DEFAULT_RECHECK_DAILY_LIMIT
        camp_name = camp.name if camp else f"#{selected_campaign_id}"
        return camp_name, current

    camp_name, current = db_exec(_inner)
    dummy = py_types.SimpleNamespace(chat=call.message.chat, from_user=call.from_user)
    ask_step(
        dummy,
        f"Цель: {camp_name} (#{selected_campaign_id})\n"
        f"Сейчас лимит повторных проверок: {current}/сутки\n\n"
        "Введи лимит повторных проверок в сутки (целое >= 0):\n"
        "Это сколько старых заявок максимум перепроверять за день.",
        handle_set_recheck_limit,
    )


def handle_set_recheck_limit(message):
    cleanup_step(message)
    try:
        limit = int((message.text or "").strip())
        if limit < 0:
            raise ValueError()
    except Exception:
        show_menu_status(message.chat.id, "targets", "❌ Неверное число.")
        return

    selected_campaign_id = get_selected_campaign_id(message.chat.id)

    def _inner(db):
        camp = get_campaign_or_default(db, selected_campaign_id)
        if camp:
            camp.recheck_daily_limit = int(limit)
            camp.updated_at = utc_now()
            db.commit()
        set_setting(db, "recheck_daily_limit", str(limit))

    db_exec(_inner)
    show_menu_status(message.chat.id, "targets", f"✅ Лимит повторных проверок: {limit}/сутки (цель #{selected_campaign_id})")


@bot.callback_query_handler(func=lambda call: call.data == "set_daily_repeat")
@admin_only_call
def cb_set_daily_repeat(call):
    _safe_answer_callback(call)

    chat_id = call.message.chat.id
    selected_campaign_id = get_selected_campaign_id(chat_id)
    camp_info = db_exec(
        lambda db: db.query(Campaign.name, Campaign.daily_repeat_enabled).filter(Campaign.id == selected_campaign_id).first()
    )
    camp_name = (camp_info[0] if camp_info else f"#{selected_campaign_id}")
    enabled_now = bool(camp_info[1]) if camp_info else False
    dummy = py_types.SimpleNamespace(chat=call.message.chat, from_user=call.from_user)
    ask_step(
        dummy,
        f"Цель: {camp_name} (#{selected_campaign_id})\n"
        f"Сейчас: {'включен' if enabled_now else 'выключен'}\n\n"
        "Включить ежедневный повтор цели? (1=да, 0=нет):\n"
        "Если включено, на следующий день цели снова получают отправителей по правилам лимитов.",
        handle_set_daily_repeat,
    )


def handle_set_daily_repeat(message):
    cleanup_step(message)
    raw = (message.text or "").strip().lower()
    if raw not in {"0", "1", "no", "yes", "false", "true"}:
        show_menu_status(message.chat.id, "targets", "❌ Введи 1 или 0.")
        return
    enabled = raw in {"1", "yes", "true"}
    selected_campaign_id = get_selected_campaign_id(message.chat.id)

    def _inner(db):
        camp = get_campaign_or_default(db, selected_campaign_id)
        dropped_queue = 0
        created_tasks = 0
        if camp:
            camp.daily_repeat_enabled = bool(enabled)
            camp.updated_at = utc_now()
            db.commit()
            dropped_queue, created_tasks = rebuild_campaign_send_queue(db, int(camp.id), create_limit=5000)
        set_setting(db, "daily_repeat_campaign_enabled", "1" if enabled else "0")
        return int(dropped_queue), int(created_tasks)

    dropped_queue, created_tasks = db_exec(_inner)
    show_menu_status(
        message.chat.id,
        "targets",
        f"✅ Ежедневный повтор: {'включен' if enabled else 'выключен'} (цель #{selected_campaign_id})\n"
        f"Очищено старых задач: {dropped_queue}\n"
        f"Добавлено недостающих задач: {created_tasks}",
    )


@bot.callback_query_handler(func=lambda call: call.data == "proxy_add")
@admin_only_call
def cb_proxy_add(call):
    _safe_answer_callback(call)

    dummy = py_types.SimpleNamespace(chat=call.message.chat, from_user=call.from_user)
    ask_step(
        dummy,
        "Введи URL прокси (пример: http://ip:port):",
        handle_proxy_add,
    )


def handle_proxy_add(message):
    cleanup_step(message)

    proxy_url = (message.text or "").strip()
    if not proxy_url.startswith("http"):
        show_menu_status(message.chat.id, "proxy", "❌ URL должен начинаться с http://.")
        return

    def _inner(db):
        if db.query(Proxy).filter(Proxy.url == proxy_url).first():
            return False
        db.add(Proxy(url=proxy_url, enabled=True))
        db.commit()
        return True

    ok = db_exec(_inner)
    show_menu_status(message.chat.id, "proxy", "✅ Прокси добавлена" if ok else "❌ Такая прокси уже есть")


@bot.callback_query_handler(func=lambda call: call.data == "proxy_list")
@admin_only_call
def cb_proxy_list(call):
    _safe_answer_callback(call)

    proxies = db_exec(lambda db: db.query(Proxy).order_by(Proxy.id.asc()).all())
    if not proxies:
        show_screen(call.message.chat.id, "❌ Прокси не найдены.", reply_markup=kb_proxy_reply(), parse_mode=None)
        return

    lines = ["📌 **Прокси:**\n"]
    for p in proxies[:80]:
        status = "✅" if p.enabled else "❌"
        lines.append(f"{status} #{p.id} {md_inline_code(p.url)}")

    text = "\n".join(lines)
    if len(text) > 3800:
        text = text[:3790] + "…"

    show_screen(call.message.chat.id, text, reply_markup=kb_proxy_reply(), parse_mode="Markdown")


@bot.callback_query_handler(func=lambda call: call.data == "proxy_delete")
@admin_only_call
def cb_proxy_delete(call):
    _safe_answer_callback(call)

    dummy = py_types.SimpleNamespace(chat=call.message.chat, from_user=call.from_user)
    ask_step(
        dummy,
        "Введи ID прокси из списка:",
        handle_proxy_delete,
    )


def handle_proxy_delete(message):
    cleanup_step(message)

    try:
        proxy_id = int((message.text or "").strip())
    except Exception:
        show_menu_status(message.chat.id, "proxy", "❌ Неверный ID.")
        return

    def _inner(db):
        proxy = db.query(Proxy).filter(Proxy.id == proxy_id).first()
        if not proxy:
            return False
        db.delete(proxy)
        db.commit()
        return True

    ok = db_exec(_inner)
    show_menu_status(message.chat.id, "proxy", "✅ Прокси удалена" if ok else "❌ Прокси не найдена")


@bot.callback_query_handler(func=lambda call: call.data == "manage_tick")
@admin_only_call
def cb_manage_tick(call):
    _safe_answer_callback(call)

    chat_id = call.message.chat.id

    def worker():
        try:
            process_tasks_job()
            show_menu_status(chat_id, "manage", "✅ Цикл выполнен.")
        except Exception as e:
            show_menu_status(chat_id, "manage", f"❌ Ошибка цикла: {e}")

    threading.Thread(target=worker, daemon=True).start()


@bot.callback_query_handler(func=lambda call: call.data == "manage_stop")
@admin_only_call
def cb_manage_stop(call):
    _safe_answer_callback(call)
    set_processing_enabled(False)
    show_menu_status(call.message.chat.id, "manage", "⏸️ Обработка остановлена.")


@bot.callback_query_handler(func=lambda call: call.data == "manage_start")
@admin_only_call
def cb_manage_start(call):
    _safe_answer_callback(call)
    set_processing_enabled(True)
    show_menu_status(call.message.chat.id, "manage", "▶️ Обработка запущена.")


@bot.callback_query_handler(func=lambda call: call.data == "manage_status")
@admin_only_call
def cb_manage_status(call):
    _safe_answer_callback(call)
    show_processing_status(call.message.chat.id)


@bot.callback_query_handler(func=lambda call: call.data == "manage_export")
@admin_only_call
def cb_manage_export(call):
    _safe_answer_callback(call)

    chat_id = call.message.chat.id
    try:
        filename = export_results_to_excel()
        with open(filename, "rb") as f:
            bot.send_document(chat_id, f, visible_file_name="results.xlsx")
        show_menu_status(chat_id, "manage", "✅ Экспорт готов.")
    except Exception as e:
        show_menu_status(chat_id, "manage", f"❌ Ошибка экспорта: {e}")



# ============================================================
# DOC UPLOAD (Excel import)
# ============================================================

@bot.message_handler(content_types=["document"])
@admin_only
def handle_document(message):
    # чистим сообщение пользователя (если возможно)
    _safe_delete_user_message(message)

    file_info = bot.get_file(message.document.file_id)
    file_path = file_info.file_path
    file_name = os.path.basename(message.document.file_name or "file.xlsx")

    lower_name = file_name.lower()
    allowed_ext = (".xlsx", ".txt", ".csv")
    if not lower_name.endswith(allowed_ext):
        show_screen(message.chat.id, "❌ Поддерживается только .xlsx/.txt/.csv", parse_mode=None)
        return

    try:
        downloaded = bot.download_file(file_path)
        local_path = f"/tmp/{file_name}"
        with open(local_path, "wb") as f:
            f.write(downloaded)

        current_menu = get_current_menu(message.chat.id)
        if current_menu == "nick_change_import":
            added, skipped, errors = import_nickname_change_tasks(local_path, source_file=file_name)
            show_menu_status(
                message.chat.id,
                "accounts",
                f"✅ Импорт задач смены ников завершён: ➕ {added}, ⏭️ {skipped}, ❌ {errors}",
            )
            return

        if lower_name.endswith(".xlsx"):
            added, skipped, errors = import_accounts_from_excel(local_path)
            entity = "аккаунтов"
        else:
            with open(local_path, "r", encoding="utf-8-sig") as fp:
                sample = [ln.strip() for ln in fp.readlines() if ln.strip() and not ln.strip().startswith("#")]
            first_line = sample[0] if sample else ""

            if current_menu == "targets":
                try:
                    added, skipped, errors = import_targets_from_text(
                        local_path, campaign_id=get_selected_campaign_id(message.chat.id)
                    )
                except TypeError:
                    added, skipped, errors = import_targets_from_text(local_path)
                entity = "ников получателя"
            elif current_menu == "accounts":
                added, skipped, errors = import_accounts_from_text(local_path)
                entity = "аккаунтов"
            elif _parse_account_line(first_line):
                added, skipped, errors = import_accounts_from_text(local_path)
                entity = "аккаунтов"
            else:
                try:
                    added, skipped, errors = import_targets_from_text(
                        local_path, campaign_id=get_selected_campaign_id(message.chat.id)
                    )
                except TypeError:
                    added, skipped, errors = import_targets_from_text(local_path)
                entity = "ников получателя"

        if entity == "аккаунтов":
            show_menu_status(
                message.chat.id,
                "accounts",
                f"✅ Импорт аккаунтов завершён: ➕ {added}, ⏭️ {skipped}, ❌ {errors}",
            )
        else:
            show_menu_status(
                message.chat.id,
                "targets",
                f"✅ Импорт ников получателя завершён: ➕ {added}, ⏭️ {skipped}, ❌ {errors}",
            )
    except Exception as e:
        show_screen(message.chat.id, f"❌ Ошибка импорта: {str(e)[:200]}", parse_mode=None)

def create_tasks_for_new_targets_job():
    """Scheduler wrapper для распределения ников получателя по активным целям."""
    def _inner(db):
        if not get_setting_bool(db, "processing_enabled", True):
            return 0
        if not is_new_send_requests_enabled(db):
            return 0
        campaign_ids = [int(x[0]) for x in db.query(Campaign.id).filter(Campaign.enabled == True).all()]
        total = 0
        for cid in campaign_ids:
            total += int(create_tasks_for_new_targets(db, limit=1000, campaign_id=cid) or 0)
        return total
    created = db_exec(_inner)
    if created > 0:
        log_event("info", f"⚙️ Авто-распределение: создано задач {created}")


def create_recheck_tasks_job():
    """
    Плановое создание check_status задач для уже отправленных заявок.
    Идёт сверх ежедневной отправки send_request, но с отдельным суточным лимитом.
    """

    def _inner(db):
        if not get_setting_bool(db, "processing_enabled", True):
            return 0
        today = utc_today().isoformat()
        total_created = 0
        total_counter_value = 0

        # Per-goal counters/limits: each goal has its own recheck budget per day.
        campaigns = db.query(Campaign).filter(Campaign.enabled == True).all()
        for camp in campaigns:
            camp_id = int(camp.id)
            day_key = f"recheck_counter_day_{camp_id}"
            val_key = f"recheck_counter_value_{camp_id}"
            counter_day = get_setting(db, day_key, "")
            counter_val = get_setting_int(db, val_key, 0)
            if counter_day != today:
                counter_day = today
                counter_val = 0

            daily_limit = max(0, int(camp.recheck_daily_limit or DEFAULT_RECHECK_DAILY_LIMIT))
            remaining = max(0, daily_limit - counter_val)
            if remaining <= 0:
                set_setting(db, day_key, counter_day)
                set_setting(db, val_key, str(counter_val))
                continue

            rows = (
                db.query(Task.account_id, Task.target_id, Target.username)
                .join(Target, Target.id == Task.target_id)
                .filter(
                    Task.task_type == "send_request",
                    Task.status == TaskStatus.DONE.value,
                    Target.status.in_(TARGET_RECHECK_ELIGIBLE_STATUSES),
                    or_(Task.campaign_id == camp_id, and_(Task.campaign_id.is_(None), Target.campaign_id == camp_id)),
                )
                .order_by(Task.completed_at.desc().nullslast(), Task.id.desc())
                .limit(max(remaining * 5, 100))
                .all()
            )
            send_mode = get_campaign_send_mode(db, camp_id)
            if send_mode == "target_first":
                planner = RecheckQueuePlanner(
                    mode="nickname",
                    shuffle_groups=True,          # random nickname order
                    shuffle_inside_group=False,   # keep sender order inside nickname group
                    seed=random.randint(1, 10**9),
                )
            else:
                planner = RecheckQueuePlanner(
                    mode="sender",
                    shuffle_groups=False,         # keep sender sequence
                    shuffle_inside_group=True,    # random nicknames for each sender
                    seed=random.randint(1, 10**9),
                )
            planner.build(
                RecheckPair(
                    account_id=int(account_id),
                    target_id=int(target_id),
                    nickname=str(username or ""),
                )
                for account_id, target_id, username in rows
            )
            ordered_pairs = planner.pop_many(len(rows))

            created = 0
            now = utc_now()
            for pair in ordered_pairs:
                if created >= remaining:
                    break
                account_id = int(pair.account_id)
                target_id = int(pair.target_id)

                has_active_check = db.query(Task.id).filter(
                    Task.task_type == "check_status",
                    Task.campaign_id == camp_id,
                    Task.account_id == account_id,
                    Task.target_id == target_id,
                    Task.status.in_(ACTIVE_SEND_TASK_STATUSES),
                ).first()
                if has_active_check:
                    continue

                acc = db.query(Account).filter(Account.id == account_id).first()
                if not acc or not acc.enabled or acc.status != AccountStatus.ACTIVE.value:
                    continue
                if not (acc.epic_account_id and acc.device_id and acc.device_secret):
                    continue

                db.add(
                    Task(
                        task_type="check_status",
                        status=TaskStatus.QUEUED.value,
                        campaign_id=camp_id,
                        account_id=account_id,
                        target_id=target_id,
                        scheduled_for=now + timedelta(seconds=random.randint(
                            max(0, int(camp.jitter_min_sec or DEFAULT_SEND_JITTER_MIN_SEC)),
                            max(
                                max(0, int(camp.jitter_min_sec or DEFAULT_SEND_JITTER_MIN_SEC)),
                                int(camp.jitter_max_sec or DEFAULT_SEND_JITTER_MAX_SEC),
                            ),
                        )),
                        max_attempts=5,
                    )
                )
                created += 1

            new_counter_val = counter_val + created
            set_setting(db, day_key, counter_day)
            set_setting(db, val_key, str(new_counter_val))
            total_created += created
            total_counter_value += new_counter_val

        # Legacy compatibility: keep aggregate counters updated for diagnostics/tools.
        set_setting(db, "recheck_counter_day", today)
        set_setting(db, "recheck_counter_value", str(total_counter_value))
        return total_created

    created = db_exec(_inner)
    if created > 0:
        log_event("info", f"🔁 Авто-recheck: создано задач {created}")
    return int(created or 0)


def start_scheduler(
    enable_reset: bool = True,
    enable_process: bool = True,
    enable_verify: bool = True,
    enable_distribute: bool = True,
    enable_recheck: bool = True,
    enable_prune: bool = True,
) -> BackgroundScheduler:
    """Инициализировать и запустить scheduler."""
    job_defaults = {
        "coalesce": True,
        "max_instances": 1,
        "misfire_grace_time": 120,
    }
    sched = BackgroundScheduler(job_defaults=job_defaults)

    if enable_reset:
        sched.add_job(reset_daily_counters_job, 'cron', hour=DAILY_RESET_HOUR_UTC, minute=0)
    if enable_process:
        sched.add_job(process_tasks_job, 'interval', seconds=PROCESS_TICK_SECONDS, max_instances=2)
        sched.add_job(process_nickname_change_tasks_job, 'interval', seconds=max(10, PROCESS_TICK_SECONDS), max_instances=2)
    if enable_verify:
        sched.add_job(verify_accounts_health_job, 'interval', seconds=300)
    if enable_distribute:
        sched.add_job(create_tasks_for_new_targets_job, 'interval', seconds=PROCESS_TICK_SECONDS * 2, max_instances=2)
    if enable_recheck:
        sched.add_job(create_recheck_tasks_job, 'interval', seconds=max(120, PROCESS_TICK_SECONDS * 3), max_instances=2)
    if enable_prune:
        sched.add_job(prune_log_events_job, 'interval', hours=12)

    sched.start()
    log_event(
        "info",
        "🚀 Scheduler активирован "
        f"(reset={enable_reset}, process={enable_process}, verify={enable_verify}, "
        f"distribute={enable_distribute}, recheck={enable_recheck}, prune={enable_prune})",
    )
    return sched


scheduler = None

# ============================================================
# MAIN
# ============================================================

def run_bot_ui():
    # Prevent accidental multiple polling instances (causes Telegram 409 and duplicate messages).
    _acquire_polling_lock()
    print("🤖 Режим: bot-ui")
    print(f"Admin ID: {ADMIN_ID}")
    print(f"DB: {DB_URL}")
    _polling_with_retry()


def run_worker_loop():
    # worker: только обработка очереди задач, без reset/verify/distribute/prune
    local_scheduler = start_scheduler(
        enable_reset=False,
        enable_process=True,
        enable_verify=False,
        enable_distribute=False,
        enable_recheck=False,
        enable_prune=False,
    )
    print("⚙️ Режим: worker")
    print(f"DB: {DB_URL}")
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        print("\n⏹️ Worker остановлен пользователем")
    finally:
        if local_scheduler is not None:
            print("🛑 Останавливаем scheduler...")
            local_scheduler.shutdown(wait=False)
        print("✅ Worker остановлен")


def run_all_in_one():
    global scheduler
    _acquire_polling_lock()
    scheduler = start_scheduler()
    print("🚀 Epic Games Account Manager запущен")
    print(f"Admin ID: {ADMIN_ID}")
    print(f"DB: {DB_URL}")
    print(f"Scheduler: {'✅ Активирован' if scheduler else '❌ Отключен'}")
    try:
        _polling_with_retry()
    except KeyboardInterrupt:
        print("\n⏹️ Остановлен пользователем")
    finally:
        if scheduler is not None:
            print("🛑 Останавливаем scheduler...")
            scheduler.shutdown(wait=False)
        print("✅ Бот остановлен")


def _polling_with_retry():
    """
    Poll Telegram with retry to avoid crash-loops on transient 409 conflicts.
    409 happens when another getUpdates request is active (e.g. another instance
    just shut down or is running elsewhere). In Docker a crash-loop makes this
    worse, so we back off and retry.
    """
    retry_sec = int(os.getenv("BOT_POLLING_RETRY_SEC", "5"))
    # Avoid processing stale pending updates on restarts (can lead to duplicate UI messages).
    while True:
        try:
            bot.infinity_polling(skip_pending=True)
            return
        except ApiTelegramException as e:
            if getattr(e, "error_code", None) == 409:
                logger.warning(f"Telegram polling conflict (409). Retry in {retry_sec}s.")
                time.sleep(max(1, retry_sec))
                continue
            raise
        except Exception as e:
            # Network hiccups shouldn't kill the whole app in production.
            logger.warning(f"Telegram polling error: {e}. Retry in {retry_sec}s.")
            time.sleep(max(1, retry_sec))
            continue


def run_app():
    ensure_runtime_settings()
    mode = os.getenv("APP_MODE", "all").strip().lower()
    if mode == "bot":
        run_bot_ui()
        return
    if mode == "worker":
        run_worker_loop()
        return
    if mode == "scheduler":
        # scheduler: reset/verify/distribute/prune, без process_tasks
        local_scheduler = start_scheduler(
            enable_reset=True,
            enable_process=False,
            enable_verify=True,
            enable_distribute=True,
            enable_recheck=True,
            enable_prune=True,
        )
        print("🗓️ Режим: scheduler")
        print(f"DB: {DB_URL}")
        try:
            while True:
                time.sleep(1)
        except KeyboardInterrupt:
            print("\n⏹️ Scheduler остановлен пользователем")
        finally:
            print("🛑 Останавливаем scheduler...")
            local_scheduler.shutdown(wait=False)
            print("✅ Scheduler остановлен")
        return
    if mode == "all":
        run_all_in_one()
        return
    raise SystemExit("❌ APP_MODE должен быть одним из: all, bot, worker, scheduler")


_POLLING_LOCK_FH = None


def _acquire_polling_lock():
    """
    Use a best-effort process lock to prevent running multiple bot polling instances locally.
    On macOS/Linux uses fcntl.
    """
    global _POLLING_LOCK_FH
    lock_path = os.getenv("BOT_POLLING_LOCK_PATH", "/tmp/egs_bot_polling.lock").strip() or "/tmp/egs_bot_polling.lock"
    try:
        import fcntl
    except Exception:
        return

    try:
        fh = open(lock_path, "w")
        fcntl.flock(fh, fcntl.LOCK_EX | fcntl.LOCK_NB)
        fh.write(str(os.getpid()))
        fh.flush()
        _POLLING_LOCK_FH = fh
    except Exception:
        raise SystemExit(f"❌ Другой экземпляр бота уже запущен (lock: {lock_path}). Останови его и попробуй снова.")


if __name__ == "__main__":
    run_app()
